"""
Import-only, by optimisation_model.

Summary cover sheet + figures for the optimisation results workbook.

Workbook sheets (in order):
    Cover            key inputs & assumptions (economics, price bands, capex, grid limits, run metadata)
    Status           solver status by district x activity, one small-multiple panel per heating
                     system, with each cell's per-building grid-headroom margin as cell text;
                     feasible cells shaded light -> dark green by NPV savings
    Top 10 NPV       bar chart — the 10 (district, activity, heating) scenarios with the highest NPV savings
    Rankings         mean NPV savings by activity and by district (marginal headlines)
    NPV Heatmap      district × activity grid, coloured by best-achievable NPV savings
    Self-supply      self-consumption / self-sufficiency vs recommended PV, by class + roof-area utilisation build-up 
    Scenario Ranking the full ranked table (data backing)
"""

import os
import sys

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")                       # headless: no GUI backend
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from matplotlib.colors import (LinearSegmentedColormap, ListedColormap, Normalize,
                               TwoSlopeNorm, to_rgba)
from matplotlib.cm import ScalarMappable
from matplotlib.ticker import FuncFormatter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill

import model_params as mp
import demand_profile_model as dm
from demand_profile_model import DISTRICT_STATIONS

# Display orders + short labels.
DIST_ORDER  = list(DISTRICT_STATIONS.keys())
ACT_SHORT   = {"Health: Health centre": "Health centre", "Health: Hospital": "Hospital",
               "Office: A/C standard": "Office (A/C)", "Retail: Department store": "Dept store"}
ACT_ORDER   = [ACT_SHORT.get(a, a) for a in mp.ROOF_PROPERTIES]
HEAT_ORDER  = ["Gas Boiler", "ASHP", "GSHP (vertical)", "GSHP (horizontal)"]
HEAT_SHORT  = {"Gas Boiler": "Gas", "ASHP": "ASHP",
               "GSHP (vertical)": "GSHP-V", "GSHP (horizontal)": "GSHP-H"}
HEAT_COLORS = {"Gas Boiler": "#9e9e9e", "ASHP": "#1f77b4",
               "GSHP (vertical)": "#2ca02c", "GSHP (horizontal)": "#17becf"}
ACT_COLORS  = {"Health centre": "#4c72b0", "Hospital": "#c44e52",
               "Office (A/C)": "#dd8452", "Dept store": "#55a868"}

# Both cost rounds write into the same run's charts/ folder, so every figure filename carries the
# round as a prefix ("deterministic_npv_top10.png", "stochastic_npv_top10.png"). Without it the
# second round silently overwrote the first and charts/ only ever held one round's figures.
# Set by write_report() for the duration of one workbook build; "" means unprefixed.
_CHART_PREFIX = ""

# Battery recommendations below this are numerical dribble from the LP rather than a design
# decision, and are plotted as "no storage" (see _fig_self_supply).
_BATT_FLOOR_KWH = 1.0

_HDR_FILL = PatternFill("solid", fgColor="DDE7F0")
_TITLE_FONT, _SECT_FONT, _BOLD = Font(bold=True, size=14), Font(bold=True, size=11), Font(bold=True)


def _act_area_map() -> dict:
    # {act_short: BEES median premises floor area m²}. Used to append the floor area to activity-class labels on categorical axes.
    areas = getattr(dm, "bees_floor_areas", None) or {}
    return {ACT_SHORT.get(a, a): v for a, v in areas.items()}


def _act_label(act_short: str, area_map: dict) -> str:
    # Append the building floor area to an activity-class label: "Hospital" -> "Hospital (4,150 m²)".
    a = area_map.get(act_short)
    return f"{act_short} ({a:,.0f} m²)" if a else act_short


def _act_labels(area_map: dict) -> list:
    # ACT_ORDER short labels with area appended — for categorical activity axes.
    return [_act_label(a, area_map) for a in ACT_ORDER]


def _gbp(x: float) -> str:
    # Compact GBP label.
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "—"
    a = abs(x)
    if a >= 1e6: return f"£{x/1e6:.2f}m"
    if a >= 1e3: return f"£{x/1e3:.0f}k"
    return f"£{x:.0f}"


def _pct(x: float) -> str:
    return f"{x*100:+.1f}%"


def _tco2e(x: float) -> str:
    # Compact tCO2e label.
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "—"
    return f"{x/1e3:.1f}k t" if abs(x) >= 1e3 else f"{x:.0f} t"


def _dual_text(value_fmt, v, cost=None, carbon=None) -> str:
    # Data-callout text for one bar/cell: "£X, Y t".
    if cost is not None and carbon is not None:
        return f"{_gbp(cost)}, {_tco2e(carbon)}"
    return value_fmt(v)


def _spec_line(r) -> str:
    # Compact technical-performance summary for one design: PV size + capacity factor (site/weather dependent) + battery & thermal-store capacity.
    pv   = getattr(r, "pv_kwp", 0.0) or 0.0
    gen  = getattr(r, "annual_pv_gen_kwh", 0.0) or 0.0
    cf   = (gen / (pv * 8760.0)) if pv > 0 else 0.0
    batt = getattr(r, "e_batt_kwh", 0.0) or 0.0
    tst  = getattr(r, "e_th_kwh", 0.0) or 0.0
    pv_txt = f"PV {pv:.0f}kWp (CF {cf:.0%})" if pv > 0 else "PV 0kWp"
    return f"{pv_txt} · Batt {batt:.0f}kWh · TST {tst:.0f}kWh"


def _spec_line_wrapped(r) -> str:
    pv   = getattr(r, "pv_kwp", 0.0) or 0.0
    gen  = getattr(r, "annual_pv_gen_kwh", 0.0) or 0.0
    cf   = (gen / (pv * 8760.0)) if pv > 0 else 0.0
    batt = getattr(r, "e_batt_kwh", 0.0) or 0.0
    tst  = getattr(r, "e_th_kwh", 0.0) or 0.0
    pv_txt = f"PV {pv:.0f}kWp (CF {cf:.0%})" if pv > 0 else "PV 0kWp"
    return f"{pv_txt}\nBatt {batt:.0f}kWh  TST {tst:.0f}kWh"


# FIGURES 
def _fig_top10(dfo: pd.DataFrame, charts_dir: str, *, value_col: str, value_fmt,
               value_label: str, title: str, fname: str,
               cost_col: str = None, carbon_col: str = None) -> str:
    top = dfo.nlargest(10, value_col).iloc[::-1]      # reverse: best at top of barh
    AM = _act_area_map()
    labels = [f"{r.district} · {_act_label(r.act_short, AM)}\n{_spec_line(r)}" for r in top.itertuples()]
    vals   = top[value_col].to_numpy()
    colors = [HEAT_COLORS[h] for h in top["heating"]]
    cost_vals   = top[cost_col].to_numpy()   if cost_col   else None
    carbon_vals = top[carbon_col].to_numpy() if carbon_col else None

    fig, ax = plt.subplots(figsize=(10, 6.8))
    y = np.arange(len(top))
    ax.barh(y, vals, color=colors)
    ax.set_yticks(y); ax.set_yticklabels(labels, fontsize=8, linespacing=1.6)
    ax.axvline(0, color="k", lw=0.6)
    ax.set_xlabel(value_label)
    ax.set_title(title, fontweight="bold")
    ax.xaxis.set_major_formatter(FuncFormatter(lambda v, _: value_fmt(v)))
    pad = 0.01 * (np.nanmax(vals) - min(0, np.nanmin(vals)) + 1)
    for i, v in enumerate(vals):
        txt = _dual_text(value_fmt, v, cost_vals[i] if cost_vals is not None else None,
                          carbon_vals[i] if carbon_vals is not None else None)
        ax.text(v + (pad if v >= 0 else -pad), i, txt,
                va="center", ha="left" if v >= 0 else "right", fontsize=8)
    ax.legend(handles=[Patch(color=HEAT_COLORS[h], label=HEAT_SHORT[h]) for h in HEAT_ORDER],
              title="Heating", fontsize=8, title_fontsize=8, loc="lower right")
    fig.tight_layout()
    return _save(fig, charts_dir, fname)


def _fig_marginal(dfo: pd.DataFrame, charts_dir: str, *, value_col: str, value_fmt,
                  value_label: str, fname: str,
                  cost_col: str = None, carbon_col: str = None) -> str:
    # Headline ranking = best ACHIEVABLE value per activity / per district.
    def _best_per(group_col, order):
        idx = dfo.groupby(group_col)[value_col].idxmax()
        return dfo.loc[idx.values].set_index(group_col).reindex(order).sort_values(value_col)

    by_act  = _best_per("act_short", ACT_ORDER)
    by_dist = _best_per("district", DIST_ORDER)

    AM = _act_area_map()
    fig, (a1, a2) = plt.subplots(1, 2, figsize=(15, 6.5))
    for ax, best, other_col, color, ttl, index_is_act in (
            (a1, by_act, "district", "#4c72b0", "by activity", True),
            (a2, by_dist, "act_short", "#55a868", "by district", False)):
        vals = best[value_col].to_numpy()
        y = np.arange(len(best))
        ax.barh(y, vals, color=color)
        ticklabels = []
        for name, r in zip(best.index.astype(str), best.itertuples()):
            name_lbl  = _act_label(name, AM) if index_is_act else name       # index is activity in a1
            other_val = str(getattr(r, other_col))
            other_lbl = other_val if index_is_act else _act_label(other_val, AM)   # other is activity in a2
            ticklabels.append(f"{name_lbl}\n{other_lbl} · {HEAT_SHORT[r.heating]}\n{_spec_line(r)}")
        ax.set_yticks(y, ticklabels, fontsize=7, linespacing=1.5)
        ax.axvline(0, color="k", lw=0.6)
        ax.set_xlabel(value_label + " (best across heating systems)")
        ax.set_title(ttl, fontweight="bold")
        ax.xaxis.set_major_formatter(FuncFormatter(lambda v, _: value_fmt(v)))
        cost_vals   = best[cost_col].to_numpy()   if cost_col   else None
        carbon_vals = best[carbon_col].to_numpy() if carbon_col else None
        for i, v in enumerate(vals):
            txt = _dual_text(value_fmt, v, cost_vals[i] if cost_vals is not None else None,
                              carbon_vals[i] if carbon_vals is not None else None)
            ax.text(v, i, "  " + txt, va="center",
                    ha="left" if v >= 0 else "right", fontsize=8)
    best = dfo.loc[dfo[value_col].idxmax()]
    best_txt = _dual_text(value_fmt, best[value_col],
                          best[cost_col] if cost_col else None,
                          best[carbon_col] if carbon_col else None)
    fig.suptitle(f"Best single scenario: {_act_label(best.act_short, AM)} · {best.district} · "
                 f"{HEAT_SHORT[best.heating]}  ({best_txt})", fontsize=10)
    fig.tight_layout(rect=(0, 0, 1, 0.96))
    return _save(fig, charts_dir, fname)


def _fig_heatmap(dfo: pd.DataFrame, charts_dir: str, *, value_col: str, value_fmt,
                 title: str, fname: str,
                 cost_col: str = None, carbon_col: str = None) -> str:
    # idxmax (not a plain pivot max) keeps the winning row per cell so it can be labelled with the
    # heating system + PV/battery/store spec that achieves the cell's best value.
    idx  = dfo.groupby(["district", "act_short"])[value_col].idxmax()
    best = dfo.loc[idx.values].set_index(["district", "act_short"])
    piv  = (dfo.pivot_table(index="district", columns="act_short",
                           values=value_col, aggfunc="max")
              .reindex(index=DIST_ORDER, columns=ACT_ORDER))
    data = piv.to_numpy(dtype=float)
    vmax = np.nanmax(np.abs(data)) or 1.0

    fig, ax = plt.subplots(figsize=(11, 9))
    im = ax.imshow(data, cmap="RdYlGn", vmin=-vmax, vmax=vmax, aspect="auto")
    ax.set_xticks(range(len(ACT_ORDER)), _act_labels(_act_area_map()))
    ax.set_yticks(range(len(DIST_ORDER)), DIST_ORDER)
    ax.set_title(title, fontweight="bold")
    for i, d in enumerate(DIST_ORDER):
        for j, a in enumerate(ACT_ORDER):
            v = data[i, j]
            if np.isnan(v):
                continue
            r = best.loc[(d, a)]
            val_txt = _dual_text(value_fmt, v, r[cost_col] if cost_col else None,
                                  r[carbon_col] if carbon_col else None)
            label = f"{val_txt} - {HEAT_SHORT[r.heating]}\n{_spec_line_wrapped(r)}"
            ax.text(j, i, label, ha="center", va="center", fontsize=6.5, color="black", linespacing=1.4)
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04,
                 format=FuncFormatter(lambda v, _: value_fmt(v)))
    fig.tight_layout()
    return _save(fig, charts_dir, fname)


# Solver status category colours.
_STATUS_COLORS = {
    "Optimal":                        "#4caf50",
    "No incumbent":                   "#9e9e9e",
    "Infeasible":                     "#e53935",
}
_STATUS_ORDER = list(_STATUS_COLORS)

# Feasible (Optimal) cells keep the green identity but are shaded light -> dark by NPV savings, so
# the grid reads on two levels: hue says feasible/infeasible, depth says how good the project is.
# Ramp starts at a pale (not white) green so the lightest feasible cell is still unmistakably green
# against the white "not solved" background.
_NPV_GREENS = LinearSegmentedColormap.from_list("feasible_npv", ["#e5f5e0", "#00441b"])
_NPV_SHADE_FLOOR = 0.12          # never fully pale: the worst feasible cell still reads as green


def _npv_shade(v, vmin: float, vmax: float) -> tuple:
    # Green shade for one feasible cell. Falls back to the flat status green when the cell has no
    # NPV (e.g. a status grid built from a results set without the column) or the range is degenerate.
    if v is None or vmin is None or vmax is None or not np.isfinite([v, vmin, vmax]).all():
        return to_rgba(_STATUS_COLORS["Optimal"])
    t = 0.5 if vmax <= vmin else (v - vmin) / (vmax - vmin)
    return _NPV_GREENS(_NPV_SHADE_FLOOR + (1.0 - _NPV_SHADE_FLOOR) * t)


def _text_color(rgba: tuple) -> str:
    # Black on pale cells, white on dark ones — the NPV ramp spans both ends of the contrast range.
    r, g, b = rgba[:3]
    return "black" if (0.299 * r + 0.587 * g + 0.114 * b) > 0.55 else "white"


def _edge_margin_label(r) -> str:
    # One cell's grid-connection headroom margin as text, on grid_sensitivity.py's signed scale:
    # positive = demand-growth margin today's DNO ceiling can still absorb (cell Optimal today);
    # negative = how much bigger the ceiling would need to be to reach feasibility (cell infeasible).
    # Cells whose bisection ran out of search range never measured an edge labelled as a "non-result".
    if not bool(r["search_bounded"]):
        return "unbounded"
    v = r["edge_margin_pct"]
    if pd.isna(v):
        return "N/A"
    mult = 1.0 + abs(v) / 100.0
    mult_txt = f"{mult:.1f}x" if mult < 10.0 else f"{mult:.0f}x"
    kw = r["edge_margin_kw"]
    kw_txt = f"{kw:+,.0f} kW" if pd.notna(kw) else "? kW"
    # One line, multiplier first with the absolute headroom in brackets — e.g. "41x (+11,084 kW)".
    # Same string in the chart cells and the Grid Sensitivity Data column, so the two read alike.
    return f"{mult_txt} ({kw_txt})"


def _fig_status_grid(df_raw: pd.DataFrame, charts_dir: str, *, title: str, fname: str,
                     grid_sensitivity: pd.DataFrame = None, npv_col: str = "npv_savings_GBP") -> str:
    # One district x activity panel per heating system. Cell text carries each cell's grid-headroom margin.
    # Heat-pump heatings only: Gas Boiler draws no heat-pump electricity so it has no margin row.
    # Feasible cells are additionally shaded by npv_col (light -> dark green = worse -> better NPV);
    # infeasible/no-incumbent cells keep their flat category colour.
    d = df_raw.copy()
    # Every flavour of infeasible reads as one category here: the qualified strings the engine can
    # still emit ("Infeasible (solver tolerance)", and "Infeasible (land)" on runs predating the
    # rename) say why, not whether, and splitting the legend on that made a heating type look like
    # two different outcomes. The raw status survives on the data sheets for diagnosis.
    d["status"] = d["status"].where(~d["status"].astype(str).str.startswith("Infeasible"),
                                    "Infeasible")
    d["act_short"] = d["activity"].map(lambda a: ACT_SHORT.get(a, a))
    other = [s for s in d["status"].unique() if s not in _STATUS_COLORS]
    status_colors = dict(_STATUS_COLORS)
    status_order = list(_STATUS_ORDER)
    for s in other:                             # tolerate any status not in the known palette
        status_colors[s] = "#6a1b9a"
        status_order.append(s)
    act_labels = _act_labels(_act_area_map())

    # NPV shading range is taken across every feasible cell in the figure (all four heating panels),
    # so a shade means the same thing in each panel and heating systems stay comparable by eye.
    has_npv = npv_col in d.columns
    npv_vals = pd.to_numeric(d.loc[d["status"] == "Optimal", npv_col], errors="coerce").dropna() \
        if has_npv else pd.Series(dtype=float)
    vmin = float(npv_vals.min()) if not npv_vals.empty else None
    vmax = float(npv_vals.max()) if not npv_vals.empty else None

    margins = {}
    if grid_sensitivity is not None and not grid_sensitivity.empty:
        g = grid_sensitivity.copy()
        g["act_short"] = g["activity"].map(lambda a: ACT_SHORT.get(a, a))
        for _, r in g.iterrows():
            margins[(r["heating"], r["district"], r["act_short"])] = _edge_margin_label(r)

    fig, axes = plt.subplots(2, 2, figsize=(15, 13))
    for ax, h in zip(axes.flat, HEAT_ORDER):
        sub = d[d["heating"] == h]
        piv = (sub.pivot_table(index="district", columns="act_short", values="status",
                               aggfunc="first")
                  .reindex(index=DIST_ORDER, columns=ACT_ORDER))
        npv_piv = (sub.pivot_table(index="district", columns="act_short", values=npv_col,
                                   aggfunc="max")
                      .reindex(index=DIST_ORDER, columns=ACT_ORDER)) if has_npv else None
        # Painted as an explicit RGBA image rather than a categorical colormap: feasible cells each
        # need their own shade, so there is no longer one colour per status to look up.
        rgba = np.ones((len(DIST_ORDER), len(ACT_ORDER), 4))     # white = cell absent from results
        for i in range(len(DIST_ORDER)):
            for j in range(len(ACT_ORDER)):
                s = piv.iat[i, j]
                if not isinstance(s, str):
                    continue
                rgba[i, j] = (_npv_shade(npv_piv.iat[i, j] if npv_piv is not None else None, vmin, vmax)
                              if s == "Optimal" else to_rgba(status_colors[s]))
        ax.imshow(rgba, aspect="auto")
        ax.set_xticks(range(len(ACT_ORDER)), act_labels, fontsize=8)
        ax.set_yticks(range(len(DIST_ORDER)), DIST_ORDER, fontsize=8)
        n_bad = (sub["status"] == "Infeasible").sum()
        ax.set_title(f"{h}  ({n_bad} infeasible)", fontweight="bold", fontsize=10)
        for i, dist in enumerate(DIST_ORDER):
            for j, act in enumerate(ACT_ORDER):
                s = piv.iat[i, j]
                if not isinstance(s, str):
                    continue
                npv_v = npv_piv.iat[i, j] if npv_piv is not None else None
                # Feasible cells lead with the NPV the shade encodes, so cells can also be ranked
                # by reading rather than by comparing shades.
                head = [_gbp(npv_v)] if (s == "Optimal" and npv_v is not None and np.isfinite(npv_v)) else \
                       ([] if s == "Optimal" else [s])
                # Margin only on cells that solved: the bisection never measures an edge for a cell
                # that is infeasible anyway, so printing its "unbounded" non-result on every one of
                # them just repeats noise over the whole panel.
                show_margin = s == "Optimal" and (h, dist, act) in margins
                parts = head + ([margins[(h, dist, act)]] if show_margin else [])
                if parts:
                    ax.text(j, i, "\n".join(parts), ha="center", va="center", fontsize=6.5,
                            color=_text_color(rgba[i, j]), linespacing=1.2, wrap=True)
    shaded = vmin is not None
    # Each line is emitted only when that content is actually on the chart, so the title never
    # promises cell text a --skip-grid-sensitivity run did not produce.
    notes = []
    if shaded:
        notes.append("Cell text (Line 1): NPV savings vs gas boiler BAU")
    # These multipliers come from a fixed-sizing bisection (baseline PV, no battery, baseline thermal store),
    # because a free re-solve lets the optimiser buy storage purely to hide the peak.
    # They are also per-building: today's spare headroom allocated entirely to this one site.
    if margins:
        notes.append("Cell text (Line 2): demand growth each building can add before the DNO "
                     "import ceiling binds")
    if notes:
        title += "\n" + "\n".join(notes)
    fig.suptitle(title, fontweight="bold", fontsize=13)
    # With the gradient active, one flat green patch would misrepresent the feasible cells, so the
    # "Optimal" entry moves out of the categorical legend and onto the colourbar below it.
    handles = [Patch(color=status_colors[s], label=s) for s in status_order
               if not (shaded and s == "Optimal")]
    fig.legend(handles=handles, title="Solver status", fontsize=9, title_fontsize=9,
               loc="upper left", bbox_to_anchor=(1.0, 0.95))
    fig.tight_layout(rect=(0, 0, 0.98, 0.95))    # before add_axes: tight_layout ignores manual axes
    if shaded:
        # Colourbar spans the same floor-to-1 slice of the ramp the cells are painted from, so the
        # bar's endpoints are the actual lightest and darkest cell colours in the figure.
        sm = ScalarMappable(norm=Normalize(vmin=vmin, vmax=vmax),
                            cmap=LinearSegmentedColormap.from_list(
                                "feasible_npv_bar",
                                _NPV_GREENS(np.linspace(_NPV_SHADE_FLOOR, 1.0, 256))))
        cax = fig.add_axes((1.005, 0.35, 0.012, 0.32))
        cb  = fig.colorbar(sm, cax=cax, format=FuncFormatter(lambda v, _: _gbp(v)))
        cb.set_label("Optimal (feasible) — NPV savings vs gas-boiler BAU", fontsize=9)
        cb.ax.tick_params(labelsize=8)
    return _save(fig, charts_dir, fname)


_STATUS_TITLE = "Solver status by district × activity × heating"


def _status_figs(dfs: dict, charts_dir: str, grid_sensitivity) -> list:
    # Feasibility does not depend on objective; cost and emissions status grids typically identical.
    key  = ["district", "activity", "heating"]
    npv, carb = dfs["NPV"], dfs["Carbon"]
    merged = npv[key + ["status"]].merge(carb[key + ["status"]], on=key, suffixes=("_n", "_c"))
    same = len(merged) and (merged["status_n"] == merged["status_c"]).all()

    if same:
        figs = [_fig_status_grid(npv, charts_dir, title=_STATUS_TITLE, fname="status_grid.png",
                                 grid_sensitivity=grid_sensitivity)]
    else:
        figs = [_fig_status_grid(npv, charts_dir, title=_STATUS_TITLE + " (cost objective)",
                                 fname="status_grid_cost.png", grid_sensitivity=grid_sensitivity),
                _fig_status_grid(carb, charts_dir, title=_STATUS_TITLE + " (emissions objective)",
                                 fname="status_grid_carbon.png", grid_sensitivity=grid_sensitivity)]
    return figs


_MAC_TITLE = "Marginal Abatement Cost by district × activity × heating"


def _model_carbon_value(d: pd.DataFrame) -> float:
    # The model's own discounted average carbon value (£/tCO2e), recovered from the results rather
    # than re-derived from CARBON_VALUES, so it tracks whatever trajectory the run actually used:
    # carbon_value_npv_GBP is that trajectory applied to the design's own lifetime emissions.
    need = {"carbon_value_npv_GBP", "lifetime_emissions_tco2e"}
    if not need <= set(d.columns):
        return float("nan")
    num = pd.to_numeric(d["carbon_value_npv_GBP"], errors="coerce")
    den = pd.to_numeric(d["lifetime_emissions_tco2e"], errors="coerce")
    return float(np.nanmedian((num / den.where(den > 0)).to_numpy(dtype=float)))


def _fig_mac_grid(dfo: pd.DataFrame, charts_dir: str, *, fname: str = "mac_grid.png") -> str:
    # Built from the COST sweep only. MAC divides a design's extra cost over BAU by the tonnes it
    # abates, so it is only a *cost* of abatement when the design was chosen under cost pressure.
    # The emissions sweep's equivalent column is cost-blind by construction and is deliberately not
    # the source here (it ships as cost_of_emissions_optimal_design_GBP_per_tco2e instead).
    col = "mac_GBP_per_tco2e"
    if col not in dfo.columns:
        return None
    d = dfo.copy()
    d[col] = pd.to_numeric(d[col], errors="coerce").replace([np.inf, -np.inf], np.nan)
    vals = d[col].dropna()
    if vals.empty:
        return None

    cval   = _model_carbon_value(d)
    centre = cval if np.isfinite(cval) else float(vals.median())
    # The ramp floors at £0 rather than at the most negative MAC. A design that abates AND saves
    # money is already a no-regret call, so how far below zero it sits carries no extra decision
    # weight — while letting it set the bound (gas+PV reaches about -£3,200/t) would compress the
    # entire £0-£500 band where the heat pumps actually separate into one indistinguishable shade.
    # Sub-zero cells clip to the green end; the printed value still shows the true figure.
    lo = 0.0 if float(vals.min()) < 0.0 else float(np.nanpercentile(vals, 5))
    lo = min(lo, centre - 1.0)
    # Percentile bound, not max: one runaway cell would otherwise flatten every other colour.
    hi = max(float(np.nanpercentile(vals, 95)), centre + 1.0)
    norm = TwoSlopeNorm(vmin=lo, vcenter=centre, vmax=hi)
    cmap = plt.get_cmap("RdYlGn_r")          # reversed: cheap abatement green, dear abatement red

    fig, axes = plt.subplots(2, 2, figsize=(15, 13))
    for ax, h in zip(axes.flat, HEAT_ORDER):
        sub = d[d["heating"] == h]
        piv = (sub.pivot_table(index="district", columns="act_short", values=col, aggfunc="min")
                  .reindex(index=DIST_ORDER, columns=ACT_ORDER))
        data = piv.to_numpy(dtype=float)
        rgba = np.ones((len(DIST_ORDER), len(ACT_ORDER), 4))    # white = no optimal design here
        for i in range(len(DIST_ORDER)):
            for j in range(len(ACT_ORDER)):
                v = data[i, j]
                if np.isfinite(v):
                    rgba[i, j] = cmap(float(np.clip(norm(v), 0.0, 1.0)))
        ax.imshow(rgba, aspect="auto")
        ax.set_xticks(range(len(ACT_ORDER)), _act_labels(_act_area_map()), fontsize=8)
        ax.set_yticks(range(len(DIST_ORDER)), DIST_ORDER, fontsize=8)
        n_cells = int(np.isfinite(data).sum())
        if not n_cells:
            ax.set_title(f"{h}  (no optimal design)", fontweight="bold", fontsize=10)
        elif np.isfinite(cval):
            n_below = int((data < cval).sum())
            ax.set_title(f"{h}  ({n_below}/{n_cells} below the carbon value)",
                         fontweight="bold", fontsize=10)
        else:
            ax.set_title(h, fontweight="bold", fontsize=10)
        for i in range(len(DIST_ORDER)):
            for j in range(len(ACT_ORDER)):
                v = data[i, j]
                if np.isfinite(v):
                    ax.text(j, i, f"£{v:,.0f}/t", ha="center", va="center", fontsize=7,
                            color=_text_color(rgba[i, j]))

    # Title only. What the colour split means (the carbon value), and why a negative MAC pays for
    # itself, are covered in the write-up rather than repeated on the figure — the per-panel counts
    # and the colourbar still carry the reference on the chart itself.
    fig.suptitle(_MAC_TITLE, fontweight="bold", fontsize=13)
    # Headroom sized for the one-line title. The 0.93 the status grid uses would leave a band of
    # empty figure under it, since that grid carries three lines of title and this one carries one.
    fig.tight_layout(rect=(0, 0, 0.98, 0.965))
    cax = fig.add_axes((1.005, 0.35, 0.012, 0.32))
    cb  = fig.colorbar(ScalarMappable(norm=norm, cmap=cmap), cax=cax,
                       format=FuncFormatter(lambda v, _: f"£{v:,.0f}"))
    cb.set_label("Marginal abatement cost (£/tCO2e) — scale floors at £0", fontsize=9)
    cb.ax.tick_params(labelsize=8)
    if np.isfinite(cval) and lo <= cval <= hi:
        cb.ax.axhline(cval, color="black", linewidth=1.2)
    return _save(fig, charts_dir, fname)


def _fig_robustness(dfo: pd.DataFrame, charts_dir: str) -> str:
    # Robustness of top NPV designs under the import-price scenario set.
    need = {"npv_savings_min_GBP", "npv_savings_max_GBP"}
    if not need.issubset(dfo.columns):
        return None                                    # deterministic run (single scenario) — skip
    top = dfo.nlargest(12, "npv_savings_GBP").iloc[::-1]
    AM = _act_area_map()
    labels = [f"{r.district} · {_act_label(r.act_short, AM)}" for r in top.itertuples()]
    exp = top["npv_savings_GBP"].to_numpy()
    lo  = top["npv_savings_min_GBP"].to_numpy()
    hi  = top["npv_savings_max_GBP"].to_numpy()
    colors = [HEAT_COLORS[h] for h in top["heating"]]

    fig, ax = plt.subplots(figsize=(10, 6.5))
    y = np.arange(len(top))
    ax.barh(y, exp, color=colors, zorder=2)
    ax.errorbar(exp, y, xerr=[exp - lo, hi - exp], fmt="none", ecolor="0.25",
                elinewidth=1.1, capsize=3, zorder=3)
    ax.set_yticks(y); ax.set_yticklabels(labels, fontsize=9)
    ax.axvline(0, color="k", lw=0.6)
    ax.set_xlabel("15-year NPV savings vs BAU — expected (bar), worst..best case (whisker)")
    ax.set_title("Robustness of top designs under electricity-price uncertainty", fontweight="bold")
    ax.xaxis.set_major_formatter(FuncFormatter(lambda v, _: _gbp(v)))
    ax.legend(handles=[Patch(color=HEAT_COLORS[h], label=HEAT_SHORT[h]) for h in HEAT_ORDER],
              title="Heating", fontsize=8, title_fontsize=8, loc="lower right")
    fig.tight_layout()
    return _save(fig, charts_dir, "npv_robustness.png")


# SELF-CONSUMPTION / SELF-SUFFICIENCY vs RECOMMENDED SIZING
def _fig_self_supply(dfo: pd.DataFrame, charts_dir: str) -> str:
    # How NPV-recommended equipment sizing drives the building's self-consumption rate (share of PV 
    # used on site) and self-sufficiency rate (share of demand met on site), across activity classes. 
    need = {"self_consumption_rate", "self_sufficiency_rate", "pv_kwp", "e_batt_kwh", "npv_savings_GBP"}
    if not need.issubset(dfo.columns):
        return None
    idx  = dfo.groupby(["district", "act_short"])["npv_savings_GBP"].idxmax()
    best = dfo.loc[idx].copy()
    best = best[best["pv_kwp"] > 0]
    if best.empty:
        return None
    bmax = float(best["e_batt_kwh"].max())
    # Sub-kWh recommendations are LP dribble, not a design decision — floor them to "no storage"
    # so a ~0.3 kWh cell cannot claim the whole marker-area scale.
    if not np.isfinite(bmax) or bmax < _BATT_FLOOR_KWH:
        bmax = 0.0                                          # no battery recommended anywhere
    S_MIN, S_SPAN = 70.0, 290.0                             # marker area for 0 kWh, and growth to bmax

    def _batt_area(kwh):                                    # marker area ∝ battery kWh (70..360)
        if bmax <= 0:
            return S_MIN                                    # every design at 0 kWh -> all markers equal
        return S_MIN + S_SPAN * (np.asarray(kwh, dtype=float) / bmax)

    fig, (a1, a2) = plt.subplots(1, 2, figsize=(15, 6), sharex=True)
    for ax, col, ttl in ((a1, "self_consumption_rate", "Self-consumption rate"),
                         (a2, "self_sufficiency_rate", "Self-sufficiency rate")):
        for cls in ACT_ORDER:
            sub = best[best["act_short"] == cls]
            if sub.empty:
                continue
            ax.scatter(sub["pv_kwp"], sub[col] * 100.0, s=_batt_area(sub["e_batt_kwh"].to_numpy()),
                       color=ACT_COLORS.get(cls, "#777777"), alpha=0.72,
                       edgecolors="k", linewidths=0.4, label=cls, zorder=2)
        ax.set_xlabel("Recommended PV size (kWp)")
        ax.set_ylabel(ttl + " (%)")
        ax.set_title(ttl, fontweight="bold")
        ax.grid(lw=0.3, alpha=0.5)
        ax.set_ylim(0, 100)
    AM = _act_area_map()
    act_handles = [a1.scatter([], [], s=S_MIN, color=ACT_COLORS.get(cls, "#777777"),
                               alpha=0.72, edgecolors="k", linewidths=0.4, label=_act_label(cls, AM))
                   for cls in ACT_ORDER if not best[best["act_short"] == cls].empty]
    # Legends sit to the right of both panels.
    fig.legend(handles=act_handles, title="Activity class", fontsize=8, title_fontsize=8,
               loc="upper left", bbox_to_anchor=(0.845, 0.88))
    # Battery-area reference legend. Always drawn — a lone "0 kWh" entry is the finding that no
    # design in the sweep recommended storage (above _BATT_FLOOR_KWH), not a missing series.
    # Ticks are unrounded so each handle's area is the one the scale actually assigns to its label.
    batt_kwh = [0.0] if bmax <= 0 else sorted({0.0, bmax / 2, bmax})
    batt_handles = [a2.scatter([], [], s=_batt_area(kwh), color="0.6",
                                edgecolors="k", linewidths=0.4, label=f"{kwh:.3g} kWh")
                     for kwh in batt_kwh]
    fig.legend(handles=batt_handles, title="Battery size (marker area)", fontsize=8,
               title_fontsize=8, loc="upper left", bbox_to_anchor=(0.845, 0.66),
               labelspacing=1.1, borderpad=0.9)
    fig.suptitle("Self-consumption and self-sufficiency of NPV-optimal designs against "
                 "PV (kWp) and battery (kWh)", fontsize=11, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 0.83, 0.95))
    return _save(fig, charts_dir, "self_supply.png")


def _roof_utilisation_rows() -> list:
    # Effective usable fraction of the roof footprint available for PV, split by roof type. 
    # Both types share the same per-activity usable-area fraction. They differ only in the roof-type factor:
    #   flat    = usable-area fraction × inter-row spacing (self-shading between tilted module rows);
    #   pitched = usable-area fraction × sec(pitch) slope-area gain × orientation/margin fraction. Pitched modules mount flush → no inter-row derate.
    areas     = getattr(dm, "bees_floor_areas", None) or {}
    pitch_geo = (1.0 / np.cos(np.radians(mp.ROOF_PITCH_DEG))) * mp.PITCHED_USABLE_SLOPE_FRAC
    rows = []
    for a in mp.ROOF_PROPERTIES:
        u = mp.ROOF_PROPERTIES[a]["pv_usable_frac"]
        r = mp.ROOF_PROPERTIES[a]["pv_inter_row_frac"]
        rows.append({"activity": ACT_SHORT.get(a, a), "activity_full": a,
                     "floor_area_m2": areas.get(a),
                     "utilisation_rate": u, "inter_row_spacing_factor": r,
                     "effective_utilisation_flat": u * r,
                     "effective_utilisation_pitched": u * pitch_geo})
    return rows


def _roof_utilisation_table() -> pd.DataFrame:
    df = pd.DataFrame(_roof_utilisation_rows())
    for c in ("utilisation_rate", "inter_row_spacing_factor",
              "effective_utilisation_flat", "effective_utilisation_pitched"):
        df[c] = df[c].round(4)
    return df.drop(columns=["activity_full"])


def _fig_roof_utilisation(charts_dir: str) -> str:
    # Roof-area utilisation (bar chart + companion table).
    rows   = _roof_utilisation_rows()
    AM     = _act_area_map()
    labels = [_act_label(r["activity"], AM) for r in rows]
    util   = np.array([r["utilisation_rate"] for r in rows])
    row    = np.array([r["inter_row_spacing_factor"] for r in rows])
    eff_f  = np.array([r["effective_utilisation_flat"] for r in rows])
    eff_p  = np.array([r["effective_utilisation_pitched"] for r in rows])

    fig, (ax, axt) = plt.subplots(2, 1, figsize=(11.5, 8.5),
                                  gridspec_kw={"height_ratios": [3, 1]})
    x = np.arange(len(rows)); w = 0.2
    series = [("Utilisation rate (usable area)", util,  "#4c72b0"),
              ("Inter-row spacing factor",        row,   "#dd8452"),
              ("Effective (flat roof)",           eff_f, "#55a868"),
              ("Effective (pitched roof)",        eff_p, "#8172b3")]
    for k, (lbl, vals, col) in enumerate(series):
        ax.bar(x + (k - 1.5) * w, vals, w, color=col, label=lbl, zorder=2)
        for xi, v in zip(x + (k - 1.5) * w, vals):
            ax.text(xi, v + 0.012, f"{v * 100:.1f}%", ha="center", va="bottom", fontsize=7.5)
    ax.set_xticks(x, labels)
    ax.set_ylim(0, 1.0)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{v * 100:.0f}%"))
    ax.set_ylabel("Fraction of roof footprint")
    ax.set_title("PV roof-area utilisation by activity class", fontweight="bold")
    ax.grid(axis="y", lw=0.3, alpha=0.5)
    ax.legend(fontsize=8, loc="upper left", ncol=2)

    axt.axis("off")
    cell = [[f"{util[i] * 100:.1f}%", f"{row[i] * 100:.0f}%",
             f"{eff_f[i] * 100:.1f}%", f"{eff_p[i] * 100:.1f}%"]
            for i in range(len(rows))]
    tbl = axt.table(cellText=cell,
                    rowLabels=[r["activity"] for r in rows],
                    colLabels=["Utilisation rate", "Inter-row spacing",
                               "Effective (flat roof)", "Effective (pitched roof)"],
                    cellLoc="center", rowLoc="left", loc="center")
    tbl.auto_set_font_size(False); tbl.set_fontsize(9); tbl.scale(1, 1.6)
    for (i, _), c in tbl.get_celld().items():
        if i == 0:
            c.set_text_props(fontweight="bold"); c.set_facecolor("#DDE7F0")
    fig.tight_layout()
    return _save(fig, charts_dir, "roof_utilisation.png")


def _rq1_table(npv: pd.DataFrame) -> pd.DataFrame:
    # NPV-optimal (recommended) design per district × class, with sizing and self-supply rates. 
    cols = ["district", "act_short", "heating", "pv_kwp", "e_batt_kwh", "annual_demand_kwh",
            "annual_pv_gen_kwh", "annual_export_kwh", "self_consumption_rate",
            "self_sufficiency_rate", "npv_savings_GBP"]
    have = [c for c in cols if c in npv.columns]
    idx  = npv.groupby(["district", "act_short"])["npv_savings_GBP"].idxmax()
    t = npv.loc[idx, have].sort_values(["act_short", "district"]).reset_index(drop=True)
    return t.rename(columns={"act_short": "activity"})


# PARETO FIGURES

def _fig_pareto_scatter(dfp: pd.DataFrame, charts_dir: str) -> str:
    # Every candidate design in (lifetime emissions, NPV savings) space, coloured by heating,
    # sized by total capex (PV/battery/thermal-store + heat-plant). Non-dominated designs are
    # ring-marked, annotated with a compact spec line, and joined by a step line (the frontier).
    total_capex = (dfp["capex_GBP"].fillna(0.0) + dfp["heat_capex_GBP"].fillna(0.0))
    cmin, cmax = total_capex.min(), total_capex.max()
    def _bubble_size(cap: pd.Series) -> pd.Series:
        if cmax <= cmin:
            return pd.Series(60.0, index=cap.index)
        return 15.0 + 185.0 * (cap - cmin) / (cmax - cmin)

    fig, ax = plt.subplots(figsize=(12.5, 7))
    heat_present = [h for h in HEAT_ORDER if not dfp[dfp["heating"] == h].empty]
    for h in heat_present:
        sub = dfp[dfp["heating"] == h]
        cap = sub["capex_GBP"].fillna(0.0) + sub["heat_capex_GBP"].fillna(0.0)
        ax.scatter(sub["lifetime_emissions_tco2e"], sub["npv_savings_GBP"],
                   s=_bubble_size(cap), c=HEAT_COLORS[h], alpha=0.55, zorder=2)
    pf = dfp[dfp["pareto_optimal"]].sort_values("lifetime_emissions_tco2e")
    ax.scatter(pf["lifetime_emissions_tco2e"], pf["npv_savings_GBP"],
               facecolors="none", edgecolors="k", s=90, linewidths=1.2, zorder=3)
    ax.axhline(0, color="k", lw=0.6)                       # BAU NPV reference
    ax.set_xlabel("Lifetime emissions (tCO₂e, lower = better →← )")
    ax.set_ylabel("15-year NPV savings vs BAU")
    ax.set_title("Cost-carbon performance across all candidate designs", fontweight="bold")
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: _gbp(v)))
    ax.xaxis.set_major_formatter(FuncFormatter(lambda v, _: _tco2e(v)))
    # Both legends sit to the right of the axes. 
    heat_handles = [ax.scatter([], [], s=70.0, c=HEAT_COLORS[h], alpha=0.55, label=HEAT_SHORT[h])
                    for h in heat_present]
    heat_handles.append(ax.scatter([], [], facecolors="none", edgecolors="k", s=70.0,
                                   linewidths=1.2, label="Pareto-optimal"))
    ax.legend(handles=heat_handles, title="Heating", fontsize=8, title_fontsize=8,
              loc="upper left", bbox_to_anchor=(1.02, 1.0))
    ref = pd.Series([100e3, 750e3, 3.4e6])
    cap_handles = [ax.scatter([], [], s=float(_bubble_size(ref).iloc[i]), c="0.6", alpha=0.55,
                              label=_gbp(float(ref.iloc[i]))) for i in range(len(ref))]
    fig.legend(handles=cap_handles, title="Total capex (bubble area)", fontsize=8, title_fontsize=8,
               loc="upper left", bbox_to_anchor=(0.795, 0.62), labelspacing=1.1, borderpad=0.9)
    fig.tight_layout(rect=(0, 0, 0.78, 1))
    return _save(fig, charts_dir, "pareto_scatter.png")


def _fig_pareto_heating(dfp: pd.DataFrame, charts_dir: str) -> str:
    # Non-dominated design with the lowest absolute MAC (£/tCO₂e).
    pf = dfp[dfp["pareto_optimal"]].copy()
    pf["abs_mac"] = pf["mac_GBP_per_tco2e"].abs().fillna(np.inf)
    idx  = pf.groupby(["district", "act_short"])["abs_mac"].idxmin()
    best = pf.loc[idx].set_index(["district", "act_short"])
    piv  = (best.reset_index().pivot(index="district", columns="act_short", values="heating")
                .reindex(index=DIST_ORDER, columns=ACT_ORDER))
    code = {h: i for i, h in enumerate(HEAT_ORDER)}
    codes = np.full(piv.shape, np.nan)
    for i in range(piv.shape[0]):
        for j in range(piv.shape[1]):
            h = piv.iat[i, j]
            if isinstance(h, str):
                codes[i, j] = code[h]

    fig, ax = plt.subplots(figsize=(11, 9))
    ax.imshow(codes, cmap=ListedColormap([HEAT_COLORS[h] for h in HEAT_ORDER]),
              vmin=-0.5, vmax=len(HEAT_ORDER) - 0.5, aspect="auto")
    ax.set_xticks(range(len(ACT_ORDER)), _act_labels(_act_area_map()))
    ax.set_yticks(range(len(DIST_ORDER)), DIST_ORDER)
    ax.set_title("Best-value Pareto heating (lowest |MAC|)\nby district × activity", fontweight="bold")
    for i, d in enumerate(DIST_ORDER):
        for j, a in enumerate(ACT_ORDER):
            h = piv.iat[i, j]
            if isinstance(h, str):
                r = best.loc[(d, a)]
                label = f"£{r.mac_GBP_per_tco2e:.0f}/t - {HEAT_SHORT[h]}\n{_spec_line_wrapped(r)}"
                ax.text(j, i, label, ha="center", va="center", fontsize=6.5, color="white",
                        linespacing=1.4)
    ax.legend(handles=[Patch(color=HEAT_COLORS[h], label=h) for h in HEAT_ORDER],
              title="Heating", fontsize=8, title_fontsize=8,
              loc="upper left", bbox_to_anchor=(1.02, 1.0))
    fig.tight_layout()
    return _save(fig, charts_dir, "pareto_heating.png")


def _save(fig, charts_dir: str, name: str) -> str:
    path = os.path.join(charts_dir, _CHART_PREFIX + name)
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return path


def _stack_images(ws, pngs: list, start_row: int = 2, col: str = "B", pad_rows: int = 3) -> None:
    # Stack chart PNGs down one sheet, spacing each by its own pixel height (~20px per Excel row).
    r = start_row
    for png in pngs:
        img = XLImage(png)
        ws.add_image(img, f"{col}{r}")
        r += int(img.height / 20) + pad_rows


# COVER SHEET

def _build_cover(book, run_meta: dict) -> None:
    ws = book.create_sheet("Cover", 0)
    for col, w in zip("ABCD", (34, 20, 16, 16)):
        ws.column_dimensions[col].width = w
    r = [1]
    def _row(): v = r[0]; r[0] += 1; return v
    def title(t):
        c = ws.cell(_row(), 1, t); c.font = _TITLE_FONT
    def section(t):
        v = _row(); c = ws.cell(v, 1, t); c.font = _SECT_FONT
        for col in range(1, 5): ws.cell(v, col).fill = _HDR_FILL
    def kv(k, val):
        v = _row(); ws.cell(v, 1, k); ws.cell(v, 2, val)
    def blank(): _row()
    def table(headers, rows):
        v = _row()
        for j, h in enumerate(headers, 1): c = ws.cell(v, j, h); c.font = _BOLD
        for row in rows:
            v = _row()
            for j, val in enumerate(row, 1): ws.cell(v, j, val)

    tc = mp.TECH_COSTS
    title("Building PV · Battery · Heat — Optimisation Results")
    kv("Run", run_meta.get("timestamp", ""))
    kv("Scenarios", f"{run_meta.get('n_feasible','?')} / {run_meta.get('n_scenarios','?')} feasible")
    # No MIP gap quoted: every sizing variable is continuous and the dispatch mutex binaries are
    # off by default, so each cell is a pure LP solved to optimality — a relative gap would only
    # understate the precision of the result.
    kv("Solver", f"{run_meta.get('solver','HiGHS')}")
    blank()

    section("Economics (real / constant-2025 prices)")
    kv("Discount rate (WACC, real)", _pct(tc["discount_rate"]))
    kv("Horizon", f"{run_meta.get('horizon_years','?')} years")
    kv("General inflation (real O&M)", _pct(tc["general_inflation"]))
    kv("Electricity price growth (real)", _pct(tc["elec_price_growth"]) + "/yr")
    kv("Gas price growth (real)", _pct(tc["gas_price_growth"]) + "/yr")
    kv("Export price (SEG)", f"£{tc['elec_export_price']:.2f}/kWh")
    blank()

    section("Active scenarios")
    kv("PV capex", f"{mp.PV_CAPEX_ACTIVE} (£{mp.PV_CAPEX_SCENARIOS[mp.PV_CAPEX_ACTIVE]:.0f}/kWp)")
    kv("Export price", f"{mp.ELEC_EXPORT_ACTIVE} (£{mp.ELEC_EXPORT_SCENARIOS[mp.ELEC_EXPORT_ACTIVE]:.2f}/kWh)")
    kv("Energy price growth", mp.ENERGY_GROWTH_ACTIVE)
    blank()

    price_scen = run_meta.get("price_scenarios")
    if price_scen:
        section("Uncertainty — electricity import-price scenarios (TSSP stage-2)")
        kv("Method", "2-D Latin Hypercube → k-medoid reduction (uncertainty.py)")
        kv("Uncertain input", "electricity import price (level × escalation); export & gas central")
        table(["Scenario", "Weight", "Yr-0 level", "Escalation/yr"],
              [[sid, f"{w:.3f}", f"×{lvl:.3f}", _pct(g)] for (sid, w, lvl, g) in price_scen])
        # The level band is right-skewed (triangular, low 0.67 / mode 1.00 / high 2.27), so its mean
        # sits well above its mode and the scenario set is NOT mean-preserving against the
        # deterministic round. Stated explicitly because the two rounds' workbooks are read side by
        # side: most of the stochastic round's higher cost is this level shift, not a cost of
        # ignoring uncertainty. What IS attributable to uncertainty is that first-stage sizing
        # barely moves between the rounds.
        _wsum = sum(w for (_, w, _, _) in price_scen)
        if _wsum > 0:
            _mean_lvl = sum(w * lvl for (_, w, lvl, _) in price_scen) / _wsum
            kv("Weighted mean level", f"×{_mean_lvl:.3f}  (deterministic round = ×1.000)")
        blank()

    section("Capex (headline)")
    kv("PV", f"£{tc['pv']['capex_per_kwp']:.0f}/kWp")
    kv("Battery", f"£{tc['battery']['energy_capex_per_kwh']:.0f}/kWh + £{tc['battery']['power_capex_per_kw']:.0f}/kW")
    kv("Thermal store", f"£{mp.THERMAL_STORE['energy_capex_per_kwh']:.0f}/kWh")
    kv("Heat plant (£/kW_th)", " · ".join(
        f"{HEAT_SHORT[h]} £{mp.HEAT_COSTS[h]['capex_per_kwth']:.0f}" for h in HEAT_ORDER))
    blank()

    section("Electricity import price bands (DESNZ non-domestic, incl CCL)")
    table(["Band", "MWh/yr", "£/kWh"],
          [[b["name"], _band_range(b), round(b["price"], 4)] for b in mp.ELEC_IMPORT_BANDS])
    blank()
    section("Gas import price bands (DESNZ non-domestic, incl CCL)")
    table(["Band", "MWh/yr", "£/kWh"],
          [[b["name"], _band_range(b), round(b["price"], 4)] for b in mp.GAS_IMPORT_BANDS])
    blank()

    section("Grid connection limits (DNO network headroom, present-day)")
    table(["District", "Import kW", "Export kW"],
          [[d, mp.GRID_LIMITS[d].get("import_kw"), mp.GRID_LIMITS[d].get("export_kw")]
           for d in DIST_ORDER if d in mp.GRID_LIMITS])


def _band_range(b: dict) -> str:
    lo, hi = b["lo"], b["hi"]
    return f"{lo:g}+" if hi == float("inf") else f"{lo:g}–{hi:g}"


# ENTRY POINT
def write_report(dfs: dict, out_path: str, *,
                 run_meta: dict, charts_dir: str, dispatch_png: str = None,
                 demand_png: str = None, front_png: str = None,
                 grid_sensitivity: pd.DataFrame = None, chart_prefix: str = "") -> list:
    # Build the three-objective workbook + standalone PNGs.
    # dispatch_png (optional):      pre-rendered stage-2 dispatch of the featured-cell knee design ("Dispatch" tab).
    # demand_png (optional):        underlying heat vs non-heat electricity demand for the same featured cell/months ("Dispatch" tab).
    # front_png (optional):         the featured-cell continuous cost/carbon frontier + MAC chart ("Pareto" tab).
    # grid_sensitivity (optional):  the district x activity x heat-pump-heating grid-headroom bisection ("Status" tab).
    # chart_prefix (optional):      "deterministic_" / "stochastic_" — keeps both rounds' PNGs in one charts/ folder.
    os.makedirs(charts_dir, exist_ok=True)
    global _CHART_PREFIX
    _CHART_PREFIX = chart_prefix
    try:
        return _write_report(dfs, out_path, run_meta=run_meta, charts_dir=charts_dir,
                             dispatch_png=dispatch_png, demand_png=demand_png,
                             front_png=front_png, grid_sensitivity=grid_sensitivity)
    finally:
        _CHART_PREFIX = ""


def _write_report(dfs: dict, out_path: str, *,
                  run_meta: dict, charts_dir: str, dispatch_png: str = None,
                  demand_png: str = None, front_png: str = None,
                  grid_sensitivity: pd.DataFrame = None) -> list:

    def _prep(df, value_col):
        d = df[df["status"] == "Optimal"].dropna(subset=[value_col]).copy()
        if d.empty:
            raise ValueError(f"No optimal scenarios with {value_col} to report.")
        d["act_short"] = d["activity"].map(lambda a: ACT_SHORT.get(a, a))
        return d

    npv    = _prep(dfs["NPV"],    "npv_savings_GBP")
    carbon = _prep(dfs["Carbon"], "emissions_saving_tco2e")
    pareto = _prep(dfs["Pareto"], "npv_savings_GBP")     

    chart_sets = {                              
        "Status": _status_figs(dfs, charts_dir, grid_sensitivity),
        "NPV": [
            _fig_top10(npv, charts_dir, value_col="npv_savings_GBP", value_fmt=_gbp,
                       value_label="15-year NPV savings vs gas-boiler BAU",
                       title="Top 10 scenarios by NPV savings", fname="npv_top10.png",
                       cost_col="npv_savings_GBP", carbon_col="emissions_saving_tco2e"),
            _fig_marginal(npv, charts_dir, value_col="npv_savings_GBP", value_fmt=_gbp,
                          value_label="NPV savings", fname="npv_marginal.png",
                          cost_col="npv_savings_GBP", carbon_col="emissions_saving_tco2e"),
            _fig_heatmap(npv, charts_dir, value_col="npv_savings_GBP", value_fmt=_gbp,
                         title="Best NPV savings by district × activity\n(best heating per cell)",
                         fname="npv_heatmap.png",
                         cost_col="npv_savings_GBP", carbon_col="emissions_saving_tco2e"),
            _fig_robustness(npv, charts_dir),     # None on a deterministic run
        ],
        "Self-supply": [
            _fig_self_supply(npv, charts_dir),   
            _fig_roof_utilisation(charts_dir),    
        ],
        "Carbon": [
            _fig_top10(carbon, charts_dir, value_col="emissions_saving_tco2e", value_fmt=_tco2e,
                       value_label="Lifetime carbon savings vs gas-boiler BAU",
                       title="Top 10 scenarios by carbon savings", fname="carbon_top10.png",
                       cost_col="npv_savings_GBP", carbon_col="emissions_saving_tco2e"),
            _fig_marginal(carbon, charts_dir, value_col="emissions_saving_tco2e", value_fmt=_tco2e,
                          value_label="Carbon savings", fname="carbon_marginal.png",
                          cost_col="npv_savings_GBP", carbon_col="emissions_saving_tco2e"),
            _fig_heatmap(carbon, charts_dir, value_col="emissions_saving_tco2e", value_fmt=_tco2e,
                         title="Best carbon savings by district × activity\n(best heating per cell)",
                         fname="carbon_heatmap.png",
                         cost_col="npv_savings_GBP", carbon_col="emissions_saving_tco2e"),
            # Sourced from the cost sweep, not `carbon` — see _fig_mac_grid on why a MAC read off
            # the emissions sweep would not be a marginal abatement cost.
            _fig_mac_grid(npv, charts_dir),
        ],
        "Pareto": [
            _fig_pareto_scatter(pareto, charts_dir),
            _fig_pareto_heating(pareto, charts_dir),
        ],
    }
    if front_png and os.path.exists(front_png):
        chart_sets["Pareto"].append(front_png)             # featured-cell continuous frontier + MAC
    if dispatch_png and os.path.exists(dispatch_png):
        chart_sets["Dispatch"] = [dispatch_png]            # stage-2 dispatch of the featured-cell knee
    if demand_png and os.path.exists(demand_png):
        chart_sets.setdefault("Dispatch", []).append(demand_png)   # underlying heat vs non-heat demand

    chart_sets = {tab: [p for p in pngs if p] for tab, pngs in chart_sets.items()}   # drop skipped figs
    chart_sets = {tab: pngs for tab, pngs in chart_sets.items() if pngs}   # drop tabs left with nothing

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        dfs["NPV"].to_excel(xw,    sheet_name="NPV Data",    index=False)
        # On the emissions sweep the £/tCO2e figure is NOT a marginal abatement cost: the design it
        # divides through was chosen with no cost pressure at all, so it is an upper bound on what
        # the abatement costs, not the cheapest way to buy it. Renamed here to say so. The real MAC
        # lives on the NPV sheet (cost-optimal design vs BAU) and, as a curve, on Pareto Front Data.
        dfs["Carbon"].rename(
            columns={"mac_GBP_per_tco2e": "cost_of_emissions_optimal_design_GBP_per_tco2e"}
        ).to_excel(xw, sheet_name="Carbon Data", index=False)
        dfs["Pareto"].to_excel(xw, sheet_name="Pareto Data", index=False)
        if "Pareto front" in dfs and not dfs["Pareto front"].empty:
            dfs["Pareto front"].to_excel(xw, sheet_name="Pareto Front Data", index=False)
        _rq1_table(npv).to_excel(xw, sheet_name="Self-supply Data", index=False)
        _roof_utilisation_table().to_excel(xw, sheet_name="Roof Util Data", index=False)
        if grid_sensitivity is not None and not grid_sensitivity.empty:
            gs = grid_sensitivity.copy()
            gs["edge_margin_label"] = [_edge_margin_label(r) for _, r in gs.iterrows()]
            gs.to_excel(xw, sheet_name="Grid Sensitivity Data", index=False)
        book = xw.book
        _build_cover(book, run_meta)
        for tab, pngs in chart_sets.items():
            _stack_images(book.create_sheet(tab), pngs)
        order = ["Cover", "Status", "NPV", "Self-supply", "Carbon", "Pareto", "Dispatch",
                 "NPV Data", "Carbon Data", "Pareto Data", "Pareto Front Data",
                 "Self-supply Data", "Roof Util Data", "Grid Sensitivity Data"]
        book._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    return [p for pngs in chart_sets.values() for p in pngs]
