"""
Import-only. Rendering layer for demand_profile_model - builds Excel sheets and demand plots. 

Imported by demand_profile_model (build_*_sheet) and by optimisation_model (the demand chart set).

generate_demand_plots() renders one district; generate_all_demand_plots() renders every district
into a single `demand/` folder and is what both demand_profile_model.main() and
optimisation_model.main() call.
"""

import os
import re
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import demand_profile_model as dpm
from demand_profile_model import (
    peak_scale_factor, half_hourly_kw_per_sqm, annual_demand_kwh,
    seasonal_demand_kwh, sample_buildings,
    HEATING_OPTIONS, HEATING_SYSTEMS, SEASON_ORDER_BENCHMARK, SEASON_ORDER_DEMAND,
    MONTHS_ORDER, MONTH_SEASON, OUTPUTS_DIR,
)


# 1 - CONSTANTS
ALT_COLORS = ["F2F7FB", "FFFFFF"]
ACT_PLOT_COLORS = {
    "Health: Health centre":    "#C0392B",
    "Health: Hospital":         "#E67E22",
    "Office: A/C standard":     "#2C3E50",
    "Retail: Department store": "#8E44AD",
}
ACT_PLOT_STYLES = {
    "Health: Health centre":    "-",
    "Health: Hospital":         "-.",
    "Office: A/C standard":     "--",
    "Retail: Department store": ":",
}
PLOT_COMBINATIONS = [   # (heating, energy_type) combinations that produce non-zero plots
    ("Gas Boiler",        "Electricity"),
    ("Gas Boiler",        "Gas"),
    ("ASHP",              "Electricity"),
    ("GSHP (vertical)",   "Electricity"),
    ("GSHP (horizontal)", "Electricity"),
]


# 2 - OUTPUT WORKBOOK BUILDERS
def header_style(cell):
    cell.fill      = PatternFill("solid", start_color="1F4E79")
    cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def data_style(cell, color=None):
    cell.font   = Font(name="Arial", size=10)
    cell.border = thin_border()
    if color:
        cell.fill = PatternFill("solid", start_color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def build_benchmark_sheet(ws, degree_days: dict, daily_hdd: float, activities=None, district: str = None):
    row1  = ["NDB_ActivityClass", "NDB_HeatingOption", "Energy Demand Type", "BaseDiurnalTimeSlice"]
    row1 += ["BaseCDay", "Benchmark kW per sqm"] * len(SEASON_ORDER_BENCHMARK)
    for col_idx, val in enumerate(row1, 1):
        header_style(ws.cell(row=1, column=col_idx, value=val))

    col_widths = [22, 18, 20, 22] + [14, 22] * len(SEASON_ORDER_BENCHMARK)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    row_num = 2
    acts = activities or list(dpm.base_load_fracs.keys())
    for act in acts:
        for heating in HEATING_OPTIONS:
            peak_scale   = peak_scale_factor(act, heating, degree_days, daily_hdd, district)
            energy_types = ["Electricity"] if HEATING_SYSTEMS[heating]["is_heat_pump"] else ["Electricity", "Gas"]
            for etype in energy_types:
                profiles = {}
                for s in SEASON_ORDER_BENCHMARK:
                    p = half_hourly_kw_per_sqm(act, heating, etype, s, degree_days, daily_hdd, district)
                    profiles[s] = p * peak_scale if s == "Peak Winter" else p

                for ts in range(48):
                    color    = ALT_COLORS[row_num % 2]
                    row_data = [act, heating, etype, ts]
                    for s in SEASON_ORDER_BENCHMARK:
                        row_data += [s, round(profiles[s][ts], 5)]
                    for col_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_idx, value=val)
                        data_style(cell, color)
                        if isinstance(val, float):
                            cell.number_format = "0.00000"
                    row_num += 1
    ws.freeze_panes = "A2"

def build_demand_sheet(ws, degree_days: dict, daily_hdd: float, activities=None, district: str = None):
    cols = [
        "NDBuildingID", "ActivityName", "FloorArea (m²)", "NDB_RetrofittingOption",
        "AnnualEnergyDemandTotal (kWh)", "AnnualEnergyDemandElectricity (kWh)", "AnnualEnergyDemandGas (kWh)",
    ] + [
        f"{s.replace(' ', '')}{sfx}"
        for s in SEASON_ORDER_DEMAND
        for sfx in ("EnergyDemandTotal (kWh)", "EnergyDemandElectricity (kWh)", "EnergyDemandGas (kWh)")
    ]
    col_widths = [16, 28, 14, 28, 36, 36, 36] + [40, 44, 40] * len(SEASON_ORDER_DEMAND)
    for col_idx, (h, w) in enumerate(zip(cols, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 30

    acts     = activities or list(dpm.base_load_fracs.keys())
    filtered = [b for b in sample_buildings() if b["activity"] in acts]

    for row_offset, bldg in enumerate(filtered, 2):
        color   = ALT_COLORS[row_offset % 2]
        act     = bldg["activity"]
        heating = bldg["heating"]
        area    = bldg["area"]

        annual   = annual_demand_kwh(act, heating, daily_hdd)
        ann_elec = round(annual["electricity"] * area)
        ann_gas  = round(annual["gas"]         * area)

        row_vals = [bldg["id"], bldg["name"], area, heating.upper().replace(" ", "_"),
                    ann_elec + ann_gas, ann_elec, ann_gas]
        for s in SEASON_ORDER_DEMAND:
            row_vals += list(seasonal_demand_kwh(act, heating, s, area, degree_days, daily_hdd, district))

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_offset, column=col_idx, value=val)
            data_style(cell, color)
            if isinstance(val, (int, float)) and col_idx >= 3:
                cell.number_format = "#,##0"
    ws.freeze_panes = "A2"


# 3 - PLOTTING
def _to_hourly_kwh(half_hour_kw_per_sqm: np.ndarray, activity: str) -> np.ndarray:
    """kWh per hour = (slot_a + slot_b) × 0.5 h × floor_area."""
    p = np.asarray(half_hour_kw_per_sqm)
    return (p[::2] + p[1::2]) * 0.5 * dpm.bees_floor_areas[activity]

def _render_daily_profile_charts(periods, get_profile, n_rows, fig_h, subdir,
                                 slug_prefix, label, activities, file_prefix=""):
    """One stacked-subplot chart per (heating, energy_type) combination."""
    os.makedirs(subdir, exist_ok=True)
    hours   = np.arange(24)
    xlabels = [f"{h:02d}:00" for h in hours]
    acts    = activities or list(dpm.base_load_fracs.keys())

    for heating, etype in PLOT_COMBINATIONS:
        fig, axes = plt.subplots(n_rows, 1, figsize=(14, fig_h), sharex=True)
        fig.patch.set_facecolor("white")
        fig.suptitle(f"Daily Demand Profiles — {etype}  [{label}]\n{heating}",
                     fontsize=13, fontweight="bold")

        for row_i, (ax, period) in enumerate(zip(axes, periods)):
            for act in acts:
                profile    = get_profile(act, heating, etype, period)
                hourly_kwh = _to_hourly_kwh(profile, act)
                ax.plot(hours, hourly_kwh,
                        color=ACT_PLOT_COLORS[act], linestyle=ACT_PLOT_STYLES[act],
                        linewidth=2.0, alpha=0.9, label=act)
            ax.set_title(period, fontsize=10, fontweight="bold", loc="left", pad=4)
            ax.set_ylabel("kWh per hour", fontsize=9)
            ax.set_ylim(bottom=0)
            ax.grid(True, alpha=0.35, linestyle="--", linewidth=0.6)
            ax.set_facecolor("#F8F9FA")
            if row_i < n_rows - 1:
                ax.tick_params(labelbottom=False)

        axes[-1].set_xticks(hours[::2])
        axes[-1].set_xticklabels(xlabels[::2], rotation=45, ha="right", fontsize=8)
        axes[-1].set_xlabel("Hour of Day", fontsize=10)

        handles, labels_leg = axes[0].get_legend_handles_labels()
        fig.legend(handles, labels_leg, title="Activity Class",
                   fontsize=9, title_fontsize=9, framealpha=0.9,
                   edgecolor="#CCCCCC", loc="upper right",
                   bbox_to_anchor=(1.18, 0.97))

        plt.tight_layout(rect=[0, 0, 0.85, 1])
        slug  = f"{file_prefix}{slug_prefix}_{heating.replace(' ', '_')}_{etype}"
        fpath = os.path.join(subdir, f"{slug}.png")
        fig.savefig(fpath, dpi=150, bbox_inches="tight")
        plt.close(fig)
        print(f"Saved: {fpath}")

def _render_wd_we_heatmap(activities, monthly_dd, daily_hdd, subdir, label, district: str = None,
                          file_prefix=""):
    """One 24×24 heatmap (12 months × WD/WE × 24 h) per (heating, energy_type)."""
    os.makedirs(subdir, exist_ok=True)
    acts    = activities or list(dpm.base_load_fracs.keys())
    hours   = np.arange(24)
    xlabels = [f"{h:02d}:00" for h in hours]
    y_labels = []
    for m in MONTHS_ORDER:
        y_labels += [f"{m[:3]} WD", f"{m[:3]} WE"]

    for heating, etype in PLOT_COMBINATIONS:
        all_z = {}
        for act in acts:
            # Same WD/WE split the optimiser uses (dm.wd_we_factors is the single source).
            wd_fac, we_fac = dpm.wd_we_factors(act)
            z_rows = []
            for m in MONTHS_ORDER:
                parent  = MONTH_SEASON[m]
                profile = half_hourly_kw_per_sqm(
                    act, heating, etype, parent,
                    {parent: monthly_dd[m]}, daily_hdd, district,
                )
                hourly_kwh = _to_hourly_kwh(profile, act)
                z_rows.append(hourly_kwh * wd_fac)
                z_rows.append(hourly_kwh * we_fac)
            all_z[act] = np.array(z_rows)  # (24, 24)

        vmax  = max(z.max() for z in all_z.values())
        n     = len(acts)
        ncols = min(n, 2)
        nrows = (n + ncols - 1) // ncols
        fig, axes_grid = plt.subplots(nrows, ncols,
                                      figsize=(9 * ncols, 7 * nrows),
                                      constrained_layout=True)
        axes_flat = np.atleast_1d(axes_grid).flatten()
        fig.patch.set_facecolor("white")
        fig.suptitle(f"Monthly WD/WE Demand Profiles — {etype}  [{label}]\n{heating}",
                     fontsize=13, fontweight="bold")

        im_ref = None
        for ax, act in zip(axes_flat, acts):
            im_ref = ax.imshow(all_z[act], aspect="auto", cmap="RdYlGn_r",
                               vmin=0, vmax=vmax, interpolation="nearest")
            ax.set_title(act, fontsize=10, fontweight="bold", pad=6)
            ax.set_xticks(np.arange(0, 24, 2))
            ax.set_xticklabels(xlabels[::2], rotation=45, ha="right", fontsize=7)
            ax.set_yticks(np.arange(len(y_labels)))
            ax.set_yticklabels(y_labels, fontsize=7)
            ax.set_xlabel("Hour of Day", fontsize=9)
            ax.set_ylabel("Month  (WD = Weekday, WE = Weekend)", fontsize=8)
            for m_i in range(1, 12):
                ax.axhline(m_i * 2 - 0.5, color="white", linewidth=0.8)

        cbar = fig.colorbar(im_ref, ax=axes_flat.tolist(), label="kWh per hour", shrink=0.8)
        cbar.ax.tick_params(labelsize=8)
        slug  = f"{file_prefix}WD_WE_{heating.replace(' ', '_')}_{etype}"
        fpath = os.path.join(subdir, f"{slug}.png")
        fig.savefig(fpath, dpi=150, bbox_inches="tight")
        plt.close(fig)
        print(f"Saved: {fpath}")

def generate_demand_plots(degree_days: dict, daily_hdd: float, label: str, monthly_dd: dict = None, activities: list = None,
                          out_dir: str = OUTPUTS_DIR, district: str = None, file_prefix: str = ""):
    """One chart for each district."""
    os.makedirs(out_dir, exist_ok=True)

    def seasonal_profile(act, heating, etype, season):
        p = half_hourly_kw_per_sqm(act, heating, etype, season, degree_days, daily_hdd, district).copy()
        if season == "Peak Winter":
            p *= peak_scale_factor(act, heating, degree_days, daily_hdd, district)
        return p

    _render_daily_profile_charts(SEASON_ORDER_BENCHMARK, seasonal_profile,
                                 5, 22, out_dir, "Seasonal", label, activities, file_prefix)

    if monthly_dd is not None:
        def monthly_profile(act, heating, etype, month):
            parent = MONTH_SEASON[month]
            m_dd   = {parent: monthly_dd[month]}
            return half_hourly_kw_per_sqm(act, heating, etype, parent, m_dd, daily_hdd, district).copy()

        _render_daily_profile_charts(MONTHS_ORDER, monthly_profile,
                                     12, 44, out_dir, "Monthly", label, activities, file_prefix)
        _render_wd_we_heatmap(activities, monthly_dd, daily_hdd, out_dir, label, district, file_prefix)


def _slug(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_")


def generate_all_demand_plots(out_dir: str, districts: list = None, activities: list = None) -> str:
    """Render the complete demand chart set into a single `demand/` folder under out_dir."""
    if dpm.degree_days_by_district is None:
        dpm.initialize()
    demand_dir = os.path.join(out_dir, "demand")
    os.makedirs(demand_dir, exist_ok=True)
    dists = districts or list(dpm.degree_days_by_district.keys())
    for i, d in enumerate(dists, 1):
        print(f"[demand charts {i}/{len(dists)}] {d}")
        generate_demand_plots(dpm.degree_days_by_district[d], dpm.daily_hdd_by_district[d], d,
                              monthly_dd=dpm.monthly_dd_by_district[d], activities=activities,
                              out_dir=demand_dir, district=d, file_prefix=f"{_slug(d)}_")
    n = len([f for f in os.listdir(demand_dir) if f.endswith(".png")])
    print(f"Demand charts: {n} files in {demand_dir}")
    return demand_dir
