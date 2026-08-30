"""PV technology sensitivity. 
Runs as part of the pipeline (optimisation_model.main() calls run_pv_sensitivity() after the cost rounds), 
or standalone against the most recent completed run:
    python pv_panel1_sensitivity.py [--limit N] [--time-limit S] [--jobs N]
Note: Running standalone against an earlier run predating a sizing bound change makes every row infeasible. 

Compares the model's current PV spec (CEP Technology Library v1.02, Panels 16-18 average) against
Panel 1 (low-end) for every Optimal design in the deterministic cost round.

Method: for each (district, activity, heating) design, first-stage sizing (e_batt, o_batt,
q_heat_cap, e_th) is fixed at the saved deterministic-round solution and only the stage-2 dispatch
is re-solved with Panel 1's efficiency / temperature coefficient / module kWp / module area
substituted for Panels 16-18,

PV is area-matched: n_pv is rescaled so the Panel 1 array occupies the same roof area as the saved design,
with capex_per_kwh scaled by Panel 1's relative price discount.

Output: <run folder>/pv_panel1_sensitivity.xlsx, one row per design, in the same NPV-savings-ranked
row order as the source deterministic sweep.
"""
import argparse
import glob
import multiprocessing as mp
import os

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import demand_profile_model as dm
import optimisation_engine as oe
from optimisation_config import resolve_jobs

# CEP Technology Library v1.02, "PV (monocrystalline silicon)" sheet:
# Panels 16-18 (current model default) = eff 20.9%, temp coeff -0.26%/K,
# 0.365 kWp, 1.75 m^2, price/Wp mean(GBP0.6389, GBP0.6438, GBP0.7081). Panel 1 (low-end) below.
PANEL_1 = dict(efficiency=0.1752, temp_coeff_per_C=-0.0040, module_kwp=0.285, module_area_m2=1.63)
PANEL_1_PRICE_PER_WP = 0.4737
PANELS_16_18_MEAN_PRICE_PER_WP = (0.6389 + 0.6438 + 0.7081) / 3.0
PRICE_RATIO = PANEL_1_PRICE_PER_WP / PANELS_16_18_MEAN_PRICE_PER_WP   # applied to capex_per_kwp

RESULT_COLS = ["status", "pv_kwp", "annual_pv_gen_kwh", "annual_import_kwh", "annual_export_kwh",
               "capex_GBP", "opex_npv_GBP", "total_cost_npv_GBP", "npv_savings_GBP",
               "lifetime_emissions_tco2e"]

BASE_RENAME = {
    "pv_kwp": "base_pv_kwp", "annual_pv_gen_kwh": "base_annual_pv_gen_kwh",
    "annual_import_kwh": "base_annual_import_kwh", "annual_export_kwh": "base_annual_export_kwh",
    "capex_GBP": "base_capex_GBP", "opex_npv_GBP": "base_opex_npv_GBP",
    "total_cost_npv_GBP": "base_total_cost_npv_GBP", "npv_savings_GBP": "base_npv_savings_GBP",
    "lifetime_emissions_tco2e": "base_lifetime_emissions_tco2e",
}

DISPLAY_COLS = [
    "district", "activity", "heating", "panel1_status",
    "n_pv", "panel1_n_pv",          # count-matched vs area-matched module counts
    "base_total_cost_npv_GBP", "panel1_total_cost_npv_GBP", "total_cost_npv_GBP_delta",
    "base_lifetime_emissions_tco2e", "panel1_lifetime_emissions_tco2e", "lifetime_emissions_tco2e_delta",
    "cheap_panel_pays_off",
]

_RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_FONT   = Font(color="9C0006")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(color="006100")

DELTA_COLS = ["total_cost_npv_GBP_delta", "lifetime_emissions_tco2e_delta"]
BOOL_COL = "cheap_panel_pays_off"


def apply_conditional_formatting(ws, columns: list, n_rows: int) -> None:
    # Delta columns: positive (Panel 1 worse) = red, negative (Panel 1 better) = green.
    # BOOL_COL: TRUE (pays off) = green, FALSE = red. NA cells (Infeasible on Panel 1) match neither.
    for col_name in DELTA_COLS:
        col_letter = get_column_letter(columns.index(col_name) + 1)
        cell_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="greaterThan", formula=["0"], fill=_RED_FILL, font=_RED_FONT))
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="lessThan", formula=["0"], fill=_GREEN_FILL, font=_GREEN_FONT))

    col_letter = get_column_letter(columns.index(BOOL_COL) + 1)
    cell_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=["FALSE"], fill=_RED_FILL, font=_RED_FONT))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=["TRUE"], fill=_GREEN_FILL, font=_GREEN_FONT))


def _latest_run_dir() -> str:
    root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
    runs = glob.glob(os.path.join(root, "Optimisation (*)"))
    if not runs:
        raise FileNotFoundError(f"No 'Optimisation (*)' run folders found under {root}")
    return max(runs, key=os.path.getmtime)


def _read_deterministic_rows(run_dir: str) -> pd.DataFrame:
    # The deterministic sweep rows from the workbook written at the end of a run.
    wb = os.path.join(run_dir, "Optimisation Results (deterministic).xlsx")
    if not os.path.exists(wb):
        raise FileNotFoundError(f"No deterministic workbook in {run_dir}.")
    return pd.read_excel(wb, sheet_name="NPV Data")


def _pool_init():
    # Runs once per worker process, so the Panel 1 PV-spec swap happens once here rather than per task.
    dm.initialize()
    base_capex_per_kwp = oe.TECH_COSTS["pv"]["capex_per_kwp"]
    oe.TECH_COSTS["pv"].update(PANEL_1)
    oe.TECH_COSTS["pv"]["capex_per_kwp"] = base_capex_per_kwp * PRICE_RATIO


def _area_matched_n_pv(activity: str, n_pv: int, base_module_area_m2: float,
                       base_n_pv_max: int) -> int:
    # AREA-matched, not COUNT-matched
    panel1_n_pv_max = oe.n_pv_max_for_activity(activity)          # recomputed on Panel 1's geometry
    if n_pv >= base_n_pv_max:
        return panel1_n_pv_max
    scaled = int(round(n_pv * base_module_area_m2 / oe.TECH_COSTS["pv"]["module_area_m2"]))
    return max(1, min(scaled, panel1_n_pv_max))                   # never exceed the roof cap


def _solve_one(task: tuple) -> dict:
    (district, activity, heating, n_pv, e_batt, o_batt, q_heat_cap, e_th,
     base_module_area_m2, base_n_pv_max, threads, time_limit_s) = task
    n_pv_matched = _area_matched_n_pv(activity, n_pv, base_module_area_m2, base_n_pv_max)
    prob, V = oe.build_lp(district, activity, heating,
                            objective="cost", scenarios=[oe.central_scenario()])
    prob += V["n_pv"]       == n_pv_matched, "fix_n_pv_area_matched"
    prob += V["e_batt"]     == e_batt,     "fix_e_batt"
    prob += V["o_batt"]     == o_batt,     "fix_o_batt"
    # Floors, not equalities to prevent infeasibility from rounding. 
    prob += V["q_heat_cap"] >= q_heat_cap, "floor_q_heat_cap"
    prob += V["e_th"]       >= e_th,       "floor_e_th"
    solver = oe._make_solver(solver_msg=False, time_limit_s=time_limit_s, threads=threads)
    status = prob.solve(solver)
    res = oe._extract_results(prob, V, status, district, activity, heating, oe.HORIZON_YEARS)
    return {"district": district, "activity": activity, "heating": heating,
            "panel1_n_pv": n_pv_matched,
            **{f"panel1_{c}": res.get(c) for c in RESULT_COLS}}


def run_pv_sensitivity(all_rows: pd.DataFrame, run_dir: str, *, time_limit_s: int = 600,
                       n_jobs: int = None, limit: int = None) -> pd.DataFrame:
    print(f"Panel 1 / Panels 16-18 price ratio applied to capex_per_kwp: {PRICE_RATIO:.4f}")

    base = all_rows[all_rows["status"] == "Optimal"].reset_index(drop=True)
    base["_rank"] = base.index   # preserves the source file's NPV-savings ranking through the merge
    if limit:
        base = base.iloc[:limit].copy()
    if base.empty:
        print("No Optimal designs to test — skipping Panel 1 sensitivity.")
        return pd.DataFrame()
    print(f"{len(base)} of {len(all_rows)} designs Optimal -- solving Panel 1 dispatch for each")

    # Same worker/thread sizing rule as the rest of the pipeline.
    n_jobs, threads_per_worker = resolve_jobs(n_jobs, len(base))

    # Denominator of the area match.
    base_module_area_m2 = float(oe.TECH_COSTS["pv"]["module_area_m2"])

    tasks = [(r.district, r.activity, r.heating, r.n_pv, r.e_batt_kwh, r.o_batt_kw,
              r.q_heat_cap_kwth, r.e_th_kwh, base_module_area_m2,
              oe.n_pv_max_for_activity(r.activity), threads_per_worker, time_limit_s)
             for r in base.itertuples()]

    print(f"Solving {len(tasks)} designs across {n_jobs} processes x {threads_per_worker} solver threads...")
    with mp.Pool(processes=n_jobs, initializer=_pool_init) as pool:
        panel1_rows = []
        for i, row in enumerate(pool.imap(_solve_one, tasks), 1):
            print(f"  [{i:>3}/{len(tasks)}] {row['activity']!r:27s} {row['heating']!r:16s} "
                  f"{row['district']!r:28s} {row['panel1_status']}")
            panel1_rows.append(row)
    panel1_df = pd.DataFrame(panel1_rows)

    base_ren = base.rename(columns=BASE_RENAME)
    keep = ["district", "activity", "heating", "_rank", "n_pv", "e_batt_kwh", "o_batt_kw",
            "q_heat_cap_kwth", "e_th_kwh"] + list(BASE_RENAME.values())
    out = base_ren[keep].merge(panel1_df, on=["district", "activity", "heating"], how="left")
    out = out.sort_values("_rank").drop(columns="_rank").reset_index(drop=True)

    out["pv_kwp_pct_change"]              = 100 * (out["panel1_pv_kwp"] - out["base_pv_kwp"]) / out["base_pv_kwp"]
    out["annual_pv_gen_kwh_delta"]        = out["panel1_annual_pv_gen_kwh"] - out["base_annual_pv_gen_kwh"]
    out["annual_pv_gen_pct_change"]       = 100 * out["annual_pv_gen_kwh_delta"] / out["base_annual_pv_gen_kwh"]
    out["annual_import_kwh_delta"]        = out["panel1_annual_import_kwh"] - out["base_annual_import_kwh"]
    out["annual_export_kwh_delta"]        = out["panel1_annual_export_kwh"] - out["base_annual_export_kwh"]
    out["capex_GBP_delta"]                = out["panel1_capex_GBP"] - out["base_capex_GBP"]
    out["opex_npv_GBP_delta"]             = out["panel1_opex_npv_GBP"] - out["base_opex_npv_GBP"]
    out["total_cost_npv_GBP_delta"]       = out["panel1_total_cost_npv_GBP"] - out["base_total_cost_npv_GBP"]
    out["lifetime_emissions_tco2e_delta"] = out["panel1_lifetime_emissions_tco2e"] - out["base_lifetime_emissions_tco2e"]
    # NA (not False) when Panel 1's fixed-sizing dispatch is Infeasible
    is_optimal = out["panel1_status"] == "Optimal"
    out["cheap_panel_pays_off"] = pd.Series(pd.NA, index=out.index, dtype="boolean")
    out.loc[is_optimal, "cheap_panel_pays_off"] = out.loc[is_optimal, "total_cost_npv_GBP_delta"] < 0

    out_path = os.path.join(run_dir, "pv_panel1_sensitivity.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        out[DISPLAY_COLS].to_excel(xw, sheet_name="Panel1 vs 16-18", index=False)
        apply_conditional_formatting(xw.sheets["Panel1 vs 16-18"], DISPLAY_COLS, len(out))
        notes = pd.DataFrame({
            "Note": ["Source run", "Panel 1 (low-end)", "Panels 16-18 (model default)",
                     "Price ratio applied to capex_per_kwp", "Method"],
            "Value": [
                run_dir,
                "eff 17.52%, temp coeff -0.40%/K, 0.285 kWp, 1.63 m^2 -- CEP Technology Library v1.02, Panel 1",
                "eff 20.9%, temp coeff -0.26%/K, 0.365 kWp, 1.75 m^2 -- CEP Technology Library v1.02, mean of Panels 16-18",
                f"{PRICE_RATIO:.4f} (Panel 1 / mean(Panels 16-18) GBP/Wp, same source+year)",
                "First-stage sizing (e_batt, o_batt, q_heat_cap, e_th) fixed at the saved "
                "deterministic-round solution; only stage-2 dispatch is re-solved with Panel 1's PV "
                "params swapped in. PV is AREA-matched rather than count-matched: n_pv is rescaled by "
                "1.75/1.63 m^2 (capped at the Panel 1 roof limit) so both panels occupy the same roof area.",
            ],
        })
        notes.to_excel(xw, sheet_name="Assumptions", index=False)

    n_infeasible = int((out["panel1_status"] != "Optimal").sum())
    n_worse      = int((out["cheap_panel_pays_off"] == False).sum())
    n_better     = int((out["cheap_panel_pays_off"] == True).sum())
    print(f"\nWrote {out_path}")
    print(f"{len(out)} designs total: {n_infeasible} Infeasible on Panel 1 under the fixed sizing "
          f"(can't run without a resize -- see panel1_status), {n_worse} solved but net worse off "
          f"over 15 yr despite the cheaper panel, {n_better} solved and net better off.")
    return out


def main():
    ap = argparse.ArgumentParser(description="Standalone re-run against an already-completed "
                                             "optimisation run. The pipeline calls "
                                             "run_pv_sensitivity() directly.")
    ap.add_argument("--limit", type=int, default=None,
                    help="Only solve the first N designs (for a quick test run).")
    ap.add_argument("--time-limit", type=int, default=600, dest="time_limit", metavar="S",
                    help="per-design solver time limit in seconds (default: 600)")
    ap.add_argument("--jobs", type=int, default=None, metavar="N",
                    help="parallel worker processes (default: a quarter of the logical cores; "
                         "see optimisation_config.resolve_jobs)")
    args = ap.parse_args()

    run_dir = _latest_run_dir()
    print(f"Latest run: {run_dir}")
    all_rows = _read_deterministic_rows(run_dir)
    run_pv_sensitivity(all_rows, run_dir, time_limit_s=args.time_limit,
                       n_jobs=args.jobs, limit=args.limit)


if __name__ == "__main__":
    main()
