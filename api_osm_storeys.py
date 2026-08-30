"""
Run-only. Queries national storey statistics from OpenStreetMap (OSM) building:levels data.
Run: python api_osm_storeys.py → data/api_osm_storeys.xlsx. 

No. of above-ground storeys used in roof-area calculation is pulled from OSM via public Overpass API, 
filtered by activity class and sampled across 9 representative cities. 
All samples pooled nationally and summarised per activity class.

A parallel survey pulls `roof:shape` and `roof:angle` for the same activity classes and cities, 
to test the flat-roof assumption used in the PV roof-area model. 

Outputs: data/api_osm_storeys.xlsx with five sheets —
  - "Summary": national median / mean / min / max / n storeys per activity class.
  - "Storey Data": one row per building, with activity class, representative station, name, storeys. 
    `Storeys (raw)` is the exact OSM value; `Storeys` is rounded to the nearest integer.
  - "Roof Shape Summary": per activity, the count of each roof:shape value, total n, the flat-roof share
    by count (Flat %) and by footprint area (Flat % area-wtd), and median roof:angle.
  - "Flat by Footprint": per activity x footprint-size band, the flat-roof share and sample size.
  - "Roof Shape Data": one row per roof:shape-tagged building, with shape, angle and footprint.

Note:
  - OSM coverage is reliant on user contributions, and may not be a fully accurate representation of the true stock.
  - Department stores are sparse in OSM; `building=retail` is a broader category that captures all retail.
  - Scotland W uses Glasgow city (instead of Prestwick) for an adequate building sample.
  - `roof:shape` is better populated than `roof:angle`; expect thin angle samples.

Usage:
    pip install requests pandas openpyxl
    python api_osm_storeys.py
"""
import os
import time
import math
import requests
import pandas as pd

import datasets   # footprint-band definition, shared with the optimisation model
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 1 - OVERALL SETUP
# Public Overpass endpoints, tried in order. The script backs off during rate limits (HTTP 429) or timeouts (504/503).
OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://maps.mail.ru/osm/tools/overpass/api/interpreter",
]
MAX_RETRIES   = 5     # per endpoint
BACKOFF_BASE  = 10    # seconds; doubles each retry 
PAUSE_BETWEEN = 8     # seconds interval between successful queries 

HEADERS    = {"User-Agent": "research-storey-survey/1.0 (academic use)"}
OUTPUT_XLSX   = os.path.join("data", "api_osm_storeys.xlsx")        # to be consumed by optimisation_model 
RAW_CSV       = os.path.join("cache", "osm_storeys_raw.csv")        # storey pull (resumable scratch)
ROOFSHAPE_CSV = os.path.join("cache", "osm_roofshape_raw.csv")      # roof:shape pull (resumable scratch)

# Representative-station cities, one per Met Office district in districts.py. bbox = (south, west, north, east).
# NOTE: Scotland W surveys Glasgow rather than the district's Prestwick station, for sample size.
STATIONS = {
    "East Anglia":              ("Norwich",    (52.58,  1.21, 52.68,  1.36)),
    "England E and NE":         ("Newcastle",  (54.94, -1.72, 55.03, -1.50)),
    "England NW and N Wales":   ("Manchester", (53.40, -2.30, 53.52, -2.16)),
    "England SE and Central S": ("London",     (51.28, -0.51, 51.69,  0.33)),
    "England SW and S Wales":   ("Exeter",     (50.68, -3.58, 50.76, -3.45)),
    "Midlands":                 ("Birmingham", (52.40, -1.98, 52.55, -1.80)),
    "Scotland E":               ("Edinburgh",  (55.90, -3.33, 55.99, -3.08)),
    "Scotland N":               ("Inverness",  (57.44, -4.27, 57.51, -4.16)),
    "Scotland W":               ("Glasgow",    (55.80, -4.34, 55.90, -4.12)),
}

# Activity classes named to match demand_profile_model.py. 
ACTIVITY_FILTERS = {
    "Office: A/C standard":     ['"building"="office"', '"office"'],
    "Retail: Department store": ['"shop"="department_store"', '"building"="retail"'],
    "Health: Health centre":    ['"amenity"="clinic"', '"healthcare"="centre"', '"amenity"="doctors"'],
    "Health: Hospital":         ['"amenity"="hospital"', '"building"="hospital"'],
}

# Activity order for the summary sheet.
ACTIVITY_ORDER = list(ACTIVITY_FILTERS.keys())

# Tag keys (for the "Type Tags" column).
DESCRIPTOR_KEYS = ("building", "amenity", "office", "shop", "healthcare")

# 2 - QUERY SETUP
def fetch(query):
    for endpoint in OVERPASS_ENDPOINTS:
        host = endpoint.split("/")[2]
        for attempt in range(MAX_RETRIES):
            try:
                r = requests.post(endpoint, data={"data": query}, headers=HEADERS, timeout=300)
            except requests.RequestException as exc:
                wait = BACKOFF_BASE * (2 ** attempt)
                print(f"    [{host}] network error ({exc}); retry in {wait}s")
                time.sleep(wait)
                continue
            if r.status_code == 200:
                return r.json().get("elements", [])
            if r.status_code in (429, 503, 504):
                retry_after = r.headers.get("Retry-After")
                wait = int(retry_after) if (retry_after and retry_after.isdigit()) \
                    else BACKOFF_BASE * (2 ** attempt)
                print(f"    [{host}] HTTP {r.status_code}; retry in {wait}s")
                time.sleep(wait)
                continue
            r.raise_for_status()  # surface other errors 
        print(f"    [{host}] exhausted retries; trying next mirror")
    raise RuntimeError("All Overpass endpoints failed after retries")

def build_query(bbox, filters, key="building:levels", out="tags"):
    # Union of elements carrying `key` and matching any of the activity `filters`.
    s, w, n, e = bbox
    b = f"({s},{w},{n},{e})"
    parts = []
    for f in filters:
        parts.append(f'way["{key}"][{f}]{b};')
        parts.append(f'relation["{key}"][{f}]{b};')
    return f"[out:json][timeout:180];({''.join(parts)});out {out};"

def parse_storeys(raw):
    """Return a clean storey count from a raw building:levels value, or None if invalid."""
    if raw is None:
        return None
    try:
        val = float(str(raw).split(";")[0].strip())   # handle "2;3" -> 2
    except ValueError:
        return None
    return val if 0 < val <= 100 else None            # drop junk values

def descriptor(tags):
    """Compact 'key=value; ...' string of the identifying tags."""
    return "; ".join(f"{k}={tags[k]}" for k in DESCRIPTOR_KEYS if k in tags)

def collect_rows():
    """Query every (activity x station) combination and return a list of row dicts."""
    rows = []
    for activity, filters in ACTIVITY_FILTERS.items():
        for district, (city, bbox) in STATIONS.items():
            elements = fetch(build_query(bbox, filters))
            n = 0
            for el in elements:
                tags = el.get("tags", {})
                storeys = parse_storeys(tags.get("building:levels"))
                if storeys is None:
                    continue
                rows.append({
                    "Activity Class":       activity,
                    "District":             district,
                    "Representative Station": city,
                    "OSM Type":             el.get("type", ""),
                    "OSM ID":               el.get("id", ""),
                    "Building Name":        tags.get("name", ""),
                    "Storeys (raw)":        storeys,   # exact OSM building:levels value
                    "Type Tags":            descriptor(tags),
                    "OSM URL":              f"https://www.openstreetmap.org/{el.get('type')}/{el.get('id')}",
                })
                n += 1
            print(f"  {activity:26s} {city:11s} n={n:4d}")
            time.sleep(PAUSE_BETWEEN)  
    return rows


def parse_roof_shape(raw):
    """Normalise a roof:shape value to a lowercase string, or None if absent."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    return s or None

def parse_roof_angle(raw):
    """Return roof:angle in degrees as a float, or None if absent/invalid."""
    if raw is None:
        return None
    try:
        return float(str(raw).split(";")[0].strip())
    except ValueError:
        return None

_EARTH_R = 6_371_000.0  # m

def _ring_area_m2(coords):
    if len(coords) < 3:
        return 0.0
    lat0 = sum(c[0] for c in coords) / len(coords)
    cos0 = math.cos(math.radians(lat0))
    pts  = [(math.radians(lon) * _EARTH_R * cos0, math.radians(lat) * _EARTH_R)
            for lat, lon in coords]
    s = 0.0
    for i in range(len(pts)):
        x1, y1 = pts[i]
        x2, y2 = pts[(i + 1) % len(pts)]
        s += x1 * y2 - x2 * y1
    return abs(s) / 2.0

def footprint_m2(el):
    """Building footprint area (m2) from an Overpass element."""
    t = el.get("type")
    if t == "way":
        geom = el.get("geometry") or []
        return _ring_area_m2([(g["lat"], g["lon"]) for g in geom])
    if t == "relation":
        total = 0.0
        for mem in el.get("members", []):
            geom = mem.get("geometry") or []
            if len(geom) < 3:
                continue
            a = _ring_area_m2([(g["lat"], g["lon"]) for g in geom])
            total += a if mem.get("role") == "outer" else -a
        return max(0.0, total)
    return 0.0

def collect_roof_rows():
    """Query roof:shape per (activity x station); one row per tagged building."""
    rows = []
    for activity, filters in ACTIVITY_FILTERS.items():
        for district, (city, bbox) in STATIONS.items():
            elements = fetch(build_query(bbox, filters, key="roof:shape", out="geom"))
            n = 0
            for el in elements:
                tags  = el.get("tags", {})
                shape = parse_roof_shape(tags.get("roof:shape"))
                if shape is None:
                    continue
                rows.append({
                    "Activity Class":       activity,
                    "District":             district,
                    "Representative Station": city,
                    "OSM Type":             el.get("type", ""),
                    "OSM ID":               el.get("id", ""),
                    "Building Name":        tags.get("name", ""),
                    "Roof Shape":           shape,
                    "Roof Angle":           parse_roof_angle(tags.get("roof:angle")),
                    "Footprint m2":         round(footprint_m2(el), 1),
                    "Type Tags":            descriptor(tags),
                    "OSM URL":              f"https://www.openstreetmap.org/{el.get('type')}/{el.get('id')}",
                })
                n += 1
            print(f"  [roof] {activity:26s} {city:11s} n={n:4d}")
            time.sleep(PAUSE_BETWEEN)  
    return rows


# 3 - EXCEL SETUP
COLUMN_ORDER = [
    "Activity Class", "District", "Representative Station", "OSM Type", "OSM ID", "Building Name", "Storeys", "Storeys (raw)", "Type Tags", "OSM URL",
]


def prepare(df):
    """OSM `building:levels` is defined as a positive integer; fractions typically represent partial floors. 
    Exact values are kept in `Storeys (raw)` and rounded to the nearest integer. 
    Also handles older cache files that stored the raw value under `Storeys`."""
    if "Storeys (raw)" not in df.columns and "Storeys" in df.columns:
        df = df.rename(columns={"Storeys": "Storeys (raw)"})        # identify cache files
    df["Storeys"] = df["Storeys (raw)"].round().astype(int)         # round storey values to nearest integer
    for col in ("Building Name", "Type Tags"):                      # CSV NaN -> blank
        if col in df.columns:
            df[col] = df[col].fillna("")
    df = df[[c for c in COLUMN_ORDER if c in df.columns]]
    return df.sort_values(["Activity Class", "District", "Storeys"]).reset_index(drop=True)

def summarise(raw_df):
    """National median / mean / min / max / n per activity class (uses rounded `Storeys`)."""
    g = raw_df.groupby("Activity Class")["Storeys"]
    summary = pd.DataFrame({
        "Sample Size (n)": g.count(),
        "Median Storeys":  g.median().round(1),   # median of integers can be x.5
        "Mean Storeys":    g.mean().round(2),
        "Min Storeys":     g.min().astype("Int64"),
        "Max Storeys":     g.max().astype("Int64"),
    })
    # Keep the four activities in a fixed order; show classes with no data as blank rows.
    summary = summary.reindex(ACTIVITY_ORDER)
    return summary.reset_index().rename(columns={"index": "Activity Class"})

ROOF_COLUMN_ORDER = [
    "Activity Class", "District", "Representative Station", "OSM Type", "OSM ID",
    "Building Name", "Roof Shape", "Roof Angle", "Footprint m2", "Type Tags", "OSM URL",
]

# Footprint-size bands (m2) to evaluate proportion of flat roofs per building size.
# Single source of truth, shared with optimisation_engine (which reads these bands back).
FOOTPRINT_BINS   = datasets.FOOTPRINT_BINS
FOOTPRINT_LABELS = datasets.FOOTPRINT_LABELS

def prepare_roof(df):
    """Restore blank text cells (CSV NaN -> ''), order columns, and sort."""
    for col in ("Building Name", "Type Tags", "Roof Shape"):
        if col in df.columns:
            df[col] = df[col].fillna("")
    df = df[[c for c in ROOF_COLUMN_ORDER if c in df.columns]]
    return df.sort_values(["Activity Class", "District", "Roof Shape"]).reset_index(drop=True)

def summarise_roof_shape(raw_df):
    """Per activity: count of each roof:shape, total n, flat share by count and by footprint area,
    and median roof:angle."""
    ct = pd.crosstab(raw_df["Activity Class"], raw_df["Roof Shape"])
    shape_cols = sorted(ct.columns, key=lambda c: ct[c].sum(), reverse=True)  # most common first
    ct = ct[shape_cols]
    ct["n"]      = ct.sum(axis=1)
    flat         = ct["flat"] if "flat" in ct.columns else 0
    ct["Flat %"] = (flat / ct["n"] * 100).round(1)

    # Footprint-area-weighted flat share.
    sized = raw_df[raw_df["Footprint m2"] > 0].copy()
    sized["flat_area"] = sized["Footprint m2"] * (sized["Roof Shape"] == "flat")
    g  = sized.groupby("Activity Class")
    aw = (g["flat_area"].sum() / g["Footprint m2"].sum() * 100).round(1)
    ct["Flat % (area-wtd)"] = aw

    # roof:angle is sparse; report it separately.
    ang = raw_df.dropna(subset=["Roof Angle"]).groupby("Activity Class")["Roof Angle"]
    ct["n (angle)"]    = ang.count()
    ct["Median angle"] = ang.median().round(1)

    ct = ct.reindex(ACTIVITY_ORDER)
    return ct.reset_index().rename(columns={"index": "Activity Class"})

def summarise_roof_by_footprint(raw_df):
    """Flat-roof share by (activity x footprint-size band) — tests whether larger buildings are flatter."""
    d = raw_df[raw_df["Footprint m2"] > 0].copy()
    d["is_flat"]        = (d["Roof Shape"] == "flat").astype(float)
    d["Footprint band"] = pd.cut(d["Footprint m2"], bins=FOOTPRINT_BINS,
                                 labels=FOOTPRINT_LABELS, right=False)
    grp = d.groupby(["Activity Class", "Footprint band"], observed=False)
    out = pd.DataFrame({
        "n":                 grp.size(),
        "Flat %":            (grp["is_flat"].mean() * 100).round(1),
        "Mean footprint m2": grp["Footprint m2"].mean().round(0),
    }).reset_index()
    out["Activity Class"] = pd.Categorical(out["Activity Class"], categories=ACTIVITY_ORDER, ordered=True)
    return out.sort_values(["Activity Class", "Footprint band"]).reset_index(drop=True)

def style_sheet(ws, n_cols):
    """Bold header, centre, freeze top row, and roughly auto-size columns."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        width = max((len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
                    default=10)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 60)

def _load_or_collect(csv_path, collector, label, required_cols=()):
    """Reuse a cached pull if present and contains required columns, else collect and checkpoint it.
    Delete the CSV / run with a cache missing required columns to re-query."""
    if os.path.exists(csv_path):
        df = pd.read_csv(csv_path)
        missing = [c for c in required_cols if c not in df.columns]
        if not missing:
            print(f"Reusing cached {label}: {csv_path}  (delete this file to re-query OSM)")
            return df
        print(f"Cached {label} missing {missing}; re-querying OSM ...")
    df = pd.DataFrame(collector())
    df.to_csv(csv_path, index=False)   # checkpoint before any summary/Excel step
    print(f"Cached {label}: {csv_path}")
    return df

def main():
    os.makedirs(os.path.dirname(OUTPUT_XLSX), exist_ok=True)
    os.makedirs(os.path.dirname(RAW_CSV), exist_ok=True)   # checkpoint cache dir (cache/)

    # Storey survey (building:levels)
    raw_df     = prepare(_load_or_collect(RAW_CSV, collect_rows, "storey data"))
    summary_df = summarise(raw_df)

    # Roof-shape survey (roof:shape / roof:angle + footprint) — tests the flat-roof assumption.
    roof_df         = prepare_roof(_load_or_collect(ROOFSHAPE_CSV, collect_roof_rows,
                                                    "roof-shape data", required_cols=("Footprint m2",)))
    roof_summary_df = summarise_roof_shape(roof_df)
    roof_size_df    = summarise_roof_by_footprint(roof_df)

    print("\nNational storey summary:")
    print(summary_df.to_string(index=False))
    print("\nNational roof-shape summary:")
    print(roof_summary_df.to_string(index=False))
    print("\nFlat-roof share by footprint band:")
    print(roof_size_df.to_string(index=False))

    sheets = [
        ("Summary",            summary_df),
        ("Storey Data",        raw_df),
        ("Roof Shape Summary", roof_summary_df),
        ("Flat by Footprint",  roof_size_df),
        ("Roof Shape Data",    roof_df),
    ]
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for name, dfx in sheets:
            dfx.to_excel(writer, sheet_name=name, index=False)
            style_sheet(writer.sheets[name], dfx.shape[1])
    print(f"\nSaved: {OUTPUT_XLSX}  ({len(raw_df)} buildings, {len(roof_df)} with roof shape)")


if __name__ == "__main__":
    main()
