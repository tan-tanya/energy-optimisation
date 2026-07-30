"""
Wholesale + regional DUoS import-price build-up.

Year-0 intra-day, regional, seasonal import price per (month, day_type, half-hour slot):
- wholesale·HH-shape + per-DNO DUoS Red/Amber/Green + CCL + band-specific residual.

DUoS unit rates and the Red/Amber/Green time-band windows are transcribed per district from each
DNO's 2025/26 CDCM Schedule of Charges; wholesale is real Elexon MID 2025 (via api_wholesale_prices.py).

The parsed build-up inputs are cached once per process. The scenario import-price PATH (escalation)
is applied downstream by the optimisation model; this module only produces the year-0 central level.
"""

import demand_profile_model as dm
from optimisation_config import HH_PER_DAY, T_RES_H, S_KEYS

_WD_BUILDUP = None   # cached parsed build-up inputs {wholesale, ccl, residual_by_band, bands, duos, shape}


def _load_wholesale_duos(path=None):
    global _WD_BUILDUP
    if _WD_BUILDUP is not None:
        return _WD_BUILDUP
    import api_wholesale_prices as wp
    import openpyxl
    from model_params import PARAMS_XLSX
    wb = openpyxl.load_workbook(path or PARAMS_XLSX, data_only=True)
    ws = wb["Energy Prices"]
    hdr, cols = None, None
    for r in range(1, 9):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if "Group" in vals and "Parameter" in vals and "Value" in vals:
            hdr = r
            cols = {v: i + 1 for i, v in enumerate(vals) if v}
            break
    if hdr is None:
        raise ValueError("Energy Prices sheet: header row (Group/Parameter/Value) not found")
    ccl = None
    residual_by_band = {}        # {DESNZ import band name -> band-specific non-commodity residual}
    bands, duos = {}, {}
    for r in range(hdr + 1, ws.max_row + 1):
        g     = ws.cell(row=r, column=cols["Group"]).value
        key   = ws.cell(row=r, column=cols["Key"]).value
        param = ws.cell(row=r, column=cols["Parameter"]).value
        val   = ws.cell(row=r, column=cols["Value"]).value
        if val is None or param is None:
            continue
        if   g == "Retail Non-Commodity"  and param == "ccl":                  ccl = float(val)
        elif g == "Retail Non-Commodity"  and param == "non_commodity_other":  residual_by_band[key] = float(val)
        elif g == "DUoS Time Band":        bands.setdefault(key, {})[param] = float(val)   # per-district (DNO) windows
        elif g == "DUoS Unit Charge":      duos.setdefault(key, {})[param] = float(val)
    wb.close()
    missing = [n for n, v in (("ccl", ccl), ("non_commodity_other (band rows)", residual_by_band or None)) if v is None]
    if missing:
        raise KeyError(f"Wholesale/DUoS build-up: missing 'Energy Prices' rows {missing}")

    wholesale = wp.load_level()
    shape     = wp.load_shape()
    _WD_BUILDUP = {"wholesale": wholesale, "ccl": ccl, "residual_by_band": residual_by_band,
                   "bands": bands, "duos": duos, "shape": shape}
    return _WD_BUILDUP


def _wholesale_shape(month: str, day_type: str, shape: dict) -> list:
    # Wholesale multiplier (48 slots) for a (month, day_type) via its season.
    key = f"{dm.MONTH_SEASON[month]}_{day_type}"
    prof = shape.get(key)
    return prof if prof else [1.0] * HH_PER_DAY


def _duos_band_for_slot(day_type: str, t: int, bands: dict) -> str:
    # DUoS Red/Amber/Green band of half-hour slot t (0..47) on a WD/WE day.
    h = t * T_RES_H   # slot start hour, 0.0 .. 23.5
    if day_type == "WE":
        we_s, we_e = bands.get("duos_we_amber_start_h", 0.0), bands.get("duos_we_amber_end_h", 0.0)
        return "amber" if (we_e > we_s and we_s <= h < we_e) else "green"
    if bands["duos_red_start_h"] <= h < bands["duos_red_end_h"]:
        return "red"
    if (bands["duos_amber_am_start_h"] <= h < bands["duos_amber_am_end_h"]
            or bands["duos_amber_pm_start_h"] <= h < bands["duos_amber_pm_end_h"]):
        return "amber"
    return "green"


def import_price_slots_central(district: str, band_name: str) -> dict:
    # Year-0 CENTRAL import price (GBP/kWh) for every (month, day_type, slot) from the build-up:
    wd = _load_wholesale_duos()
    du = wd["duos"].get(district)
    if du is None:
        raise KeyError(f"No 'DUoS Unit Charge' rows for district {district!r} on the Energy Prices sheet")
    residual = wd["residual_by_band"].get(band_name)
    if residual is None:
        raise KeyError(f"No 'non_commodity_other' residual for import band {band_name!r}; add a "
                       f"'Retail Non-Commodity' row keyed by that band on the Energy Prices sheet")
    unit  = {"red": du["duos_red_unit"], "amber": du["duos_amber_unit"], "green": du["duos_green_unit"]}
    bands_d = wd["bands"].get(district)
    if bands_d is None:
        raise KeyError(f"No 'DUoS Time Band' rows for district {district!r} on the Energy Prices sheet")
    fixed = wd["ccl"] + residual
    raw = {}
    for (m, d) in S_KEYS:
        wshape = _wholesale_shape(m, d, wd["shape"])
        for t in range(HH_PER_DAY):
            raw[(m, d, t)] = (wd["wholesale"] * wshape[t]
                              + unit[_duos_band_for_slot(d, t, bands_d)]
                              + fixed)
    return raw
