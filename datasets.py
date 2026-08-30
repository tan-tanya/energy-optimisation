"""
Import-only.

One get_*() per source, returning a clean object (pd.DataFrame / pd.ExcelFile) and hiding
formatting behind a single import. File-based sources only.

Sources:
  inputs_workbook()            data/inputs.xlsx                          -> pd.ExcelFile
  get_degree_days(icao)        data/hdd/{icao}_HDD_15.5C.csv             -> DataFrame[date, hdd]
  get_ukcp(region)             data/climateprojections/UKCP_{region}.csv -> DataFrame[year, month, delta_T_mean]
  get_sunshine(district)       data/sunlighthours/{sunshine_file}        -> DataFrame (lowercased cols)
  get_osm_summary()            data/api_osm_storeys.xlsx :: Summary          -> DataFrame
  get_osm_flat_by_footprint()  data/api_osm_storeys.xlsx :: Flat by Footprint-> DataFrame
  get_temperature_profile()    data/api_temperature_profiles.xlsx :: HourlyTemp -> np.ndarray(24)
  get_temperature_anomaly()    data/api_temperature_profiles.xlsx :: HourlyTemp -> np.ndarray(24)

Also owns the footprint-size band definition (FOOTPRINT_BINS / FOOTPRINT_LABELS) shared by the
producer of the OSM survey (api_osm_storeys.py) and its consumer (optimisation_engine.py).
"""

import os
import numpy as np
import openpyxl
import pandas as pd
from districts import SUNSHINE_FILE
from seasons import HOURS_PER_DAY

# 1 - SETUP
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

def _path(*parts):
    return os.path.join(DATA_DIR, *parts)

INPUTS_XLSX       = _path("inputs.xlsx")
OSM_STOREYS_XLSX  = _path("api_osm_storeys.xlsx")
TEMPERATURE_PROFILES_XLSX = _path("api_temperature_profiles.xlsx")

# Footprint-size bands for the OSM roof survey. 
FOOTPRINT_BINS   = [0, 250, 1000, 5000, float("inf")]
FOOTPRINT_LABELS = ["<250", "250-1,000", "1,000-5,000", ">=5,000"]

# Degree-days baseline window: average HDD data over the last N calendar years.
HDD_BASELINE_YEARS = 3

# Sunshine-hours baseline window: average seasonal sunshine over the last N complete years
# Used by demand_profile_model.load_light_factors.
SUNSHINE_BASELINE_YEARS = 10

def read_dated_csv(path, **read_csv_kwargs):
    """Read a CSV whose real header row starts with 'Date', preceded by metadata lines."""
    with open(path) as f:
        skip = next(i for i, line in enumerate(f) if line.startswith("Date"))
    return pd.read_csv(path, skiprows=skip, **read_csv_kwargs)

def inputs_workbook():
    """data/inputs.xlsx as a pd.ExcelFile."""
    return pd.ExcelFile(INPUTS_XLSX)


# 2 - HDD
def get_degree_days(icao, last_years=None):
    """Clean daily HDD series for an ICAO station: DataFrame[date (datetime), hdd (float)].
    last_years=N keeps a rolling N-year window ending at the latest date in the file (default: all).
    Rolling so 36 months over 4 partial calendar years still yields ~N samples for every month."""
    df = read_dated_csv(_path("hdd", f"{icao}_HDD_15.5C.csv"), usecols=[0, 1])
    df.columns = ["date", "hdd"]
    df["hdd"]  = pd.to_numeric(df["hdd"], errors="coerce")
    df = df.dropna(subset=["date", "hdd"]).copy()
    df["date"] = pd.to_datetime(df["date"])
    if last_years is not None:
        cutoff = df["date"].max() - pd.DateOffset(years=last_years)
        df = df[df["date"] > cutoff].copy()
    return df


# 3 - UKCP18 RCP8.5 TEMPERATURE ANOMALIES
def get_ukcp(region):
    """Ensemble-mean monthly UKCP18 anomalies for a region: DataFrame[year, month, delta_T_mean]."""
    df = read_dated_csv(_path("climateprojections", f"UKCP_{region}.csv"))
    df = df.rename(columns={df.columns[0]: "date"})
    df["date"]  = pd.to_datetime(df["date"])
    df["year"]  = df["date"].dt.year
    df["month"] = df["date"].dt.month
    members = [c for c in df.columns if c not in ("date", "year", "month")]
    df[members] = df[members].apply(pd.to_numeric, errors="coerce")
    df["delta_T_mean"] = df[members].mean(axis=1)
    return df[["year", "month", "delta_T_mean"]]


# 4 - MET OFFICE SEASONAL SUNSHINE
def get_sunshine(district):
    """Parsed seasonal-sunshine table for a district (columns lowercased), from the registry file."""
    df = pd.read_csv(_path("sunlighthours", SUNSHINE_FILE[district]), skiprows=5, sep=r"\s+", engine="python")
    df.columns = [c.lower() for c in df.columns]
    return df


# 5 - OSM ROOF DATA (API)
def _require_osm():
    if not os.path.exists(OSM_STOREYS_XLSX):
        raise FileNotFoundError(
            f"OSM storey survey not found at {OSM_STOREYS_XLSX}. Run api_osm_storeys.py first to generate it."
        )

def get_osm_summary():
    """data/api_osm_storeys.xlsx :: 'Summary' — median/mean storeys per activity class."""
    _require_osm()
    return pd.read_excel(OSM_STOREYS_XLSX, sheet_name="Summary")

def get_osm_flat_by_footprint():
    """data/api_osm_storeys.xlsx :: 'Flat by Footprint' — flat-roof share per (activity, footprint band)."""
    _require_osm()
    return pd.read_excel(OSM_STOREYS_XLSX, sheet_name="Flat by Footprint")


# 6 - ERA5 TEMPERATURE (API) 
# Getters return None when the workbook is absent; 
# demand_profile_model falls back to idealised sinusoid instead of failing.
_TEMPERATURE_CACHE = None

def _load_temperature_profiles(path=TEMPERATURE_PROFILES_XLSX):
    """Parse the workbook into {(district, season): [24 floats]} (absolute °C). Cached."""
    global _TEMPERATURE_CACHE
    if _TEMPERATURE_CACHE is not None:
        return _TEMPERATURE_CACHE
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Temperature profiles not found at {path}. Run 'python api_temperature_profiles.py' "
            f"first to generate it."
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["HourlyTemp"]
    # locate header row containing 'District'
    hdr = None
    for rr in range(1, 8):
        if ws.cell(row=rr, column=1).value == "District":
            hdr = rr
            break
    if hdr is None:
        raise ValueError(f"Could not find a 'District' header row in {path}::HourlyTemp")
    profiles = {}
    for rr in range(hdr + 1, ws.max_row + 1):
        district = ws.cell(row=rr, column=1).value
        season   = ws.cell(row=rr, column=2).value
        if district is None or season is None:
            continue
        vals = [ws.cell(row=rr, column=3 + h).value for h in range(HOURS_PER_DAY)]
        if all(isinstance(v, (int, float)) for v in vals):
            profiles[(district, season)] = [float(v) for v in vals]
    wb.close()
    _TEMPERATURE_CACHE = profiles
    return _TEMPERATURE_CACHE

def _season_key(season):
    """Map a model season to the stored shape season."""
    return "Winter" if season == "Peak Winter" else season

def get_temperature_profile(district, season, path=TEMPERATURE_PROFILES_XLSX):
    """Absolute mean 24-hour temperature [°C] for (district, season) as np.ndarray(24), or None
    if the data is unavailable for that key."""
    try:
        profiles = _load_temperature_profiles(path)
    except FileNotFoundError:
        return None
    vals = profiles.get((district, _season_key(season)))
    return np.array(vals) if vals is not None else None

def get_temperature_anomaly(district, season, path=TEMPERATURE_PROFILES_XLSX):
    """Mean-zero diurnal anomaly for (district, season): deviation of each hour from its daily mean."""
    prof = get_temperature_profile(district, season, path)
    if prof is None:
        return None
    return prof - prof.mean()

# 7 - PROJECTION OUTPUT
ELEC_PROJECTION_CSV    = _path("electricity_projection_output.csv")
CLIMATE_PROJECTION_CSV = _path("climate_projection_output.csv")

def get_electricity_projection():
    """DESNZ non-heat electricity demand growth: DataFrame[year, electricity_twh, growth_factor]
    (relative to 2025). Generated by projections.py (run_electricity_projection)."""
    if not os.path.exists(ELEC_PROJECTION_CSV):
        raise FileNotFoundError(
            f"Electricity projection not found at {ELEC_PROJECTION_CSV}. "
            f"Run projections.py (or demand_profile_model.py) first.")
    return pd.read_csv(ELEC_PROJECTION_CSV)

def get_climate_projection():
    """UKCP18 HDD projection per (district, year, month): DataFrame with baseline_hdd_per_day and
    projected_hdd_per_day. Generated by projections.py (run_sigma_fit / run_hdd_projection)."""
    if not os.path.exists(CLIMATE_PROJECTION_CSV):
        raise FileNotFoundError(
            f"Climate projection not found at {CLIMATE_PROJECTION_CSV}. "
            f"Run projections.py (or demand_profile_model.py) first.")
    return pd.read_csv(CLIMATE_PROJECTION_CSV)
