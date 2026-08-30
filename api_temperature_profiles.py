"""
Run-only. Queries hourly, per-district air temperatures from Open-Meteo Historical Weather.
Run: python api_temperature_profiles.py → data/api_temperature_profiles.xlsx.
    Workbook is read through datasets.get_temperature_anomaly(), used by demand_profile_model's 
    hourly_temp_profile() to give intra-day temperature a real diurnal shape.
      
Pulls ERA5 hourly 2 m air temperature from the Open-Meteo Historical Weather API for each district's representative station, 
and reduces a multi-year window to an average 24-hour day per season. The model adds an hourly anomaly onto the HDD-derived 
daily-mean temperature, sharpening ASHP COP, PV cell-temperature and heat-demand shape.

Output: data/api_temperature_profiles.xlsx —
  - "HourlyTemp": one row per (District, Season) × 24 h, mean hourly ERA5 2 m temperature [°C].
  - "Meta":       source, ERA5 model, window years, timezone, pull timestamp.

Seasons follow the model's calendar: Winter = Dec/Jan/Feb, Spring = Mar/Apr/May, Summer = Jun/Jul/Aug, Autumn = Sep/Oct/Nov.
"Peak Winter" is synthetic (single coldest half-hour); loader maps it to Winter.

Notes:
  - GB timezone.
  - Window defaults to the last 10 complete calendar years (→ 2016-2025 today), matching the sunshine baseline.
  - ERA5 has ~5-day delay; the latest complete year is always fully available.
  - Pulled year-by-year per district; ~9 districts × N years calls.

Usage:
    python api_temperature_profiles.py                 
    python api_temperature_profiles.py --years 10
    python api_temperature_profiles.py --start 2016 --end 2025
"""
import os
import sys
import time
import json
import argparse
import datetime as dt
import urllib.parse
import urllib.request
import urllib.error

import openpyxl
from openpyxl.styles import Font

from districts import DISTRICTS, DISTRICT_LATLON
from seasons import MONTH_SEASON_BY_NUM as MONTH_SEASON, SEASONS, HOURS_PER_DAY

# 1 - OVERALL SETUP
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
TEMPERATURE_PROFILES_XLSX = os.path.join(DATA_DIR, "api_temperature_profiles.xlsx")

API_BASE = "https://archive-api.open-meteo.com/v1/archive"
ERA5_MODEL = "era5"           # 1940-present
TIMEZONE   = "Europe/London"  # GB local clock
# MONTH_SEASON / SEASONS / HOURS_PER_DAY come from seasons.py, shared with the demand model.

# 2 - QUERY SETUP
def _get_json(url, retries=4, backoff=5):
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=120) as r:
                return json.load(r)
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            if attempt == retries - 1:
                raise
            wait = backoff * (2 ** attempt)
            print(f"  request failed ({e}); retry {attempt+1}/{retries-1} in {wait}s", file=sys.stderr)
            time.sleep(wait)


def fetch_hourly_temps(lat, lon, start_year, end_year):
    """ERA5 hourly 2 m temperature for one point, pulled year-by-year over [start_year, end_year].
    Returns two parallel lists: ISO local timestamps and temperatures [°C]."""
    times, temps = [], []
    for yr in range(start_year, end_year + 1):
        params = {
            "latitude":  f"{lat:.4f}",
            "longitude": f"{lon:.4f}",
            "start_date": f"{yr}-01-01",
            "end_date":   f"{yr}-12-31",
            "hourly":     "temperature_2m",
            "timezone":   TIMEZONE,
            "models":     ERA5_MODEL,
        }
        url = f"{API_BASE}?{urllib.parse.urlencode(params)}"
        payload = _get_json(url)
        h = payload.get("hourly", {}) if isinstance(payload, dict) else {}
        yt, yv = h.get("time", []), h.get("temperature_2m", [])
        times.extend(yt)
        temps.extend(yv)
        print(f"    {yr}: {len(yt)} hourly records")
        time.sleep(0.25)   
    return times, temps


def build_profiles(times, temps):
    """Mean ERA5 temperature [°C] per (season, local hour), dropping nulls.
    Returns {season: [24 floats]}."""
    sums   = {(s, h): 0.0 for s in SEASONS for h in range(HOURS_PER_DAY)}
    counts = {(s, h): 0   for s in SEASONS for h in range(HOURS_PER_DAY)}
    for ts, t in zip(times, temps):
        if t is None:
            continue
        # ts format: "2016-01-01T00:00"
        month = int(ts[5:7])
        hour  = int(ts[11:13])
        season = MONTH_SEASON[month]
        sums[(season, hour)]   += float(t)
        counts[(season, hour)] += 1
    profiles = {}
    for s in SEASONS:
        profiles[s] = [round(sums[(s, h)] / counts[(s, h)], 3) if counts[(s, h)] else None
                       for h in range(HOURS_PER_DAY)]
    return profiles


# 3 - EXCEL SETUP
def write_workbook(district_profiles, start_year, end_year, path=TEMPERATURE_PROFILES_XLSX):
    """district_profiles: {district: {season: [24 floats]}}."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HourlyTemp"
    ws.cell(row=1, column=1, value=(f"ERA5 mean 2 m air temperature [°C] by district × season × local hour, "
            f"{start_year}-{end_year}. Generated by api_temperature_profiles.py; re-run to refresh."))
    hdr = 3
    cols = ["District", "Season"] + [f"h{h:02d}" for h in range(HOURS_PER_DAY)]
    for j, c in enumerate(cols, start=1):
        ws.cell(row=hdr, column=j, value=c)
    r = hdr
    for district in DISTRICTS:
        for s in SEASONS:
            r += 1
            ws.cell(row=r, column=1, value=district)
            ws.cell(row=r, column=2, value=s)
            for h in range(HOURS_PER_DAY):
                ws.cell(row=r, column=3 + h, value=district_profiles[district][s][h])
    for cell in ws[hdr]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10

    mws = wb.create_sheet("Meta")
    rows = [
        ("Parameter", "Value"),
        ("source", "Open-Meteo Historical Weather API (ERA5 reanalysis)"),
        ("source_url", API_BASE),
        ("model", ERA5_MODEL),
        ("variable", "temperature_2m"),
        ("timezone", TIMEZONE),
        ("window_start_year", start_year),
        ("window_end_year", end_year),
        ("seasons", "Winter=DJF, Spring=MAM, Summer=JJA, Autumn=SON"),
        ("pulled_utc", dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")),
        ("notes", "HourlyTemp holds absolute °C; the model uses anomaly = temp - 24h mean per "
                  "(district, season) added onto the HDD-derived daily-mean temperature. "
                  "Peak Winter maps to Winter (no own shape)."),
    ]
    for ri, (a, b) in enumerate(rows, start=1):
        mws.cell(row=ri, column=1, value=a)
        mws.cell(row=ri, column=2, value=b)
    mws.column_dimensions["A"].width = 20
    mws.column_dimensions["B"].width = 60
    mws["A1"].font = mws["B1"].font = Font(bold=True)

    os.makedirs(DATA_DIR, exist_ok=True)
    try:
        wb.save(path)
    except PermissionError as e:
        raise SystemExit(f"Could not save {path}: {e}")
    return path

def main(argv=None):
    last_complete = dt.date.today().year - 1   # current year is incomplete
    ap = argparse.ArgumentParser(description="Pull ERA5 hourly temperature -> data/api_temperature_profiles.xlsx")
    ap.add_argument("--years", type=int, default=10, help="number of complete years ending at last year (default 10)")
    ap.add_argument("--start", type=int, default=None, help="explicit start year (overrides --years)")
    ap.add_argument("--end",   type=int, default=last_complete, help=f"end year (default {last_complete})")
    ap.add_argument("--out",   default=TEMPERATURE_PROFILES_XLSX)
    args = ap.parse_args(argv)

    end_year   = args.end
    start_year = args.start if args.start is not None else end_year - args.years + 1

    print(f"Pulling ERA5 hourly temperature {start_year}-{end_year} for {len(DISTRICTS)} districts ...")
    district_profiles = {}
    for district in DISTRICTS:
        lat, lon = DISTRICT_LATLON[district]
        print(f"  {district} ({lat:.3f}, {lon:.3f}):")
        times, temps = fetch_hourly_temps(lat, lon, start_year, end_year)
        district_profiles[district] = build_profiles(times, temps)

    path = write_workbook(district_profiles, start_year, end_year, args.out)
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
