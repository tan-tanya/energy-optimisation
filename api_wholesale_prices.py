"""
Run-only.
Run: python api_wholesale_prices.py → data/api_wholesale_prices.xlsx.

Pulls GB wholesale electricity prices from the Elexon Insights Market Index Data (MID) dataset
(the market reference price) via the public REST API, and reduces a full calendar year to an 
average half-hourly day per season × day-type.

Output: data/api_wholesale_prices.xlsx —
  - "Shape":  48 half-hour rows × 8 columns (WD, WE for each season) and the mean MID price 
              [GBP/MWh] for that (season, day-type, settlement period) across the year.
  - "Meta":   provider, year, settlement-period day-count, annual mean, pull timestamp and source.

Seasons follow the model's calendar (must match demand_profile_model.MONTH_SEASON):
  Winter = Dec/Jan/Feb, Spring = Mar/Apr/May, Summer = Jun/Jul/Aug, Autumn = Sep/Oct/Nov.

Notes / caveats:
  - GB timezone; the two clock-change days have 46/50 settlement periods but only 1-48 is kept. 
  - Prices are nominal at time of trade; for a strict real-2025 model deflate the LEVEL if 
    mixing years (a single 2025 pull needs no deflation). 

Usage:
    python api_wholesale_prices.py                 # pull 2025, APXMIDP -> data/api_wholesale_prices.xlsx
    python api_wholesale_prices.py --year 2025 --provider APXMIDP
"""
import os
import sys
import time
import json
import argparse
import datetime as dt
import urllib.request
import urllib.error

import openpyxl
from openpyxl.styles import Font

# 1 - OVERALL SETUP
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
WHOLESALE_PRICES_XLSX = os.path.join(DATA_DIR, "api_wholesale_prices.xlsx")

API_BASE = "https://data.elexon.co.uk/bmrs/api/v1/datasets/MID"
DEFAULT_PROVIDER = "APXMIDP"
HH_PER_DAY = 48

# Must mirror demand_profile_model.MONTH_SEASON.
MONTH_SEASON = {1: "Winter", 2: "Winter", 3: "Spring", 4: "Spring", 5: "Spring", 6: "Summer",
                7: "Summer", 8: "Summer", 9: "Autumn", 10: "Autumn", 11: "Autumn", 12: "Winter"}
SEASONS   = ["Winter", "Spring", "Summer", "Autumn"]
DAY_TYPES = ["WD", "WE"]
COLS = [f"{s}_{d}" for s in SEASONS for d in DAY_TYPES]  


# 2 - QUERY SETUP
def _get_json(url, retries=4, backoff=5):
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=60) as r:
                return json.load(r)
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            if attempt == retries - 1:
                raise
            wait = backoff * (2 ** attempt)
            print(f"  request failed ({e}); retry {attempt+1}/{retries-1} in {wait}s", file=sys.stderr)
            time.sleep(wait)


MAX_WINDOW_DAYS = 7   

def fetch_mid_year(year, provider=DEFAULT_PROVIDER):
    """All MID records for `year` from one provider, pulled in <=7-day windows.
    Returns a list of {settlementDate, settlementPeriod, price, volume}."""
    out = []
    cur, end = dt.date(year, 1, 1), dt.date(year + 1, 1, 1)
    while cur < end:
        to = min(cur + dt.timedelta(days=MAX_WINDOW_DAYS), end)
        url = f"{API_BASE}?from={cur.isoformat()}&to={to.isoformat()}&format=json"
        payload = _get_json(url)
        recs = payload.get("data", payload) if isinstance(payload, dict) else payload
        kept = [r for r in recs if r.get("dataProvider") == provider]
        out.extend(kept)
        print(f"  {cur.isoformat()}..{to.isoformat()}: {len(recs)} rows, {len(kept)} {provider}")
        cur = to
        time.sleep(0.25)   # be polite to the public API
    return out


def build_profiles(records):
    """Mean MID price [GBP/MWh] per (season, day-type, settlement period), plus the day-weighted
    annual mean. Drops zero/dummy rows and out-of-range periods."""
    # 1) dedupe (date, sp) -> mean price (a date/period can carry >1 published row)
    by_key = {}
    for r in records:
        sp = r.get("settlementPeriod")
        price = r.get("price")
        if sp is None or price is None or not (1 <= sp <= HH_PER_DAY) or price <= 0:
            continue
        d = r.get("settlementDate")
        by_key.setdefault((d, sp), []).append(float(price))
    # 2) accumulate into (season, day-type, sp) buckets
    sums  = {(c, sp): 0.0 for c in COLS for sp in range(1, HH_PER_DAY + 1)}
    counts = {(c, sp): 0  for c in COLS for sp in range(1, HH_PER_DAY + 1)}
    grand_sum, grand_n = 0.0, 0
    for (d, sp), prices in by_key.items():
        date = dt.date.fromisoformat(d[:10])
        col = f"{MONTH_SEASON[date.month]}_{'WE' if date.weekday() >= 5 else 'WD'}"
        p = sum(prices) / len(prices)
        sums[(col, sp)]  += p
        counts[(col, sp)] += 1
        grand_sum += p
        grand_n   += 1
    profiles = {}
    for c in COLS:
        profiles[c] = [sums[(c, sp)] / counts[(c, sp)] if counts[(c, sp)] else None
                       for sp in range(1, HH_PER_DAY + 1)]
    annual_mean_mwh = grand_sum / grand_n if grand_n else float("nan")
    n_obs = {c: max(counts[(c, sp)] for sp in range(1, HH_PER_DAY + 1)) for c in COLS}
    return profiles, annual_mean_mwh, n_obs


# 3 - EXCEL SETUP
def write_workbook(profiles, annual_mean_mwh, n_obs, year, provider, path=WHOLESALE_PRICES_XLSX):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shape"
    ws.cell(row=1, column=1, value=f"GB wholesale (Elexon MID, {provider}) mean half-hourly price by season x day-type [GBP/MWh]")
    ws.cell(row=2, column=1, value=(f"Pulled for calendar {year}. Each cell = mean MID price over all days of that "
            "season/day-type at that settlement period. Seasonal level preserved (not normalised). "
            "Generated by wholesale_prices.py — do not hand-edit; re-run to refresh."))
    hdr = 4
    ws.cell(row=hdr, column=1, value="HH")
    ws.cell(row=hdr, column=2, value="Time")
    for j, c in enumerate(COLS, start=3):
        ws.cell(row=hdr, column=j, value=c)
    for i in range(HH_PER_DAY):
        r = hdr + 1 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=f"{(i//2):02d}:{'30' if i%2 else '00'}")
        for j, c in enumerate(COLS, start=3):
            v = profiles[c][i]
            ws.cell(row=r, column=j, value=round(v, 3) if v is not None else None)
    for cell in ws[hdr]:
        cell.font = Font(bold=True)

    mws = wb.create_sheet("Meta")
    rows = [
        ("Parameter", "Value"),
        ("source", "Elexon Insights - Market Index Data (MID)"),
        ("source_url", API_BASE),
        ("provider", provider),
        ("year", year),
        ("annual_mean_GBP_per_MWh", round(annual_mean_mwh, 4)),
        ("annual_mean_GBP_per_kWh", round(annual_mean_mwh / 1000.0, 6)),
        ("settlement_periods", HH_PER_DAY),
        ("pulled_utc", dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")),
        ("notes", "Shape sheet holds absolute GBP/MWh; model uses level=annual_mean and "
                  "shape=price/annual_mean (global mean 1.0). Periods 1-48 only."),
    ]
    for c, label in (("max_n_days_" + col, n_obs[col]) for col in COLS):
        rows.append((c, label))
    for ri, (a, b) in enumerate(rows, start=1):
        mws.cell(row=ri, column=1, value=a)
        mws.cell(row=ri, column=2, value=b)
    mws.column_dimensions["A"].width = 26
    mws["A1"].font = mws["B1"].font = Font(bold=True)

    os.makedirs(DATA_DIR, exist_ok=True)
    try:
        wb.save(path)
    except PermissionError as e:
        raise SystemExit(f"Could not save {path} (is it open in Excel?): {e}")
    return path


# 4 - LOADERS
_CACHE = None

def _load(path=WHOLESALE_PRICES_XLSX):
    global _CACHE
    if _CACHE is not None:
        return _CACHE
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Wholesale price data not found at {path}. Run 'python wholesale_prices.py' first to generate it."
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Shape"]
    # find header row containing 'HH'
    hdr, scols = None, None
    for rr in range(1, 8):
        row = [ws.cell(row=rr, column=c).value for c in range(1, ws.max_column + 1)]
        if "HH" in row:
            hdr = rr
            scols = {v: i + 1 for i, v in enumerate(row) if v}
            break
    profiles = {}
    for name, ci in scols.items():
        if name in ("HH", "Time"):
            continue
        col = []
        for rr in range(hdr + 1, ws.max_row + 1):
            v = ws.cell(row=rr, column=ci).value
            if isinstance(v, (int, float)):
                col.append(float(v))
        if len(col) >= HH_PER_DAY:
            profiles[name] = col[:HH_PER_DAY]
    meta = {}
    if "Meta" in wb.sheetnames:
        mws = wb["Meta"]
        for rr in range(1, mws.max_row + 1):
            k = mws.cell(row=rr, column=1).value
            v = mws.cell(row=rr, column=2).value
            if k is not None:
                meta[str(k)] = v
    wb.close()
    _CACHE = {"profiles": profiles, "meta": meta}
    return _CACHE


def load_level(path=WHOLESALE_PRICES_XLSX):
    """Annual mean GB wholesale price [GBP/kWh] — the wholesale leg's level."""
    meta = _load(path)["meta"]
    if "annual_mean_GBP_per_kWh" in meta and meta["annual_mean_GBP_per_kWh"] is not None:
        return float(meta["annual_mean_GBP_per_kWh"])
    # fallback: derive from the absolute Shape profiles
    profs = _load(path)["profiles"]
    vals = [v for col in profs.values() for v in col]
    return (sum(vals) / len(vals)) / 1000.0


def load_shape(path=WHOLESALE_PRICES_XLSX):
    """Normalised half-hourly multipliers per '{Season}_{WD|WE}' (GLOBAL mean 1.0, so per-season means
    differ — keeps seasonal level + intra-day shape). Returns {col: [48 floats]}."""
    data = _load(path)
    profs = data["profiles"]
    mean_mwh = load_level(path) * 1000.0
    if not mean_mwh:
        return {c: [1.0] * HH_PER_DAY for c in profs}
    return {c: [v / mean_mwh for v in col] for c, col in profs.items()}


def main(argv=None):
    ap = argparse.ArgumentParser(description="Pull Elexon MID wholesale prices -> data/api_wholesale_prices.xlsx")
    ap.add_argument("--year", type=int, default=2025)
    ap.add_argument("--provider", default=DEFAULT_PROVIDER, help="MID data provider (APXMIDP or N2EXMIDP)")
    ap.add_argument("--out", default=WHOLESALE_PRICES_XLSX)
    args = ap.parse_args(argv)

    print(f"Pulling Elexon MID for {args.year} (provider {args.provider}) ...")
    records = fetch_mid_year(args.year, args.provider)
    print(f"  total {args.provider} rows: {len(records)}")
    profiles, annual_mean_mwh, n_obs = build_profiles(records)
    print(f"  annual mean: {annual_mean_mwh:.2f} GBP/MWh = {annual_mean_mwh/1000:.5f} GBP/kWh")
    path = write_workbook(profiles, annual_mean_mwh, n_obs, args.year, args.provider, args.out)
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
