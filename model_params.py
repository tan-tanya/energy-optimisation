"""
Import-only, by demand_profile_model, optimisation_model, optimisation_report. 
Loads data/model_parameters.xlsx and reconstructs the parameter structures.

Each sheet has its header on row 4 and data from row 5. Column A is either "Line" (flat, scalar sheets) 
or an index column (Scope / Technology / Month — per-activity / per-tech / per-month sheets). 
The loader keys rows by the "Parameter" column and, where present, the column-A index value.
"""
import os
import re

import numpy as np
import openpyxl

PARAMS_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "data", "model_parameters.xlsx")


def _num(v):
    # Whole-number floats -> int (so year/index params can index lists).
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)) and float(v).is_integer():
        return int(v)
    return v


def _find_header_row(ws, max_scan=12):
    # Locate the header row containing both 'Parameter' and 'Value', loader robust to additional or missing rows.
    for r in range(1, min(max_scan, ws.max_row) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if "Parameter" in vals and "Value" in vals:
            return r
    raise ValueError(f"No header row (with 'Parameter' and 'Value') found in sheet {ws.title!r}")


def _read_rows(ws):
    hr = _find_header_row(ws)
    headers = [ws.cell(row=hr, column=c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(hr + 1, ws.max_row + 1):
        d = {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers)}
        if d.get("Parameter") is None:
            continue
        d["_value"] = _num(d.get("Value"))
        rows.append(d)
    return rows


def _is_active(d):
    # True if the row's Active cell flags it as the selected scenario.
    return str(d.get("Active") or "").strip().lower() in ("yes", "y", "true", "1", "x")


def load_params(path=PARAMS_XLSX):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Parameter workbook not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    R = {ws.title: _read_rows(ws) for ws in wb.worksheets if ws.title != "Notes"}
    wb.close()

    # REAL-TERMS REBASING
    # The model is a real-terms appraisal, but cost inputs were sourced in different real base years. 
    # Any row carrying a 'Price Base Year' is rebased to TARGET_PRICE_YEAR using the HM Treasury UK GDP deflator, 
    # factor = deflator[target]/deflator[base].
    # This is not forward inflation (real prices stay flat over the horizon; the real discount rate handles time value).
    TARGET_PRICE_YEAR = 2025
    defl_re = re.compile(r"^GDP_DEFLATOR \[(\d{4})\]$")
    gdp_deflator = {}
    for d in R.get("Scalars", []):
        m = defl_re.match(str(d.get("Parameter")))
        if m and d["_value"] not in (None, ""):
            gdp_deflator[int(m.group(1))] = float(d["_value"])
    if gdp_deflator:
        f_target = gdp_deflator[TARGET_PRICE_YEAR]
        for rows in R.values():
            for d in rows:
                by, v = d.get("Price Base Year"), d.get("_value")
                if by in (None, "") or not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                by = int(by)
                if by != TARGET_PRICE_YEAR:
                    d["_value"] = v * f_target / gdp_deflator[by]

    def flat(sheet):
        return {d["Parameter"]: d["_value"] for d in R[sheet]}

    # optimisation_model.py
    scal = flat("Scalars")

    # Drop the GDP_DEFLATOR [YYYY] reference rows out of scal.
    for key in [k for k in scal if defl_re.match(str(k))]:
        scal.pop(key)

    # DESNZ Reference non-heat electricity demand series, 'ELEC_DEMAND_TWH [YYYY]' rows (transcribed from Annex F). 
    edem_re = re.compile(r"^ELEC_DEMAND_TWH \[(\d{4})\]$")
    elec_demand_twh = {}
    for key in [k for k in scal if edem_re.match(str(k))]:
        m = edem_re.match(str(key))
        elec_demand_twh[int(m.group(1))] = float(scal.pop(key))

    # Per-district grid ceilings: 18 'GRID_{IMPORT,EXPORT}_LIMIT_KW [<district>]' rows on the Scalars sheet (DNO network-headroom). 
    grid_re = re.compile(r"^GRID_(IMPORT|EXPORT)_LIMIT_KW \[(.+)\]$")
    grid_limits = {}
    for key in list(scal.keys()):
        m = grid_re.match(str(key))
        if not m:
            continue
        val = scal.pop(key)
        if val in (None, ""):
            continue
        direction, district = m.group(1).lower(), m.group(2)
        grid_limits.setdefault(district, {})[f"{direction}_kw"] = float(val)

    # Primary energy factors: 'PRIMARY_ENERGY_FACTOR [<carrier>]' rows on the Scalars sheet
    # (NCM Modelling Guide, kWh primary / kWh delivered). Reporting only - not in OF1 or OF2.
    pef_re = re.compile(r"^PRIMARY_ENERGY_FACTOR \[(.+)\]$")
    primary_energy_factors = {}
    for key in list(scal.keys()):
        m = pef_re.match(str(key))
        if not m:
            continue
        val = scal.pop(key)
        if val in (None, ""):
            continue
        primary_energy_factors[m.group(1)] = float(val)

    # PV: single-value rows, plus const_per_kwp + infra_per_kwp as Low/Central/High scenarios (DESNZ).
    pv, pv_const, pv_infra = {}, {}, {}
    pv_keyed, pv_active = {}, {}            # other Low/Central/High scalars (e.g. maint_per_mw_per_yr)
    for d in R["PV"]:
        param, key = d["Parameter"], d.get("Key")
        if key in (None, ""):
            pv[param] = d["_value"]
            continue
        if param == "const_per_kwp":
            pv_const[key] = d["_value"]
        elif param == "infra_per_kwp":
            pv_infra[key] = d["_value"]
        else:
            pv_keyed.setdefault(param, {})[key] = d["_value"]
        if _is_active(d):
            pv_active[param] = key
    # capex scenario: const + infra share one DESNZ Low/Central/High switch
    pv_capex_active = pv_active.get("const_per_kwp") or pv_active.get("infra_per_kwp") or "Central"
    pv_capex_scen = {s: pv_const.get(s, 0.0) + pv_infra.get(s, 0.0) for s in set(pv_const) | set(pv_infra)}
    pv["const_per_kwp"] = pv_const[pv_capex_active]
    pv["infra_per_kwp"] = pv_infra[pv_capex_active]
    pv["capex_per_kwp"] = pv["const_per_kwp"] + pv["infra_per_kwp"]
    # any other keyed scalar (e.g. maint_per_mw_per_yr): take its Active row, else Central
    for param, scen in pv_keyed.items():
        pv[param] = scen.get(pv_active.get(param, "Central"), next(iter(scen.values())))

    batt = flat("Battery")
    batt["chg_eff"] = batt["disc_eff"] = float(np.sqrt(batt["round_trip_eff"]))   # derived
    tech_costs = {**scal, "pv": pv, "battery": batt}

    # Electricity & gas prices live on one 'Energy Prices' sheet, split by the Group column:
    #   Electricity Import    -> size-band import price (auto-selected by demand; Min/Max MWh thresholds)
    #   Gas Import            -> size-band gas price    (auto-selected by demand; Min/Max MWh thresholds)
    #   Electricity Export    -> SEG export price scenarios (Active-flagged)
    #   Energy Price Growth   -> real elec/gas escalation scenarios (Active-flagged)
    #   Import Price Scenario -> Low/Central/High level multipliers (x Central 2025 band price)
    #                            for elec & gas imports (Active-flagged). This is a present-day level band 
    #                            (independent of the Energy Price Growth scenario, so the two may be set separately). 
    #                            Central = 1.0 leaves the DESNZ 2025 band actuals unchanged.
    elec_bands, gas_bands = [], []
    export_scen, export_active = {}, None
    growth_scen, growth_active = {}, None
    import_mult_scen, import_active = {}, None
    for d in R["Energy Prices"]:
        group = d.get("Group")
        if group in ("Electricity Import", "Gas Import"):
            lo, hi = d.get("Min (MWh/yr)"), d.get("Max (MWh/yr)")
            band = {"name": d["Key"],
                    "lo": float(lo) if lo is not None else 0.0,
                    "hi": float(hi) if hi is not None else float("inf"),
                    "price": float(d["_value"])}
            (elec_bands if group == "Electricity Import" else gas_bands).append(band)
        elif group == "Electricity Export":
            export_scen[d["Key"]] = float(d["_value"])
            if _is_active(d): export_active = export_active or d["Key"]
        elif group == "Energy Price Growth":
            growth_scen.setdefault(d["Key"], {})[d["Parameter"]] = float(d["_value"])
            if _is_active(d): growth_active = growth_active or d["Key"]
        elif group == "Import Price Scenario":
            import_mult_scen.setdefault(d["Key"], {})[d["Parameter"]] = float(d["_value"])
            if _is_active(d): import_active = import_active or d["Key"]
    elec_bands.sort(key=lambda b: b["lo"])
    gas_bands.sort(key=lambda b: b["lo"])
    export_active = export_active or "Central"
    growth_active = growth_active or "Central"
    tech_costs["elec_export_price"] = export_scen[export_active]
    tech_costs["elec_price_growth"] = growth_scen[growth_active]["elec_price_growth"]
    tech_costs["gas_price_growth"]  = growth_scen[growth_active]["gas_price_growth"]

    # Apply the active import-price level multiplier to every size band's price.
    import_active = import_active or "Central"
    imp = import_mult_scen.get(import_active, {})
    elec_import_mult = float(imp.get("elec_import_multiplier", 1.0))
    gas_import_mult  = float(imp.get("gas_import_multiplier", 1.0))
    for b in elec_bands: b["price"] *= elec_import_mult
    for b in gas_bands:  b["price"] *= gas_import_mult
    tech_costs["elec_import_multiplier"] = elec_import_mult
    tech_costs["gas_import_multiplier"]  = gas_import_mult

    heat_costs = {}
    for d in R["Heating"]:
        heat_costs.setdefault(d["Scope"], {})[d["Parameter"]] = d["_value"]

    thermal = flat("Thermal Store")

    roof_props, roof_load, roof_pitch, pitched_frac = {}, None, None, None
    for d in R["Roof"]:
        p, v = d["Parameter"], d["_value"]
        if   p in ("pv_usable_frac", "pv_inter_row_frac"):
            roof_props.setdefault(d["Scope"], {})[p] = v
        elif p == "ROOF_LOAD_KG_PER_M2":       roof_load = v
        elif p == "ROOF_PITCH_DEG":            roof_pitch = float(v)
        elif p == "PITCHED_USABLE_SLOPE_FRAC": pitched_frac = v

    # Per-district daily GHI: nested {district: {month: kWh/m2/day}}
    monthly_ghi = {}
    for d in R["GHI"]:
        monthly_ghi.setdefault(d["District"], {})[d["Month"]] = d["_value"]

    # Emissions: operational carbon factors + appraisal carbon values (DESNZ Green Book data-tables-1-19).
    #   Gas Emission Factor          natural-gas combustion [kgCO2e/kWh, gross CV], constant.
    #   Electricity Emission Factor  grid intensity [kgCO2e/kWh] per calendar year (LRM, consumption-based,
    #                                Commercial/Public); trajectory reflects decarbonising grid.
    #   Carbon Value                 appraisal carbon value [GBP(2022)/tCO2e] per year, active sensitivity only.
    gas_ef, elec_ef, carbon_val = None, {}, {}
    for d in R["Emissions"]:
        group, yr, v = d.get("Group"), d.get("Year"), d["_value"]
        if v in (None, ""):
            continue
        if group == "Gas Emission Factor":
            gas_ef = float(v)
        elif group == "Electricity Emission Factor":
            elec_ef[int(yr)] = float(v)
        elif group == "Carbon Value" and _is_active(d):
            carbon_val[int(yr)] = float(v)

    # demand_profile_model.py
    heatcop = flat("Heating & COP (DM)")
    ground  = flat("Ground Loop (DM)")

    ev_park, ev_dwell, ev_scalar = {}, {}, {}
    for d in R["EV (DM)"]:
        p, v = d["Parameter"], d["_value"]
        if   p == "EV_PARKING_DENSITY": ev_park[d["Scope"]] = v
        elif p == "EV_DWELL_HOURS":     ev_dwell[d["Scope"]] = v
        else:                           ev_scalar[p] = v

    clim, diurnal, season_doy, we_load = {}, {}, {}, {}
    for d in R["Climate & Demand (DM)"]:
        p, v = d["Parameter"], d["_value"]
        if   p == "DIURNAL_AMPLITUDE": diurnal[d["Scope"]] = float(v)
        elif p == "SEASON_DOY":        season_doy[d["Scope"]] = v
        elif p == "WE_LOAD_FACTOR":    we_load[d["Scope"]] = float(v)
        else:                          clim[p] = v

    return {
        # optimisation_model.py
        "TECH_COSTS":                tech_costs,
        "HEAT_COSTS":                heat_costs,
        "THERMAL_STORE":             thermal,
        "ROOF_PROPERTIES":           roof_props,
        "ROOF_LOAD_KG_PER_M2":       roof_load,
        "ROOF_PITCH_DEG":            roof_pitch,
        "PITCHED_USABLE_SLOPE_FRAC": pitched_frac,
        "DISTRICT_MONTHLY_GHI":      monthly_ghi,
        "GAS_EMISSION_FACTOR":       gas_ef,
        "ELEC_EMISSION_FACTORS":     elec_ef,
        "CARBON_VALUES":             carbon_val,
        "GRID_LIMITS":               grid_limits,
        "PRIMARY_ENERGY_FACTORS":    primary_energy_factors,
        "ELEC_IMPORT_BANDS":         elec_bands,
        "GAS_IMPORT_BANDS":          gas_bands,
        "ELEC_EXPORT_SCENARIOS":     export_scen,
        "ELEC_EXPORT_ACTIVE":        export_active,
        "ENERGY_GROWTH_SCENARIOS":   growth_scen,
        "ENERGY_GROWTH_ACTIVE":      growth_active,
        "IMPORT_PRICE_SCENARIOS":    import_mult_scen,
        "IMPORT_PRICE_ACTIVE":       import_active,
        "GDP_DEFLATOR":              gdp_deflator,
        "ELEC_DEMAND_TWH":           elec_demand_twh,
        "PRICE_BASE_YEAR":           TARGET_PRICE_YEAR,
        "PV_CAPEX_SCENARIOS":        pv_capex_scen,
        "PV_CAPEX_ACTIVE":           pv_capex_active,
        # demand_profile_model.py
        "ETA_BOILER":                    float(heatcop["ETA_BOILER"]),
        "ashp_cop_at_7C":                float(heatcop["ashp_cop_at_7C"]),
        "ashp_cop_slope_per_C":          float(heatcop["ashp_cop_slope_per_C"]),
        "gshp_cop_at_10C":               float(heatcop["gshp_cop_at_10C"]),
        "gshp_cop_slope_per_C":          float(heatcop["gshp_cop_slope_per_C"]),
        "T_ASHP_MIN": heatcop["T_ASHP_MIN"], "T_ASHP_MAX": heatcop["T_ASHP_MAX"],
        "T_GSHP_MIN": heatcop["T_GSHP_MIN"], "T_GSHP_MAX": heatcop["T_GSHP_MAX"],
        "HORIZONTAL_LOOP_DEPTH_M":       float(ground["HORIZONTAL_LOOP_DEPTH_M"]),
        "SOIL_THERMAL_DIFFUSIVITY_M2_S": float(ground["SOIL_THERMAL_DIFFUSIVITY_M2_S"]),
        "SURFACE_TEMP_PEAK_DOY":         ground["SURFACE_TEMP_PEAK_DOY"],
        "BRINE_OFFSET_C":                float(ground["BRINE_OFFSET_C"]),
        "DEFAULT_GROUND_TEMP_C":         float(ground["DEFAULT_GROUND_TEMP_C"]),
        # Horizontal-loop land constraint (vertical boreholes need negligible surface area).
        "HORIZONTAL_COLLECTOR_M2_PER_KWTH": float(ground["HORIZONTAL_COLLECTOR_M2_PER_KWTH"]),
        "SITE_PLOT_RATIO":                  float(ground["SITE_PLOT_RATIO"]),
        "PARKING_GROSS_M2_PER_SPACE":       float(ground["PARKING_GROSS_M2_PER_SPACE"]),
        "EV_CHARGER_KW":                 float(ev_scalar["EV_CHARGER_KW"]),
        "EV_SPACE_FRACTION":             float(ev_scalar["EV_SPACE_FRACTION"]),
        "EV_PARKING_DENSITY":            ev_park,
        "EV_DWELL_HOURS":                ev_dwell,
        "PEAK_FRACTION":                 float(clim["PEAK_FRACTION"]),
        "T_SETBACK":                     float(clim["T_SETBACK"]),
        "HDD_BASE":                      float(clim["HDD_BASE"]),
        "UK_LATITUDE":                   float(clim["UK_LATITUDE"]),
        "DIURNAL_AMPLITUDE":             diurnal,
        "SEASON_DOY":                    season_doy,
        "WE_LOAD_FACTOR":                we_load,
    }


__all__ = [
    "load_params", "PARAMS_XLSX",
    "TECH_COSTS", "HEAT_COSTS", "THERMAL_STORE", "ROOF_PROPERTIES",
    "ROOF_LOAD_KG_PER_M2", "ROOF_PITCH_DEG", "PITCHED_USABLE_SLOPE_FRAC",
    "DISTRICT_MONTHLY_GHI", "GRID_LIMITS", "select_grid_limit",
    "PRIMARY_ENERGY_FACTORS",
    "GAS_EMISSION_FACTOR", "ELEC_EMISSION_FACTORS", "CARBON_VALUES",
    "EMISSIONS_BASE_YEAR", "elec_emission_factor", "carbon_value",
    "ELEC_IMPORT_BANDS", "select_elec_band",
    "GAS_IMPORT_BANDS", "select_gas_band",
    "ELEC_EXPORT_SCENARIOS", "ELEC_EXPORT_ACTIVE",
    "ENERGY_GROWTH_SCENARIOS", "ENERGY_GROWTH_ACTIVE",
    "IMPORT_PRICE_SCENARIOS", "IMPORT_PRICE_ACTIVE",
    "GDP_DEFLATOR", "PRICE_BASE_YEAR", "ELEC_DEMAND_TWH",
    "PV_CAPEX_SCENARIOS", "PV_CAPEX_ACTIVE",
    "ETA_BOILER", "ashp_cop_at_7C", "ashp_cop_slope_per_C",
    "gshp_cop_at_10C", "gshp_cop_slope_per_C",
    "T_ASHP_MIN", "T_ASHP_MAX", "T_GSHP_MIN", "T_GSHP_MAX",
    "HORIZONTAL_COLLECTOR_M2_PER_KWTH", "SITE_PLOT_RATIO", "PARKING_GROSS_M2_PER_SPACE",
    "HORIZONTAL_LOOP_DEPTH_M", "SOIL_THERMAL_DIFFUSIVITY_M2_S", "SURFACE_TEMP_PEAK_DOY",
    "BRINE_OFFSET_C", "DEFAULT_GROUND_TEMP_C",
    "EV_CHARGER_KW", "EV_SPACE_FRACTION", "EV_PARKING_DENSITY", "EV_DWELL_HOURS",
    "PEAK_FRACTION", "T_SETBACK", "HDD_BASE", "UK_LATITUDE",
    "DIURNAL_AMPLITUDE", "SEASON_DOY", "WE_LOAD_FACTOR",
]

# Bind every parameter as a module attribute so callers can import specific attributes from the module
globals().update(load_params())


# Optimisation representative year is 2025 (WD/WE day counts, prices); horizon year index y=0 maps to 2025. 
EMISSIONS_BASE_YEAR = 2025


def _series_at_year(series, year_index):
    # Value of a {calendar_year: x} series at horizon-year index y; clamps to the table's range
    # (factors are flat-extrapolated beyond the last published year).
    if not series:
        raise KeyError("Emissions series is empty — check the 'Emissions' sheet loaded.")
    yr = EMISSIONS_BASE_YEAR + int(year_index)
    yr = min(max(yr, min(series)), max(series))
    return series[yr]


def elec_emission_factor(year_index):
    # Grid electricity emission factor [kgCO2e/kWh] at horizon-year index y (LRM, consumption-based).
    return _series_at_year(ELEC_EMISSION_FACTORS, year_index)


def carbon_value(year_index):
    # Appraisal carbon value [GBP(2022)/tCO2e] at horizon-year index y (active sensitivity).
    return _series_at_year(CARBON_VALUES, year_index)


def select_grid_limit(district):
    # DNO import / export connection ceilings (kW) for a district, from the Scalars-sheet
    # GRID_{IMPORT,EXPORT}_LIMIT_KW [<district>] rows. Returns {"import_kw", "export_kw"}.
    d = GRID_LIMITS.get(district)
    if d is None:
        raise KeyError(
            f"No grid connection limits for district {district!r}. Add "
            f"'GRID_IMPORT_LIMIT_KW [{district}]' and 'GRID_EXPORT_LIMIT_KW [{district}]' to the Scalars sheet."
        )
    missing = [s for s in ("import_kw", "export_kw") if s not in d]
    if missing:
        raise KeyError(
            f"District {district!r} is missing grid limit value(s) {missing} — fill the blank "
            f"GRID_*_LIMIT_KW [{district}] cell(s) on the Scalars sheet."
        )
    return {"import_kw": d["import_kw"], "export_kw": d["export_kw"]}


def select_elec_band(annual_mwh):
    # DESNZ non-domestic electricity size band for a given gross annual electricity demand (MWh/yr).
    # Returns the band dict {name, lo, hi, price}.
    return _select_band(ELEC_IMPORT_BANDS, annual_mwh)


def select_gas_band(annual_mwh):
    # DESNZ non-domestic gas size band for a given gross annual gas demand (MWh/yr).
    # Returns the band dict {name, lo, hi, price}.
    return _select_band(GAS_IMPORT_BANDS, annual_mwh)


def _select_band(bands, annual_mwh):
    # First band whose [lo, hi) covers annual_mwh; else the highest band whose lower bound is satisfied
    # (handles the 1-MWh gaps between DESNZ gas thresholds and demand above the top band). bands sorted by lo.
    for b in bands:
        if b["lo"] <= annual_mwh < b["hi"]:
            return b
    return next((b for b in reversed(bands) if b["lo"] <= annual_mwh), bands[0])
