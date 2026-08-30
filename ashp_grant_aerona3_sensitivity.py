"""ASHP technology sensitivity. 
Runs as part of the pipeline (optimisation_model.main() calls run_ashp_sensitivity() after the cost rounds),
or standalone against the most recent completed run:
    python ashp_grant_aerona3_sensitivity.py [--limit N] [--time-limit S] [--jobs N]
Note: Running standalone against an earlier run predating a sizing bound change makes every row infeasible. 

Compares the model's current ASHP COP against a real UK product -- the Grant Aerona3 R32 (COP @ A7/W55 = 2.66) 
-- for every ASHP design in the latest deterministic cost-round optimisation run.

Method: 
For each (district, activity) design where ASHP is the winning heating system, first-stage sizing (n_pv, e_batt, 
o_batt, q_heat_cap, e_th) is fixed at the saved deterministic solution. Only the stage-2 dispatch is re-solved with:
  - ashp_cop_at_7C lowered to Grant Aerona3's 2.66 (COP/temperature SLOPE held at the model's
    existing value; cold-weather sensitivity assumed identical to the current model).
  - HEAT_COSTS["ASHP"]["capex_per_kwth"] cut by CAPEX_REDUCTION_FRAC, paired with the COP cut (~2.5%).
Baseline (current-COP) metrics are read straight from the deterministic sweep rows rather than re-solved.

Output: <run folder>/ashp_grant_aerona3_sensitivity.xlsx, one row per ASHP design, 
in the same NPV-savings-ranked row order as the source deterministic sweep.
"""
import argparse
import glob
import multiprocessing as mp
import os

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import demand_profile_model as dm
import optimisation_engine as oe
from optimisation_config import resolve_jobs

# Grant Aerona3 R32 vs the model's current Company-4-R32-average.
CURRENT_ASHP_COP_AT_7C = 2.7283
GRANT_COP_AT_7C        = 2.66
CAPEX_REDUCTION_FRAC   = 0.026   # paired capex cut, set alongside the 2.50% COP drop (2.66/2.7283)

RESULT_COLS = ["status", "q_heat_cap_kwth", "annual_import_kwh", "annual_export_kwh",
               "capex_GBP", "opex_npv_GBP", "total_cost_npv_GBP", "npv_savings_GBP",
               "lifetime_emissions_tco2e"]

BASE_RENAME = {
    "annual_import_kwh": "base_annual_import_kwh", "annual_export_kwh": "base_annual_export_kwh",
    "capex_GBP": "base_capex_GBP", "opex_npv_GBP": "base_opex_npv_GBP",
    "total_cost_npv_GBP": "base_total_cost_npv_GBP", "npv_savings_GBP": "base_npv_savings_GBP",
    "lifetime_emissions_tco2e": "base_lifetime_emissions_tco2e",
}

DISPLAY_COLS = [
    "district", "activity", "grant_status",
    "base_total_cost_npv_GBP", "grant_total_cost_npv_GBP", "total_cost_npv_GBP_delta",
    "base_lifetime_emissions_tco2e", "grant_lifetime_emissions_tco2e", "lifetime_emissions_tco2e_delta",
    "cheap_ashp_pays_off",
]

_RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_FONT   = Font(color="9C0006")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(color="006100")

DELTA_COLS = ["total_cost_npv_GBP_delta", "lifetime_emissions_tco2e_delta"]
BOOL_COL = "cheap_ashp_pays_off"


def apply_conditional_formatting(ws, columns: list, n_rows: int) -> None:
    for col_name in DELTA_COLS:
        col_letter = get_column_letter(columns.index(col_name) + 1)
        cell_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="greaterThan", formula=["0"], fill=_RED_FILL, font=_RED_FONT))
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="lessThan", formula=["0"], fill=_GREEN_FILL, font=_GREEN_FONT))

    col_letter = get_column_letter(columns.index(BOOL_COL) + 1)
    cell_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=["FALSE"], fill=_RED_FILL, font=_RED_FONT))
    ws.conditional_formatting.add(cell_range, CellIsRule(
        operator="equal", formula=["TRUE"], fill=_GREEN_FILL, font=_GREEN_FONT))


def _latest_run_dir() -> str:
    root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
    runs = glob.glob(os.path.join(root, "Optimisation (*)"))
    if not runs:
        raise FileNotFoundError(f"No 'Optimisation (*)' run folders found under {root}")
    return max(runs, key=os.path.getmtime)


def _read_deterministic_rows(run_dir: str) -> pd.DataFrame:
    # The deterministic sweep rows from the workbook written at the end of a run.
    wb = os.path.join(run_dir, "Optimisation Results (deterministic).xlsx")
    if not os.path.exists(wb):
        raise FileNotFoundError(f"No deterministic workbook in {run_dir}.")
    return pd.read_excel(wb, sheet_name="NPV Data")


def _pool_init():
    # Runs once per worker process.
    dm.initialize()
    dm.ashp_cop_at_7C = GRANT_COP_AT_7C   # slope (dm.ashp_cop_slope_per_C) left untouched, per direction
    base_capex = oe.HEAT_COSTS["ASHP"]["capex_per_kwth"]
    oe.HEAT_COSTS["ASHP"]["capex_per_kwth"] = base_capex * (1.0 - CAPEX_REDUCTION_FRAC)


def _solve_one(task: tuple) -> dict:
    district, activity, n_pv, e_batt, o_batt, q_heat_cap, e_th, threads, time_limit_s = task
    prob, V = oe.build_lp(district, activity, "ASHP",
                            objective="cost", scenarios=[oe.central_scenario()])
    prob += V["n_pv"]       == n_pv,       "fix_n_pv"
    prob += V["e_batt"]     == e_batt,     "fix_e_batt"
    prob += V["o_batt"]     == o_batt,     "fix_o_batt"
    # Flooring instead of equalities to prevent infeasibility due to rounding.
    prob += V["q_heat_cap"] >= q_heat_cap, "floor_q_heat_cap"
    prob += V["e_th"]       >= e_th,       "floor_e_th"
    solver = oe._make_solver(solver_msg=False, time_limit_s=time_limit_s, threads=threads)
    status = prob.solve(solver)
    res = oe._extract_results(prob, V, status, district, activity, "ASHP", oe.HORIZON_YEARS)
    return {"district": district, "activity": activity,
            **{f"grant_{c}": res.get(c) for c in RESULT_COLS}}


def run_ashp_sensitivity(all_rows: pd.DataFrame, run_dir: str, *, time_limit_s: int = 600,
                         n_jobs: int = None, limit: int = None) -> pd.DataFrame:
    """Called by optimisation_model.main() with the fresh deterministic cost-round DataFrame, 
    and by this module's own CLI with the same frame read back from the deterministic "NPV Data" sheet.
    Passing the in-memory frame makes this safe to run in-pipeline.
    """
    print(f"Grant Aerona3 COP@7C={GRANT_COP_AT_7C} vs current {CURRENT_ASHP_COP_AT_7C} "
          f"({(GRANT_COP_AT_7C / CURRENT_ASHP_COP_AT_7C - 1) * 100:+.2f}%); "
          f"capex cut applied: {CAPEX_REDUCTION_FRAC * 100:.1f}%")

    base = all_rows[(all_rows["status"] == "Optimal") & (all_rows["heating"] == "ASHP")].reset_index(drop=True)
    base["_rank"] = base.index   # preserves the source file's NPV-savings ranking through the merge
    if limit:
        base = base.iloc[:limit].copy()
    if base.empty:
        print("No Optimal ASHP designs to test — skipping Grant Aerona3 sensitivity.")
        return pd.DataFrame()
    print(f"{len(base)} ASHP designs Optimal (of {len(all_rows)} total rows) -- "
          f"solving Grant Aerona3 dispatch for each")

    # Same worker/thread sizing rule as the rest of the pipeline.
    n_jobs, threads_per_worker = resolve_jobs(n_jobs, len(base))

    tasks = [(r.district, r.activity, r.n_pv, r.e_batt_kwh, r.o_batt_kw,
              r.q_heat_cap_kwth, r.e_th_kwh, threads_per_worker, time_limit_s)
             for r in base.itertuples()]

    print(f"Solving {len(tasks)} designs across {n_jobs} processes x {threads_per_worker} solver threads...")
    with mp.Pool(processes=n_jobs, initializer=_pool_init) as pool:
        grant_rows = []
        for i, row in enumerate(pool.imap(_solve_one, tasks), 1):
            print(f"  [{i:>3}/{len(tasks)}] {row['activity']!r:27s} {row['district']!r:28s} "
                  f"{row['grant_status']}")
            grant_rows.append(row)
    grant_df = pd.DataFrame(grant_rows)

    base_ren = base.rename(columns=BASE_RENAME)
    keep = ["district", "activity", "_rank", "n_pv", "e_batt_kwh", "o_batt_kw",
            "q_heat_cap_kwth", "e_th_kwh"] + list(BASE_RENAME.values())
    out = base_ren[keep].merge(grant_df, on=["district", "activity"], how="left")
    out = out.sort_values("_rank").drop(columns="_rank").reset_index(drop=True)

    out["annual_import_kwh_delta"]        = out["grant_annual_import_kwh"] - out["base_annual_import_kwh"]
    out["annual_export_kwh_delta"]        = out["grant_annual_export_kwh"] - out["base_annual_export_kwh"]
    out["capex_GBP_delta"]                = out["grant_capex_GBP"] - out["base_capex_GBP"]
    out["opex_npv_GBP_delta"]             = out["grant_opex_npv_GBP"] - out["base_opex_npv_GBP"]
    out["total_cost_npv_GBP_delta"]       = out["grant_total_cost_npv_GBP"] - out["base_total_cost_npv_GBP"]
    out["lifetime_emissions_tco2e_delta"] = out["grant_lifetime_emissions_tco2e"] - out["base_lifetime_emissions_tco2e"]
    # NA (not False) when Grant Aerona3's fixed-sizing dispatch is Infeasible
    # (the lower COP pushes required grid import over the DNO limit)
    is_optimal = out["grant_status"] == "Optimal"
    out["cheap_ashp_pays_off"] = pd.Series(pd.NA, index=out.index, dtype="boolean")
    out.loc[is_optimal, "cheap_ashp_pays_off"] = out.loc[is_optimal, "total_cost_npv_GBP_delta"] < 0

    out_path = os.path.join(run_dir, "ashp_grant_aerona3_sensitivity.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        out[DISPLAY_COLS].to_excel(xw, sheet_name="Grant Aerona3 vs current", index=False)
        apply_conditional_formatting(xw.sheets["Grant Aerona3 vs current"], DISPLAY_COLS, len(out))
        notes = pd.DataFrame({
            "Note": ["Source run", "Current model ASHP COP@7C", "Grant Aerona3 R32 COP@A7/W55",
                     "COP/temperature slope", "Capex reduction applied", "Method"],
            "Value": [
                run_dir,
                f"{CURRENT_ASHP_COP_AT_7C} (CEP Technology Library v1.02, Company 4 R-32 average, units 12-17)",
                f"{GRANT_COP_AT_7C} (manufacturer EN14511 datasheet)",
                "Unchanged from the current model -- no published multi-temperature curve for Grant Aerona3,"
                "so cold-weather sensitivity is assumed identical; only the reference-point COP is lowered.",
                f"{CAPEX_REDUCTION_FRAC * 100:.1f}% cut to HEAT_COSTS['ASHP']['capex_per_kwth'], "
                f"paired with the COP drop.",
                "First-stage sizing (n_pv, e_batt, o_batt, q_heat_cap, e_th) fixed at the saved "
                "deterministic-round solution; only stage-2 dispatch is re-solved with Grant "
                "Aerona3's COP/capex swapped in. Applies only to ASHP designs.",
            ],
        })
        notes.to_excel(xw, sheet_name="Assumptions", index=False)

    n_infeasible = int((out["grant_status"] != "Optimal").sum())
    n_worse      = int((out["cheap_ashp_pays_off"] == False).sum())
    n_better     = int((out["cheap_ashp_pays_off"] == True).sum())
    print(f"\nWrote {out_path}")
    print(f"{len(out)} ASHP designs total: {n_infeasible} Infeasible on Grant Aerona3 under the "
          f"fixed sizing, {n_worse} solved but net worse off over 15 yr despite the cheaper unit, "
          f"{n_better} solved and net better off.")
    return out


def main():
    ap = argparse.ArgumentParser(description="Standalone re-run against an already-completed "
                                             "optimisation run. The pipeline calls "
                                             "run_ashp_sensitivity() directly.")
    ap.add_argument("--limit", type=int, default=None,
                    help="Only solve the first N designs (for a quick test run).")
    ap.add_argument("--time-limit", type=int, default=600, dest="time_limit", metavar="S",
                    help="per-design solver time limit in seconds (default: 600)")
    ap.add_argument("--jobs", type=int, default=None, metavar="N",
                    help="parallel worker processes (default: a quarter of the logical cores; "
                         "see optimisation_config.resolve_jobs)")
    args = ap.parse_args()

    run_dir = _latest_run_dir()
    print(f"Latest run: {run_dir}")
    all_rows = _read_deterministic_rows(run_dir)
    run_ashp_sensitivity(all_rows, run_dir, time_limit_s=args.time_limit,
                         n_jobs=args.jobs, limit=args.limit)


if __name__ == "__main__":
    main()
