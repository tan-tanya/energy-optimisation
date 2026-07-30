"""
Imported by optimisation model, but can be run as a script to generate profiles.
Model generates building energy demand profiles per UK Met Office district and activity class.

Inputs 
1. data/inputs.xlsx; benchmarks and parameters from CIBSE, NCM, TM46, BEES, UKPN, Met Office
2. data/hdd/{ICAO}_HDD_15.5C.csv; daily heating degree-days from Met Office data per district
3. data/sunlighthours/*.txt; Met Office seasonal sunshine hours per district
4. data/api_temperature_profiles.xlsx; ERA5 hourly temperature (Open-Meteo) per district/season 

Outputs (only generated when script is run directly):
1. outputs/{district} - {activity} ({timestamp})/Demand Profiles.xlsx; sheets as follows
    - Energy Profiles          (half-hourly kW/m² benchmarks per season)
    - Demand Outputs           (annual + seasonal kWh totals per sample building)
2. *.png; seasonal / monthly / WD-WE heatmap demand plots

Data layers:
    (a) CIBSE annual energy benchmarks (typical kWh/m²/y per activity class)
    (b) NCM hourly occupancy + area-weighted heating setpoints
    (c) TM46 split of demand into HDD-driven vs baseload fractions
    (d) BEES median floor area per sub-sector (sample building size)
    (e) BEES base load fractions (always-on electricity share)
    (f) UKPN EV charging profile shape + per-activity EV demand parameters
    (g) Met Office sunshine hours (monthly lighting factor)
    (h) Met Office heating degree-days (seasonal / monthly stats per district)
    (i) ERA5 hourly temperature diurnal anomaly (real intra-day shape; sinusoid is the fallback)

Model parameters:
    - Diurnal temperature amplitude + solar geometry (DIURNAL_AMPLITUDE, SEASON_DOY) — sinusoid
      fallback only; real ERA5 diurnal shape (data layer (i)) is used whenever a district is passed
    - Heating system COP models (cop_ashp, cop_gshp, HEATING_SYSTEMS)

Pipeline:
    1. initialize()   loads all source data into module-level state
    2. main()         prompts for district + activity, builds workbook and plots
"""

import functools
import os
import subprocess
import sys
from datetime import datetime
import numpy as np
import pandas as pd
import datasets   # ERA5 hourly-temperature diurnal shape (real, vs the sinusoid fallback) reads through here
import openpyxl   # main() builds the demo workbook; styling/plots live in demand_report.py
from model_params import (
    ETA_BOILER, ashp_cop_at_7C, ashp_cop_slope_per_C,
    gshp_cop_at_10C, gshp_cop_slope_per_C,
    T_ASHP_MIN, T_ASHP_MAX, T_GSHP_MIN, T_GSHP_MAX,
    HORIZONTAL_LOOP_DEPTH_M, SOIL_THERMAL_DIFFUSIVITY_M2_S,
    SURFACE_TEMP_PEAK_DOY, BRINE_OFFSET_C, DEFAULT_BRINE_TEMP_C,
    EV_CHARGER_KW, EV_SPACE_FRACTION, EV_PARKING_DENSITY, EV_DWELL_HOURS,
    PEAK_FRACTION, T_SETBACK, HDD_BASE, UK_LATITUDE,
    DIURNAL_AMPLITUDE, SEASON_DOY, WE_LOAD_FACTOR,
)

# Run directly as well as import: alias this module's __main__ entry under its real name in 
# sys.modules before the circular `from demand_report import ...` below runs. 
# Without this, demand_report's `import demand_profile_model` re-executes this whole file from
# scratch as a second module instance, which re-triggers the same import mid-definition and fails.
if __name__ == "__main__":
    sys.modules.setdefault("demand_profile_model", sys.modules[__name__])


# 1 - PATHS
OUTPUTS_DIR = "outputs"


# 2 - SOURCE DATA CONFIGURATION (STATIC)
# area_weights = assumed floor-area fractions per NCM space type
ACTIVITY_LOOKUP = {
    "Health: Health centre": {
        "cibse_category": "Primary Health Care",
        "cibse_building": "Health Centres and Clinics",
        "ncm_rows":       (148, 154),  # NCM - Occupancy Excel rows 150-155 (pandas iloc 148:154)
        "tm46_cat":       19,          # Clinic
        "bees_sector":    "Health",
        "bees_sub":       "Health centres",
        "area_weights":   [0.03, 0.20, 0.07, 0.10, 0.05, 0.55],
    },
    "Health: Hospital": {
        "cibse_category": "Hospitals",
        "cibse_building": "General Acute Hospital",
        "ncm_rows":       (74, 97),    # NCM - Occupancy Excel rows 76-98 (pandas iloc 74:97)
        "tm46_cat":       20,          # Hospital (clinical and research)
        "bees_sector":    "Health",
        "bees_sub":       "Hospitals",
        "area_weights":   [0.03, 0.12, 0.04, 0.03, 0.01, 0.02,
                           0.02, 0.02, 0.02, 0.35, 0.02, 0.08, 0.02, 0.01, 0.01,
                           0.03, 0.05, 0.02, 0.01, 0.03, 0.02, 0.03, 0.01],
    },
    "Office: A/C standard": {
        "cibse_category": "Offices",
        "cibse_building": "Air conditioned, standard",
        "ncm_rows":       (0, 11),     # NCM - Occupancy Excel rows 2-12 (pandas iloc 0:11)
        "tm46_cat":       1,           # General office
        "bees_sector":    "Office",
        "bees_sub":       "Offices (public)",
        "area_weights":   [0.03, 0.15, 0.05, 0.05, 0.03, 0.50, 0.05, 0.03, 0.03, 0.05, 0.03],
    },
    "Retail: Department store": {
        "cibse_category": "Retail",
        "cibse_building": "Department stores",
        "ncm_rows":       (56, 66),    # NCM - Occupancy Excel rows 58-67 (pandas iloc 56:66)
        "tm46_cat":       4,           # Large non-food shop (includes department stores)
        "bees_sector":    "Retail",
        "bees_sub":       "Large non-food shops",
        "area_weights":   [0.05, 0.03, 0.03, 0.03, 0.45, 0.15, 0.10, 0.08, 0.04, 0.04],
    },
}

from districts import DISTRICT_STATIONS  

# Base load (always-on electricity share) derived from BEES 2014-15 Energy Intensity survey)
BASE_LOAD_CFG = {
    "Health: Health centre":    (slice(141, 148), {"Hot water": 1.0, "Lighting": 0.5}),
    "Health: Hospital":         (slice(148, 155), {"Medical equipment": 1.0, "Cooling & humidification": 1.0, "Fans": 1.0, "Hot water": 1.0}),
    "Office: A/C standard":     (slice(246, 254), {"ICT Equipment": 1.0, "Hot water": 1.0, "Cooling & humidification": 1.0, "Fans": 1.0}),
    "Retail: Department store": (slice(300, 307), {"Cooled storage": 1.0, "Cooling & humidification": 1.0, "Fans": 1.0, "Lighting": 0.2}),
}

# EV demand inputs: EV_PARKING_DENSITY and EV_DWELL_HOURS loaded from data/Model_Parameters.xlsx via model_params
# Weekend load factors derived from TM46 operational schedules: WE_LOAD_FACTOR loaded from data/Model_Parameters.xlsx via model_params


# 3 - CALENDAR / SEASON CONSTANTS
MONTH_SEASON = {
    "January": "Winter",  "February": "Winter",  "March":     "Spring",
    "April":   "Spring",  "May":      "Spring",  "June":      "Summer",
    "July":    "Summer",  "August":   "Summer",  "September": "Autumn",
    "October": "Autumn",  "November": "Autumn",  "December":  "Winter",
}
MONTH_DAYS = {
    "January":   31, "February": 28, "March":     31, "April":   30,
    "May":       31, "June":     30, "July":      31, "August":  31,
    "September": 30, "October":  31, "November":  30, "December": 31,
}
MONTHS_ORDER           = list(MONTH_DAYS.keys())
# First day-of-year of each month. Evaluates 1.5 m ground temperature / brine COP for every calendar day in a month
MONTH_START_DOY  = {m: 1 + sum(MONTH_DAYS[mm] for mm in MONTHS_ORDER[:i])
                    for i, m in enumerate(MONTHS_ORDER)}
SEASON_ORDER_BENCHMARK = ["Summer", "Autumn", "Winter", "Peak Winter", "Spring"]  # Energy Profiles sheet
SEASON_ORDER_DEMAND    = ["Spring", "Summer", "Autumn", "Winter", "Peak Winter"]  # Demand Outputs sheet
MAIN_SEASONS           = ["Summer", "Autumn", "Winter", "Spring"]                 # excludes Peak Winter

# Solar geometry drives both intra-day temperature shape and the lighting demand shape. 
# DIURNAL_AMPLITUDE only drives temperature, and only used when no district is passed to hourly_temp_profile(), 
# or ERA5 data is unavailable for it, otherwise the real ERA5 diurnal anomaly is used instead.
# Sinusoid source: UK Met Office climate statistics; consistent with CIBSE Guide A §2.3.

# Heat-demand intra-day shape: the day's heat is distributed by both temperature and occupancy.
# Occupancy raises the indoor target (setback↔setpoint) while ERA5 diurnal temperature sets the gradient.
# Falls back to an occupancy-only shape (floored at a flat setback level) when no heating is needed all day.


# 4 - PHYSICAL / NUMERICAL CONSTANTS
REFERENCE_DISTRICT = "England SE and Central S"   # CIBSE benchmarks assumed for London Heathrow

# Ground-loop thermal model (Banks 2008 §3.3, §8.1, §8.5)
# Damping depth d = sqrt(2·α/ω); ω = 2π/year. Seasonal signal attenuates as exp(-z/d) and lags by z/d radians.
_OMEGA_YR                     = 2.0 * np.pi / (365.25 * 86400.0)
DAMPING_DEPTH_M               = np.sqrt(2.0 * SOIL_THERMAL_DIFFUSIVITY_M2_S / _OMEGA_YR)   


# 5 - HEATING SYSTEMS
# Gas Boiler: fixed efficiency; ASHP: varies with outdoor temperature; GSHP: varies with ground temperature
def cop_ashp(t_ext: float) -> float:
    """Air-source heat pump COP at outdoor temperature t_ext (°C).
    Average slope from Company 4's R-32 heat pumps in CEP Technology Library v1.02, HP air-water (Heating) sheet.
    R32 ASHPs selected to reflect increased adoption in the UK, amidst ongoing phase-out of R410a models."""
    t = max(T_ASHP_MIN, min(T_ASHP_MAX, t_ext))
    return ashp_cop_at_7C + ashp_cop_slope_per_C * (t - 7.0)

def cop_gshp(t_brine: float = DEFAULT_BRINE_TEMP_C) -> float:
    """Ground-source heat pump COP at brine inlet temperature t_brine (°C).
    Slope from Company 1's 49-75 kW commercial unit in CEP Technology Library v1.02, HP ground-water (Heating) sheet."""
    t = max(T_GSHP_MIN, min(T_GSHP_MAX, t_brine))
    return gshp_cop_at_10C + gshp_cop_slope_per_C * (t - 10.0)

def daily_ground_temperature(annual_mean_T_C: float, surface_amplitude_C: float,
                             day_of_year: int, depth_m: float = HORIZONTAL_LOOP_DEPTH_M) -> float:
    """Soil temperature at given depth and DOY from 1-D heat conduction with sinusoidal annual surface boundary (Banks 2008 §3.3, Figure 3.3).
    At z = 1.5 m: amplitude ≈ 54% of surface, phase lag ≈ 35 days; At z ≳ 5·DAMPING_DEPTH (~12 m): seasonal signal < 1%, T ≈ annual mean."""
    z = depth_m
    d = DAMPING_DEPTH_M
    return (annual_mean_T_C
            + surface_amplitude_C
              * np.exp(-z / d)
              * np.cos(2.0 * np.pi * (day_of_year - SURFACE_TEMP_PEAK_DOY) / 365.0 - z / d))

def brine_temperature(annual_mean_T_C: float, surface_amplitude_C: float,
                      day_of_year: int, loop_type: str) -> float:
    """Horizontal loops at 1.5 m carry seasonal ground-temperature variation; 
    vertical boreholes (≥80 m depth) effectively have constant brine temperatures at annual mean ground T year-round (Banks 2008 Figure 3.5)."""
    if loop_type == "horizontal":
        ground_T = daily_ground_temperature(annual_mean_T_C, surface_amplitude_C, day_of_year)
    else:
        ground_T = annual_mean_T_C
    return ground_T - BRINE_OFFSET_C

# GSHP "cop" lambda is only a fallback for callers that don't supply a district to _breakdown. When `district` is passed, 
# _breakdown computes brine_temperature(..., loop_type) → cop_gshp(t_brine) directly using district_climate_stats[district].
HEATING_SYSTEMS = {
    "Gas Boiler":        {"is_heat_pump": False, "cop": None,                "loop_type": None},
    "ASHP":              {"is_heat_pump": True,  "cop": cop_ashp,            "loop_type": None},
    "GSHP (vertical)":   {"is_heat_pump": True,  "cop": lambda _t: cop_gshp(DEFAULT_BRINE_TEMP_C - BRINE_OFFSET_C), "loop_type": "vertical"},
    "GSHP (horizontal)": {"is_heat_pump": True,  "cop": lambda _t: cop_gshp(DEFAULT_BRINE_TEMP_C - BRINE_OFFSET_C), "loop_type": "horizontal"},
}
HEATING_OPTIONS = list(HEATING_SYSTEMS)


# 7 - MODULE-LEVEL STATE (populated by initialize(); None until then)
cibse_dashboard         = None   # {activity: {electricity_typical, fossil_typical, description, category, source}}
ncm_occupancy           = None   # {activity: 24-hour occupancy fraction list}
ncm_setpoints           = None   # {activity: area-weighted heating setpoint °C}
tm46_adjustment         = None   # {activity: {elec, fossil} HDD-driven fractions}
bees_floor_areas        = None   # {activity: median floor area m²}
base_load_fracs         = None   # {activity: always-on electricity share}
ev_profile_48           = None   # (48,) half-hourly EV charging utilisation [fraction of rated capacity]
degree_days_by_district = None   # {district: {season: {hdd_per_day, n_days, light_factor}}}
daily_hdd_by_district   = None   # {district: mean daily HDD across year}
monthly_dd_by_district  = None   # {district: {month: {hdd_per_day, light_factor, n_days}}}
reference_daily_hdd     = None   # daily HDD for REFERENCE_DISTRICT
district_climate_stats  = None   # {district: {annual_mean_T_C, surface_amplitude_C}} — for GSHP ground-temp model


# 8 - DATA LAYER CALCULATIONS
def solar_elevation_profile(season: str, latitude: float = UK_LATITUDE) -> np.ndarray:
    """24-element relative solar availability (solar noon = 1) for the mid-season day of year."""
    doy      = SEASON_DOY[season]
    decl_rad = np.radians(23.45 * np.sin(np.radians(360.0 / 365.0 * (doy - 81))))
    lat_rad  = np.radians(latitude)
    h        = np.arange(24) + 0.5
    ha_rad   = np.radians((h - 12) * 15.0)
    sin_elev = (np.sin(lat_rad) * np.sin(decl_rad)
                + np.cos(lat_rad) * np.cos(decl_rad) * np.cos(ha_rad))
    profile  = np.maximum(0.0, sin_elev)
    peak     = profile.max()
    return profile / peak if peak > 0 else profile

def hourly_temp_profile(t_mean: float, season: str, latitude: float = UK_LATITUDE,
                        district: str = None) -> np.ndarray:
    """24-element hourly dry-bulb temperature [°C] swinging around the daily mean `t_mean`.

    Shape source (HDD owns the level, temperature owns the shape):
      - When `district` is given AND ERA5 profiles exist, use the REAL mean diurnal anomaly for that
        (district, season) from api_temperature_profiles.xlsx:
      - Otherwise fall back to the idealised sinusoid T(h) = t_mean + A·cos(2π(h-peak)/24), peak ~2 h
        after solar noon, amplitude DIURNAL_AMPLITUDE[season] — phase from solar geometry only."""
    if district is not None:
        anomaly = datasets.get_temperature_anomaly(district, season)
        if anomaly is not None:
            return t_mean + anomaly
    solar  = solar_elevation_profile(season, latitude)
    noon_h = float(np.argmax(solar))
    peak_h = noon_h + 2.0
    a      = DIURNAL_AMPLITUDE[season]
    h      = np.arange(24)
    return t_mean + a * np.cos(2 * np.pi * (h - peak_h) / 24)

def _ev_capacity_kw(activity: str) -> float:
    """Installed EV charging capacity [kW/m²] — nameplate, not a peak demand.

    ev_profile_48 supplies the utilisation fraction (peaks at 0.526), so the actual half-hourly
    peak demand is this × 0.526, not this.
    """
    util = min(1.0, EV_DWELL_HOURS[activity] / 8.0)
    return EV_PARKING_DENSITY[activity] * EV_SPACE_FRACTION * EV_CHARGER_KW * util

def ev_kw_per_sqm(activity: str) -> np.ndarray:
    """48-slot half-hourly EV charging demand [kW/m²]."""
    return _ev_capacity_kw(activity) * ev_profile_48

def ev_annual_kwh_per_sqm(activity: str) -> float:
    """Annual EV charging energy [kWh/m²]."""
    return float(_ev_capacity_kw(activity) * ev_profile_48.sum() * 0.5 * 365)


# 9 - CORE DEMAND CALCULATIONS
# lru_cache enables memoisation of intermediate results for each combination.
@functools.lru_cache(maxsize=None)
def _season_context(activity: str, season: str, daily_hdd: float,
                    seasonal_hdd: float, light_fac: float, district: str = None) -> dict:
    """Per-season intermediate factors used by the demand breakdown.
    `district` selects the ERA5 diurnal temperature for the degree-hour heat shape; None keeps synthetic-sinusoid temperature."""
    e            = cibse_dashboard[activity]["electricity_typical"]
    f            = cibse_dashboard[activity]["fossil_typical"]
    e_adj        = tm46_adjustment[activity]["elec"]
    f_adj        = tm46_adjustment[activity]["fossil"]
    e_base       = base_load_fracs[activity]
    hdd_ratio    = daily_hdd / reference_daily_hdd
    t_ext        = HDD_BASE - seasonal_hdd
    t_set        = ncm_setpoints[activity]
    heat_setback = max(0.0, (T_SETBACK - t_ext) / (t_set - t_ext)) if t_set > t_ext else 0.0
    occ          = np.array(ncm_occupancy[activity])
    occ_norm     = occ / occ.sum() if occ.sum() > 0 else np.ones(24) / 24
    # Occupancy-only fallback shape: heated to occupancy, floored at a flat setback level.
    heat_occ     = np.maximum(occ, heat_setback)
    heat_norm    = heat_occ / heat_occ.sum()
    # Both occupancy and temperature set the intra-day heat demand.
    gate         = np.clip(occ / occ.max(), 0.0, 1.0) if occ.max() > 0 else np.zeros(24)
    t_indoor     = T_SETBACK + (t_set - T_SETBACK) * gate
    t_out_h      = hourly_temp_profile(t_ext, season, district=district)
    heat_h       = np.maximum(0.0, t_indoor - t_out_h)
    # Fall back to the occupancy shape when no heating is needed all day.
    heat_norm    = heat_h / heat_h.sum() if heat_h.sum() > 1e-9 else heat_norm
    uniform      = np.ones(24) / 24
    solar_h      = solar_elevation_profile(season)
    w_solar      = (occ_norm * solar_h).sum()
    w_dark       = 1.0 - w_solar
    # light_fac is normalised so the day-weighted annual mean = 1.0.
    #   - light_fac <= 1 (bright months): DIM during daylight — light_fac_h = 1 - solar_h*d,
    #     d clamped to 1 so it never goes negative at solar noon.
    #   - light_fac >  1 (dark months): BOOST during darkness (1 - solar_h) rather than daylight —
    #     using solar_h here would flip the sign and put extra artificial-lighting load at solar noon instead of night.
    if light_fac <= 1.0:
        d_light     = min((1.0 - light_fac) / w_solar, 1.0) if w_solar > 1e-9 else 0.0
        light_fac_h = 1.0 - solar_h * d_light
    else:
        b_light     = (light_fac - 1.0) / w_dark if w_dark > 1e-9 else 0.0
        light_fac_h = 1.0 + (1.0 - solar_h) * b_light
    d_base       = e * e_base * (1 - e_adj) / 365
    d_occ_base   = e * (1 - e_base) * (1 - e_adj) / 365
    d_dd_elec    = e * e_adj * hdd_ratio * (seasonal_hdd / daily_hdd) / 365
    return {
        "F": f, "F_adj": f_adj,
        "seasonal_HDD": seasonal_hdd, "hdd_ratio": hdd_ratio, "T_ext": t_ext,
        "occ_norm": occ_norm, "heat_norm": heat_norm, "uniform": uniform,
        "light_fac_h": light_fac_h,
        "D_base": d_base, "D_occ_base": d_occ_base, "D_dd_elec": d_dd_elec,
    }

@functools.lru_cache(maxsize=None)
def _breakdown(activity: str, heating: str, season: str, daily_hdd: float,
               seasonal_hdd: float, light_fac: float,
               district: str = None) -> dict:
    """Half-hourly kW/m² split into end-use components.
    For GSHP, if `district` is supplied the heat-pump COP is derived from the per-(district, season, loop-type) brine inlet temperature 
    via `brine_temperature()` and `cop_gshp()`. Otherwise the heat-pump COP function from HEATING_SYSTEMS is used (ASHP uses outdoor air; 
    GSHP falls back to the annual-mean brine inlet DEFAULT_BRINE_TEMP_C − BRINE_OFFSET_C == COP ≈ 2.95)."""
    c = _season_context(activity, season, daily_hdd, seasonal_hdd, light_fac, district)
    f, f_adj                       = c["F"], c["F_adj"]
    seasonal_HDD, hdd_ratio, t_ext = c["seasonal_HDD"], c["hdd_ratio"], c["T_ext"]
    occ_norm, heat_norm, uniform   = c["occ_norm"], c["heat_norm"], c["uniform"]
    light_fac_h                    = c["light_fac_h"]
    d_base, d_occ_base, d_dd_elec  = c["D_base"], c["D_occ_base"], c["D_dd_elec"]

    lsp_kW  = d_base * uniform + d_occ_base * occ_norm * light_fac_h
    hvac_kW = d_dd_elec * occ_norm

    system = HEATING_SYSTEMS[heating]
    if system["is_heat_pump"]:
        loop_type = system["loop_type"]
        if loop_type is not None and district is not None:
            # Brine inlet temp varies seasonally for horizontal (1.5 m depth) and is flat at annual-mean ground T for vertical borefields (≥80 m depth).
            # COP is constant across hours of the day because brine temp has no diurnal variation at borehole depth.
            stats     = district_climate_stats[district]
            t_brine   = brine_temperature(stats["annual_mean_T_C"],
                                          stats["surface_amplitude_C"],
                                          SEASON_DOY[season], loop_type)
            cop_values = np.full(24, cop_gshp(t_brine))
        else:
            cop_fn     = system["cop"]
            t_hourly   = hourly_temp_profile(t_ext, season, district=district)
            cop_values = np.array([cop_fn(t) for t in t_hourly])
        q_heat_day = f * ETA_BOILER * hdd_ratio * (seasonal_HDD / daily_hdd) / 365
        hp_kW      = q_heat_day * heat_norm / cop_values
    else:
        hp_kW = np.zeros(24)

    if heating == "Gas Boiler":
        sh_kW = (f * f_adj * hdd_ratio * (seasonal_HDD / daily_hdd) / 365) * heat_norm
        hw_kW = (f * (1 - f_adj) / 365) * occ_norm
    else:
        sh_kW = np.zeros(24)
        hw_kW = np.zeros(24)

    return {
        "Lighting & Small Power": np.repeat(lsp_kW,  2),
        "HVAC":                   np.repeat(hvac_kW, 2),
        "Heat Pump":              np.repeat(hp_kW,   2),
        "EV Charging":            ev_kw_per_sqm(activity),
        "Space Heating":          np.repeat(sh_kW,   2),
        "Hot Water & Process":    np.repeat(hw_kW,   2),
    }

def half_hourly_kw_per_sqm_breakdown(activity: str, heating: str, season: str,
                                     degree_days: dict, daily_hdd: float,
                                     district: str = None) -> dict:
    """Public wrapper: dict-keyed lookup → cached implementation.
    Pass `district` to enable GSHP COP variation by district / season / loop type."""
    s = degree_days[season]
    return _breakdown(activity, heating, season, daily_hdd,
                      s["hdd_per_day"], s["light_factor"], district)

@functools.lru_cache(maxsize=None)
def _energy_total(activity: str, heating: str, energy_type: str, season: str,
                  daily_hdd: float, seasonal_hdd: float, light_fac: float,
                  district: str = None) -> np.ndarray:
    b = _breakdown(activity, heating, season, daily_hdd, seasonal_hdd, light_fac, district)
    if energy_type == "Electricity":
        return b["Lighting & Small Power"] + b["HVAC"] + b["Heat Pump"] + b["EV Charging"]
    if energy_type == "Gas":
        return b["Space Heating"] + b["Hot Water & Process"]
    raise ValueError(f"unknown energy_type: {energy_type!r}")

def half_hourly_kw_per_sqm(activity: str, heating: str, energy_type: str, season: str,
                           degree_days: dict, daily_hdd: float,
                           district: str = None) -> np.ndarray:
    """48-slot half-hourly kW/m² total demand for the given energy type.
    Pass `district` to enable GSHP COP variation by district / season / loop type."""
    s = degree_days[season]
    return _energy_total(activity, heating, energy_type, season, daily_hdd,
                         s["hdd_per_day"], s["light_factor"], district)

@functools.lru_cache(maxsize=None)
def annual_demand_kwh(activity: str, heating: str, daily_hdd: float) -> dict:
    """Annual demand [kWh/m²] split into {electricity, gas}."""
    e         = cibse_dashboard[activity]["electricity_typical"]
    f         = cibse_dashboard[activity]["fossil_typical"]
    e_adj     = tm46_adjustment[activity]["elec"]
    f_adj     = tm46_adjustment[activity]["fossil"]
    hdd_ratio = daily_hdd / reference_daily_hdd

    e_scaled = e * (1 - e_adj) + e * e_adj * hdd_ratio
    ev       = ev_annual_kwh_per_sqm(activity)

    system = HEATING_SYSTEMS[heating]
    if system["is_heat_pump"]:
        t_ext_annual = HDD_BASE - daily_hdd
        annual_cop   = system["cop"](t_ext_annual)
        return {"electricity": e_scaled + f * ETA_BOILER * hdd_ratio / annual_cop + ev, "gas": 0}
    f_scaled = f * (1 - f_adj) + f * f_adj * hdd_ratio
    return {"electricity": e_scaled + ev, "gas": f_scaled}

def building_halfhourly_kwh(activity: str, heating: str, season: str, floor_area: float,
                            degree_days: dict, daily_hdd: float,
                            district: str = None) -> dict:
    elec = half_hourly_kw_per_sqm(activity, heating, "Electricity", season, degree_days, daily_hdd, district)
    gas  = half_hourly_kw_per_sqm(activity, heating, "Gas",         season, degree_days, daily_hdd, district)
    elec_kwh  = (elec * 0.5 * floor_area).tolist()
    gas_kwh   = (gas  * 0.5 * floor_area).tolist()
    total_kwh = [e + g for e, g in zip(elec_kwh, gas_kwh)]
    return {"total": total_kwh, "electricity": elec_kwh, "gas": gas_kwh}

@functools.lru_cache(maxsize=None)
def _peak_scale_factor(activity: str, heating: str, daily_hdd: float,
                       seasonal_hdd_pw: float, light_fac_pw: float,
                       district: str = None) -> float:
    b = _breakdown(activity, heating, "Peak Winter", daily_hdd, seasonal_hdd_pw, light_fac_pw, district)
    max_total_kw = float(np.max(sum(b.values())))
    annual = annual_demand_kwh(activity, heating, daily_hdd)
    target_kw = PEAK_FRACTION * (annual["electricity"] + annual["gas"]) / 0.5
    return target_kw / max_total_kw if max_total_kw > 0 else 1.0

def peak_scale_factor(activity: str, heating: str, degree_days: dict, daily_hdd: float,
                      district: str = None) -> float:
    """Scale Peak Winter profile so max half-hourly kWh = PEAK_FRACTION of annual."""
    s = degree_days["Peak Winter"]
    return _peak_scale_factor(activity, heating, daily_hdd, s["hdd_per_day"], s["light_factor"], district)

def sample_buildings() -> list:
    """All (activity × heating) sample buildings."""
    out = []
    bid = 1
    for act in ACTIVITY_LOOKUP:
        for heating in HEATING_OPTIONS:
            out.append({
                "id":       bid,
                "activity": act,
                "name":     cibse_dashboard[act]["description"],
                "area":     bees_floor_areas[act],
                "heating":  heating,
            })
            bid += 1
    return out

def seasonal_demand_kwh(activity: str, heating: str, season: str, floor_area: float,
                        degree_days: dict, daily_hdd: float, district: str = None) -> tuple:
    """Seasonal kWh totals for one building: (total, electricity, gas), rounded to integers.
    Peak Winter is the single coldest half-hour scaled to PEAK_FRACTION of annual demand."""
    if season == "Peak Winter":
        annual           = annual_demand_kwh(activity, heating, daily_hdd)
        ann_total_per_m2 = annual["electricity"] + annual["gas"]
        peak_kwh         = PEAK_FRACTION * ann_total_per_m2 * floor_area
        b        = half_hourly_kw_per_sqm_breakdown(activity, heating, "Peak Winter", degree_days, daily_hdd, district)
        elec_pw  = b["Lighting & Small Power"] + b["HVAC"] + b["Heat Pump"] + b["EV Charging"]
        gas_pw   = b["Space Heating"] + b["Hot Water & Process"]
        total_pw = elec_pw + gas_pw
        ts       = int(np.argmax(total_pw))
        e_frac   = float(elec_pw[ts] / total_pw[ts]) if total_pw[ts] > 0 else 1.0
        return (round(peak_kwh), round(peak_kwh * e_frac), round(peak_kwh * (1.0 - e_frac)))

    kwh = building_halfhourly_kwh(activity, heating, season, floor_area, degree_days, daily_hdd, district)
    n   = degree_days[season]["n_days"]
    return (round(sum(kwh["total"])       * n),
            round(sum(kwh["electricity"]) * n),
            round(sum(kwh["gas"])         * n))


# 10. LOADERS 
# (a) CIBSE annual energy benchmarks (typical kWh/m²/y per activity)
def load_cibse_dashboard(xl: pd.ExcelFile) -> dict:
    df = xl.parse("CIBSE Dashboard", header=0)
    df.columns = ["category", "building_type", "year", "fossil_good", "fossil_typical", "elec_good", "elec_typical", "units", "source"]
    out = {}
    for act, cfg in ACTIVITY_LOOKUP.items():
        row = df[
            (df["category"]      == cfg["cibse_category"]) &
            (df["building_type"] == cfg["cibse_building"])
        ].iloc[0]
        out[act] = {
            "description":         row["building_type"],
            "category":            row["category"],
            "electricity_typical": int(row["elec_typical"]),
            "fossil_typical":      int(row["fossil_typical"]),
            "source":              row["source"],
        }
    return out

# (b) NCM hourly occupancy + area-weighted heating setpoints
def load_ncm_data(xl: pd.ExcelFile):
    """Returns (occupancy_profiles, heating_setpoints); Occupancy = area-weighted mean"""
    df_ncm = xl.parse("NCM - Occupancy", header=0)
    df_act = xl.parse("NCM - Activity",  header=0)
    hour_cols         = [c for c in df_ncm.columns if str(c).startswith("Hr ")]
    heat_setpoint_col = [c for c in df_act.columns if "Heat Setpoint" in str(c)][0]

    occ_out, set_out = {}, {}
    for act, cfg in ACTIVITY_LOOKUP.items():
        start, end = cfg["ncm_rows"]
        profiles  = df_ncm.iloc[start:end][hour_cols].values.astype(float)
        densities = df_act.iloc[start:end]["Occupancy Density (m²/person)"].values.astype(float)
        area_w    = np.array(cfg["area_weights"], dtype=float)
        weights   = area_w * (1.0 / densities)   # area × occupants/m²
        weights   = weights / weights.sum()
        occ_out[act] = (weights[:, None] * profiles).sum(axis=0).tolist()

        # Area-weighted mean heating setpoint
        setpoints = df_act.iloc[start:end][heat_setpoint_col].values.astype(float)
        area_norm = area_w / area_w.sum()
        set_out[act] = float((area_norm * setpoints).sum())
    return occ_out, set_out

# (c) TM46 split of demand into HDD-driven vs baseload fractions
def load_tm46_adjustment(xl: pd.ExcelFile) -> dict:
    df = xl.parse("CIBSE - TM46 Adjustment", header=1)
    df.columns = ["cat", "category_name", "elec_pct", "fossil_pct", "separable",
                  "occupancy_def", "ref_hours", "max_hours", "elec_increase", "fossil_increase"]

    def pct(v):
        return float(str(v).strip().replace("%", "")) / 100.0

    out = {}
    for act, cfg in ACTIVITY_LOOKUP.items():
        row = df[df["cat"] == cfg["tm46_cat"]].iloc[0]
        out[act] = {"elec": pct(row["elec_pct"]), "fossil": pct(row["fossil_pct"])}
    return out

# (d) BEES median floor area per sub-sector (used as sample building size)
def load_bees_floor_areas(xl: pd.ExcelFile) -> dict:
    """Median premises floor area per BEES sub-sector (Figure 1.4, England and Wales); header=2."""
    df = xl.parse("BEES - Floor Areas", header=2)
    df.columns = ["sector", "sub_sector", "statistic", "value"]
    out = {}
    for act, cfg in ACTIVITY_LOOKUP.items():
        row = df[
            (df["sector"]     == cfg["bees_sector"]) &
            (df["sub_sector"] == cfg["bees_sub"]) &
            (df["statistic"]  == "Median")
        ].iloc[0]
        out[act] = float(row["value"])
    return out

# (e) BEES base load fractions — always-on electricity share
def load_base_load_fracs(xl: pd.ExcelFile, verbose: bool = False) -> dict:
    df = xl.parse("BEES - Energy End Uses", header=None)
    df.columns = ["sector", "sub_sector", "end_use", "value"]
    fracs = {}
    for act, (rows, components) in BASE_LOAD_CFG.items():
        sub   = df.iloc[rows]
        uses  = dict(zip(sub["end_use"].str.strip(), sub["value"]))
        total = sum(uses.values())
        fracs[act] = round(sum(w * uses.get(k, 0) for k, w in components.items()) / total, 4)

    if verbose:
        print("Derived base load fractions (BEES 2014-15, Energy Intensity by Sub-sector):")
        for act, frac in fracs.items():
            print(f"  {act}: {frac:.4f}  ({frac * 100:.1f}%)")
    return fracs

# (f) UKPN EV charging utilisation profile
def load_ev_profile(xl: pd.ExcelFile) -> np.ndarray:
    """48-slot half-hourly EV charging utilisation [fraction of rated capacity] from the UKPN dataset.

    The UKPN 'Standard Profiles ... for Electricity Demand' columns are already percentages of RATED
    CAPACITY (0–100), not measured demand — the sheet's `Non_variable` column is a constant 100.0.
    So the correct conversion is /100, NOT /max(): the series carries its own absolute level, and
    peak-normalising would discard it. Averaged over all 365 days of 2019; EV charging is essentially
    flat across the week (weekend/weekday mean ratio 1.030), so no weekday/weekend split is applied
    here — adding one would impose a swing the source data does not show.

    Resulting profile peaks at 0.526 and averages 0.348 of rated capacity.
    """
    df = xl.parse("UKPN - EV Profiles")
    df["Timestamp"] = pd.to_datetime(df["Timestamp"], utc=True)
    df["slot"] = df["Timestamp"].dt.hour * 2 + (df["Timestamp"].dt.minute >= 30).astype(int)
    raw = df.groupby("slot")["EV_Charging"].mean().sort_index().values  # (48,) 0–100 [% of capacity]
    return raw / 100.0

# (g) Met Office sunshine hours to derive a monthly lighting factor (per district)
#
# Each month's sunshine is averaged over the last datasets.SUNSHINE_BASELINE_YEARS complete years 
# (default 10 -> 2016-2025) so the factor is not driven by an anomalous single year.
MONTH_COL = {"January": "jan", "February": "feb", "March":     "mar", "April":   "apr",
             "May":     "may", "June":     "jun", "July":      "jul", "August":  "aug",
             "September":"sep", "October":  "oct", "November":  "nov", "December":"dec"}

def load_light_factors(verbose: bool = False) -> dict:
    """{district: {month: raw_light_factor}} — inverse of monthly sunshine hours (more sunshine → less artificial demand). 
    Each month's sunshine is averaged over the last datasets.SUNSHINE_BASELINE_YEARS complete years (default 10)."""
    n_years      = datasets.SUNSHINE_BASELINE_YEARS
    district_sun = {}

    for district in DISTRICT_STATIONS:
        df = datasets.get_sunshine(district)
        # Coerce the monthly columns to numeric ("---"/blanks -> NaN), drop incomplete years
        # (e.g. the partial latest year), then average the last n_years complete rows per month.
        cols = [c for c in MONTH_COL.values() if c in df.columns]
        df   = df.copy()
        df[cols] = df[cols].apply(pd.to_numeric, errors="coerce")
        complete = df.dropna(subset=cols).sort_values("year").tail(n_years)
        if complete.empty:
            continue
        sun = {month: float(complete[col].mean())
               for month, col in MONTH_COL.items() if col in complete.columns}
        if sun:
            district_sun[district] = sun

    if verbose:
        yrs = "?"
        try:
            sample   = datasets.get_sunshine(next(iter(DISTRICT_STATIONS)))
            scols    = [c for c in MONTH_COL.values() if c in sample.columns]
            sample[scols] = sample[scols].apply(pd.to_numeric, errors="coerce")
            sample   = sample.dropna(subset=scols).sort_values("year").tail(n_years)
            yrs = f"{int(sample['year'].min())}-{int(sample['year'].max())}"
        except Exception:
            pass
        print(f"Monthly sunshine hours per district (mean of last {n_years} complete years, {yrs}):")
        for d, sun in district_sun.items():
            print(f"  {d}: { {m[:3]: round(h, 1) for m, h in sun.items()} }")

    return {d: {m: 1.0 / h for m, h in sun.items()} for d, sun in district_sun.items()}

# (h) Met Office heating degree-days → seasonal / monthly stats per district
def load_degree_days(raw_lf: dict, icao: str):
    """Returns (seasonal_stats, annual_daily_hdd, monthly_stats) for one district's ICAO station.
    Uses the last datasets.HDD_BASELINE_YEARS calendar years of available HDD data; seasonal/monthly
    means average across those years (per-day means, so partial boundary years are handled correctly)."""
    df = datasets.get_degree_days(icao, last_years=datasets.HDD_BASELINE_YEARS)
    df["month"] = df["date"].dt.month

    def season(m):
        if m in (6, 7, 8):    return "Summer"
        if m in (9, 10, 11):  return "Autumn"
        if m in (12, 1, 2):   return "Winter"
        return "Spring"
    df["season"] = df["month"].map(season)

    canonical_n = {s: sum(MONTH_DAYS[m] for m in MONTHS_ORDER if MONTH_SEASON[m] == s)
                   for s in MAIN_SEASONS}

    stats = {}
    for s in MAIN_SEASONS:
        sub = df[df["season"] == s]
        # hdd_per_day from actual data; n_days uses canonical length for annual consistency
        stats[s] = {"hdd_per_day": round(sub["hdd"].mean(), 4),
                    "n_days":      canonical_n[s]}

    # Peak Winter = single coldest half-hour; magnitude later scaled to PEAK_FRACTION of annual
    stats["Peak Winter"] = {"hdd_per_day": round(float(df["hdd"].max()), 4), "n_days": 1}

    # Lighting factors are monthly. Normalise so the day-weighted mean across the 12 months = 1.0, 
    # redistributing lighting load across the year without changing the annual total. 
    year_days  = sum(MONTH_DAYS[m] for m in MONTHS_ORDER)
    wt_avg     = sum(raw_lf[m] * MONTH_DAYS[m] for m in MONTHS_ORDER) / year_days
    month_lf   = {m: round(raw_lf[m] / wt_avg, 4) for m in MONTHS_ORDER}
    for s in MAIN_SEASONS:
        s_months = [m for m in MONTHS_ORDER if MONTH_SEASON[m] == s]
        s_days   = sum(MONTH_DAYS[m] for m in s_months)
        stats[s]["light_factor"] = round(sum(month_lf[m] * MONTH_DAYS[m] for m in s_months) / s_days, 4)
    stats["Peak Winter"]["light_factor"] = stats["Winter"]["light_factor"]

    daily_hdd = round(df["hdd"].mean(), 4)   # mean daily HDD across the baseline window (was sum/365 for a single year)

    monthly = {
        m: {"hdd_per_day": 0.0 if (sub := df[df["month"] == i]).empty else round(sub["hdd"].mean(), 4),
            "light_factor": month_lf[m],
            "n_days":       MONTH_DAYS[m]}
        for i, m in enumerate(MONTHS_ORDER, 1)
    }
    return stats, daily_hdd, monthly


# 10 - CLI / MAIN
def _prompt_choice(prompt, options):
    print(f"\n{prompt}")
    for i, opt in enumerate(options, 1):
        print(f"  {i}. {opt}")
    while True:
        try:
            sel = int(input("Enter number: ").strip())
            if 1 <= sel <= len(options):
                return options[sel - 1]
        except ValueError:
            pass
        print("Invalid selection")

def initialize(verbose: bool = False):
    """Load all source data into module-level state.
    Set verbose=True to print derived intermediates (base load fractions, sunshine hours)."""
    global cibse_dashboard, ncm_occupancy, ncm_setpoints, tm46_adjustment, bees_floor_areas
    global base_load_fracs, ev_profile_48
    global degree_days_by_district, daily_hdd_by_district, monthly_dd_by_district
    global reference_daily_hdd, district_climate_stats

    # Clear caches in case initialize() is called more than once in the same process
    _season_context.cache_clear()
    _breakdown.cache_clear()
    _energy_total.cache_clear()
    annual_demand_kwh.cache_clear()
    _peak_scale_factor.cache_clear()

    with datasets.inputs_workbook() as xl:
        cibse_dashboard              = load_cibse_dashboard(xl)
        ncm_occupancy, ncm_setpoints = load_ncm_data(xl)
        tm46_adjustment              = load_tm46_adjustment(xl)
        bees_floor_areas             = load_bees_floor_areas(xl)
        base_load_fracs              = load_base_load_fracs(xl, verbose=verbose)
        ev_profile_48                = load_ev_profile(xl)

    district_raw_lfs        = load_light_factors(verbose=verbose)
    degree_days_by_district = {}
    daily_hdd_by_district   = {}
    monthly_dd_by_district  = {}
    for district, raw_lf in district_raw_lfs.items():
        icao = DISTRICT_STATIONS[district]
        dd, daily, monthly = load_degree_days(raw_lf, icao)
        degree_days_by_district[district] = dd
        daily_hdd_by_district[district]   = daily
        monthly_dd_by_district[district]  = monthly

    reference_daily_hdd = daily_hdd_by_district[REFERENCE_DISTRICT]

    # Per-district annual mean air temperature and surface amplitude, used by the GSHP ground-temperature model. 
    district_climate_stats = {}
    for d, monthly in monthly_dd_by_district.items():
        monthly_T = np.array([HDD_BASE - monthly[m]["hdd_per_day"] for m in MONTHS_ORDER])
        n_days    = np.array([monthly[m]["n_days"]                 for m in MONTHS_ORDER])
        district_climate_stats[d] = {
            "annual_mean_T_C":     float((monthly_T * n_days).sum() / n_days.sum()),
            "surface_amplitude_C": float((monthly_T.max() - monthly_T.min()) / 2.0),
        }

def main():
    here = os.path.dirname(os.path.abspath(__file__))
    print("Running projections.py (climate + electricity) …")
    subprocess.run([sys.executable, os.path.join(here, "projections.py")],
                   check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    initialize(verbose=False)

    selected_district = _prompt_choice("Select district:", list(degree_days_by_district.keys()))
    selected_activity = _prompt_choice("Select activity class:", list(base_load_fracs.keys()))
    activities = [selected_activity]

    dd        = degree_days_by_district[selected_district]
    daily_hdd = daily_hdd_by_district[selected_district]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print(f"\nBuilding benchmark profiles [{selected_district}] …")
    build_benchmark_sheet(wb.create_sheet("Energy Profiles"), dd, daily_hdd,
                          activities=activities, district=selected_district)

    print(f"Building demand outputs [{selected_district}] …")
    build_demand_sheet(wb.create_sheet("Demand Outputs"), dd, daily_hdd,
                       activities=activities, district=selected_district)

    safe_activity = selected_activity.replace(":", " -").replace("/", "")
    timestamp     = datetime.now().strftime("%Y%m%d, %H%M")
    run_dir       = os.path.join(OUTPUTS_DIR, f"{selected_district} - {safe_activity} ({timestamp})")
    os.makedirs(run_dir, exist_ok=True)

    out_path = os.path.join(run_dir, "Demand Profiles.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")

    gen_plots = input("\nGenerate demand plots? (Y/N): ").strip().upper()
    if gen_plots == "Y":
        print("Generating demand plots …")
        generate_demand_plots(dd, daily_hdd, selected_district,
                              monthly_dd=monthly_dd_by_district[selected_district],
                              activities=activities, out_dir=run_dir,
                              district=selected_district)
    else:
        print("Skipping plot generation.")

# Imported at the bottom to avoid circular import (demand_report imports this module for the physics).
from demand_report import (
    build_benchmark_sheet, build_demand_sheet, generate_demand_plots
    )


if __name__ == "__main__":
    main()
