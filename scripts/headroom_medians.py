"""
ROLE: One-off / re-runnable provenance script. Not imported by the model.

headroom_medians.py — recomputes the 18 GRID_{IMPORT,EXPORT}_LIMIT_KW rows on the Scalars sheet of
data/model_parameters.xlsx directly from the raw DNO open-data extracts in data/headroom/.

WHY THIS EXISTS
The per-district ceilings were originally hand-computed as the MEDIAN headroom across each licence
area's primary substations. That estimator breaks down where a large share of primaries are recorded
at exactly zero headroom (fully saturated): the median falls into the zero mass and reports a value
that describes the saturated tail rather than a typical connectable substation. Two districts were
badly affected:

    Scotland N (SSEN/SHEPD)   45% of primaries at zero demand headroom -> median 110 kW
    England E and NE (NPG)    61% of primaries at zero generation headroom -> median 0 kW

110 kW is not a credible connection ceiling for a district containing hospitals whose non-heat
electrical peak alone is ~208 kW, and a literal zero export ceiling forced the optimiser to treat
batteries as PV sinks (no curtailment variable exists in the electricity balance). Both were
artifacts of the estimator, not findings about the networks.

WHAT CHANGED
The estimator is now the MEDIAN OF NON-ZERO headroom — the typical headroom among primaries that
have any at all. Applied uniformly to all nine districts, so the districts stay comparable with each
other (a Scotland-N-only rule would have traded an artifact for an inconsistency). For the six
districts with little zero mass the change is small-to-negligible; it is large only where the old
median was sitting in the zero cluster, which is precisely the case it was meant to fix.

Interpretation: the ceiling represents a typical *connectable* primary in the licence area. Sites on
an already-saturated primary are out of scope of a district-level average either way — the model has
no substation-level siting dimension to represent them.

VALIDATION
--verify re-runs the ORIGINAL plain-median estimator and checks it still reproduces the legacy
values below. This is what pins the filter/column choices to the numbers actually in the workbook:
all 16 distinct legacy values reproduce exactly. Run it after any refresh of data/headroom/ — a
failure means the upstream extract changed shape, not that the new estimator is wrong.

USAGE
    python scripts/headroom_medians.py            # print the comparison table
    python scripts/headroom_medians.py --verify   # + check legacy medians still reproduce
    python scripts/headroom_medians.py --apply    # write new values + sources into the workbook

--apply rewrites only the Value and Source cells of existing rows (no inserts), so the Scalars
sheet's whole-row conditional formatting is left intact. It backs the workbook up first and will
refuse to run while the file is open in Excel.
"""
import argparse
import os
import shutil
import sys
from datetime import datetime

import numpy as np
import pandas as pd

ROOT      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HEADROOM  = os.path.join(ROOT, "data", "headroom")
XLSX      = os.path.join(ROOT, "data", "model_parameters.xlsx")
BACKUPS   = os.path.join(ROOT, "_backups")

UKPN   = "East Anglia, England SE, Central S (UKPN) - dfes-network-headroom-report.csv"
NPG    = "England E and NE (NPG) - heatmapsubstationareas.csv"
ENWL_H = "England NW and N Wales (ENWL) - ndp-pry-bsp-headroom.csv"
ENWL_G = "England NW and N Wales (ENWL) - ndp-pry-bsp-generation.csv"
SPM    = "England NW and N Wales (SPM) - spm-nshr-data-workbook.csv"
NGED   = "England SW and S Wales, Midlands (NGED) - substations.csv"
SPD    = "Scotland E and W (SPD) - spd-nshr-data-workbook.csv"
SSEN   = "Scotland N (SSEN) - headroom-dashboard-data-march-2026.csv"

_cache = {}


def _load(fname):
    if fname not in _cache:
        _cache[fname] = pd.read_csv(os.path.join(HEADROOM, fname), low_memory=False)
    return _cache[fname]


def _num(series):
    return pd.to_numeric(series, errors="coerce").dropna()


# Per-DNO extractors. Each returns a Series of headroom values in kW, one per primary substation.
# Filters reproduce the original hand-computed selection exactly (see --verify).

def _ukpn(areas, category):
    # DFES NSHR. Demand rows carry a Physical Level tag, generation rows do not, so the primary
    # selection is done on voltage alone (11/33 kV) for both, which is what the original used.
    df = _load(UKPN)
    m = (df["LicenceArea"].isin(areas) & (df["Category"] == category)
         & (df["Scenario"] == "Counterfactual") & (df["Year"] == 2025)
         & (df["Voltage (kV)"].isin([11.0, 33.0])))
    return _num(df.loc[m, "Headroom (MW)"]) * 1000.0


def _npg(col):
    df = _load(NPG)
    m = (df["Type"] == "Primary") & (df["Downstream Voltage"].astype(str).isin(["11.0", "33.0"]))
    return _num(df.loc[m, col]) * 1000.0


def _enwl(fname, col):
    df = _load(fname)
    m = (df["GROUP"] == "PRIMARY") & (df["YEAR"] == 2025) & (df["SCENARIO"] == "4 - HOLISTIC TRANSITION")
    return _num(df.loc[m, col]) * 1000.0


def _spm_demand():
    df = _load(SPM)
    m = ((df["Headroom Type"] == "Demand") & (df["Scenario"] == "HT") & (df["DFES Year"] == 2025)
         & (df["Voltage"].astype(str).isin(["11.0", "33.0"])))
    return _num(df.loc[m, "Headroom (MW)"]) * 1000.0


def _nged(areas, col):
    df = _load(NGED)
    m = (df["type"] == "Primary") & (df["area"].isin(areas))
    return _num(df.loc[m, col]) * 1000.0


def _spd(htype):
    df = _load(SPD)
    m = (df["Headroom Type"] == htype) & (df["Scenario"] == "HT") & (df["Year"] == 2025)
    return _num(df.loc[m, "Headroom (MW)"]) * 1000.0


def _ssen(col):
    df = _load(SSEN)
    s = df[(df["Map / License Area"] == "Scotland / SHEPD") & (df["Substation Type"] == "Primary")]
    return _num(s[col]) * 1000.0


_EPN = ["Eastern Power Networks (EPN)"]
_SEC = ["London Power Networks (LPN)", "South Eastern Power Networks (SPN)"]

# (district, direction) -> (extractor, legacy plain-median value, short source label)
SPECS = {
    ("East Anglia", "import"):
        (lambda: _ukpn(_EPN, "Demand Headroom"), 7000, "UKPN DFES NSHR (EPN)"),
    ("East Anglia", "export"):
        (lambda: _ukpn(_EPN, "Gen inverter headroom"), 17710, "UKPN DFES NSHR (EPN)"),
    ("England E and NE", "import"):
        (lambda: _npg("Demand Headroom"), 5405, "NPG Heat Map (NE+Yorkshire)"),
    ("England E and NE", "export"):
        (lambda: _npg("Generation Headroom"), 0, "NPG Heat Map (NE+Yorkshire)"),
    ("England NW and N Wales", "import"):
        (lambda: pd.concat([_enwl(ENWL_H, "FIRM HEADROOM ( MVA )"), _spm_demand()]), 5920,
         "ENWL NDP + SPEN SPM NSHR pooled"),
    ("England NW and N Wales", "export"):
        (lambda: _enwl(ENWL_G, "PRY GEN HEADROOM - INVERTER BASED ( MW )"), 25700, "ENWL NDP only"),
    ("England SE and Central S", "import"):
        (lambda: _ukpn(_SEC, "Demand Headroom"), 10330, "UKPN DFES NSHR (LPN+SPN)"),
    ("England SE and Central S", "export"):
        (lambda: _ukpn(_SEC, "Gen inverter headroom"), 30940, "UKPN DFES NSHR (LPN+SPN)"),
    ("England SW and S Wales", "import"):
        (lambda: _nged(["South West", "South Wales"], "demandConnectedHeadroomMW"), 5150,
         "NGED Network Capacity Map"),
    ("England SW and S Wales", "export"):
        (lambda: _nged(["South West", "South Wales"], "generationConnectedHeadroomMW"), 6550,
         "NGED Network Capacity Map"),
    ("Midlands", "import"):
        (lambda: _nged(["West Midlands", "East Midlands"], "demandConnectedHeadroomMW"), 7550,
         "NGED Network Capacity Map"),
    ("Midlands", "export"):
        (lambda: _nged(["West Midlands", "East Midlands"], "generationConnectedHeadroomMW"), 12800,
         "NGED Network Capacity Map"),
    ("Scotland E", "import"):
        (lambda: _spd("Demand"), 11362, "SPD NSHR workbook"),
    ("Scotland E", "export"):
        (lambda: _spd("Generation (Fully Rated Converter)"), 9722, "SPD NSHR workbook"),
    ("Scotland N", "import"):
        (lambda: _ssen("Estimated Demand Headroom (MVA)"), 110, "SSEN SHEPD dashboard"),
    ("Scotland N", "export"):
        (lambda: _ssen("Estimated Generation Headroom (MW)"), 0, "SSEN SHEPD dashboard"),
}

# Scotland W shares SPD's single licence area with Scotland E — mirrored, not independently computed.
MIRRORED = {"Scotland W": "Scotland E"}


def compute():
    rows = []
    for (dist, direction), (fn, legacy, src) in SPECS.items():
        v = fn()
        nz = v[v > 0]
        rows.append({
            "district": dist, "direction": direction,
            "n": len(v), "n_zero": int((v == 0).sum()),
            "zero_pct": round(float((v == 0).mean()) * 100.0, 1),
            "median_kw": round(float(np.median(v)), 1),
            "median_nonzero_kw": round(float(np.median(nz)), 1) if len(nz) else 0.0,
            "legacy_kw": legacy, "source": src,
        })
    df = pd.DataFrame(rows)
    for mirror, of in MIRRORED.items():
        add = df[df["district"] == of].copy()
        add["district"] = mirror
        add["source"] = add["source"] + " (mirrored from %s)" % of
        df = pd.concat([df, add], ignore_index=True)
    return df.sort_values(["district", "direction"]).reset_index(drop=True)


def verify(df):
    bad = df[(df["median_kw"] - df["legacy_kw"]).abs()
             > np.maximum(2.0, 0.005 * df["legacy_kw"].abs())]
    if len(bad):
        print("\nFAIL: plain median no longer reproduces the legacy stored values:")
        print(bad[["district", "direction", "median_kw", "legacy_kw"]].to_string(index=False))
        return False
    print("\nOK: plain median reproduces all %d legacy values (filters/columns validated)." % len(df))
    return True


def apply_to_workbook(df):
    import openpyxl
    if not os.path.exists(XLSX):
        sys.exit("missing workbook: %s" % XLSX)
    if any(f.startswith("~$") and f.endswith("model_parameters.xlsx")
           for f in os.listdir(os.path.dirname(XLSX))):
        sys.exit("model_parameters.xlsx appears to be open in Excel — close it and re-run.")

    os.makedirs(BACKUPS, exist_ok=True)
    stamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    bdir   = os.path.join(BACKUPS, "%s_pre_headroom_nonzero_median" % stamp)
    os.makedirs(bdir, exist_ok=True)
    shutil.copy2(XLSX, os.path.join(bdir, "model_parameters.xlsx"))
    print("backed up -> %s" % bdir)

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Scalars"]
    idx = {(r["district"], r["direction"]): r for _, r in df.iterrows()}
    n = 0
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name or "GRID_" not in str(name):
            continue
        direction = "import" if "IMPORT" in name else "export"
        dist = str(name).split("[", 1)[1].rstrip("]")
        rec = idx.get((dist, direction))
        if rec is None:
            print("  ! no computed value for %s" % name)
            continue
        ws.cell(row, 3).value = float(rec["median_nonzero_kw"])
        ws.cell(row, 5).value = (
            "%s: MEDIAN OF NON-ZERO headroom across primary substations (n=%d, of which %d at zero "
            "= %.0f%%), present-day/2025 basis. Estimator changed from plain median: where a large "
            "share of primaries are fully saturated the plain median falls into the zero mass and "
            "describes the saturated tail rather than a typical connectable substation (plain median "
            "here = %.0f kW). Applied uniformly to all 9 districts for comparability. Reproducible "
            "via scripts/headroom_medians.py."
            % (rec["source"], rec["n"], rec["n_zero"], rec["zero_pct"], rec["median_kw"])
        )
        n += 1
    wb.save(XLSX)
    print("updated %d Scalars rows in %s" % (n, XLSX))


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[3])
    ap.add_argument("--verify", action="store_true", help="check the plain median still reproduces legacy values")
    ap.add_argument("--apply", action="store_true", help="write the new values into model_parameters.xlsx")
    args = ap.parse_args()

    df = compute()
    show = df[["district", "direction", "n", "zero_pct", "legacy_kw", "median_nonzero_kw"]].copy()
    show["change_x"] = (show["median_nonzero_kw"] / show["legacy_kw"].replace(0, np.nan)).round(2)
    print(show.to_string(index=False))

    ok = verify(df) if (args.verify or args.apply) else True
    if args.apply:
        if not ok:
            sys.exit("refusing to --apply while validation fails")
        apply_to_workbook(df)


if __name__ == "__main__":
    main()
